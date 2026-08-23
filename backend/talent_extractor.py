"""
Talent & Extras extraction for TPC Production Binder Wizard.

Converts Extreme Reach talent invoice PDFs + PTIP Excel into workbook rows
for the Talent & Extras tab.  Three intake scenarios:
  A  PDFs only (no PTIP)
  B  PTIP only (no PDFs)
  C  Both PDFs and PTIP (primary path)

The PTIP is authoritative for financial data; PDFs win on wages/misc placement
and supply work dates.  The backend returns fully-assembled rows in workbook
column order.  It never touches the .xlsx.
"""

import io
import re
import json
import base64
from collections import defaultdict
from datetime import datetime

import openpyxl
import pdfplumber

from payroll_reconciler import _AICP_CATEGORIES_TEXT


# ── Small helpers ────────────────────────────────────────────────────────────

def _to_float(v) -> float:
    if v is None:
        return 0.0
    try:
        return round(float(str(v).replace(',', '').replace('$', '').strip()), 2)
    except Exception:
        return 0.0


def _fmt_date(v) -> str:
    """Normalize any date value to MM/DD/YYYY string."""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%m/%d/%Y')
    s = str(v).strip()
    if not s or s.lower() == 'none':
        return ''
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
    # ISO format: YYYY-MM-DD or YYYY-MM-DD HH:MM:SS
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f"{m.group(2)}/{m.group(3)}/{m.group(1)}"
    return s


def _expand_date_yy(d: str) -> str:
    """Convert MM/DD/YY → MM/DD/YYYY (assumes 2000s)."""
    parts = d.split('/')
    if len(parts) == 3 and len(parts[2]) == 2:
        parts[2] = '20' + parts[2]
    return '/'.join(parts)


# ── Name utilities ───────────────────────────────────────────────────────────

_COMPANY_TOKENS = {'inc', 'llc', 'corp', 'ltd', 'agency', 'talent', 'agents',
                   'artists', 'models', 'group', 'management', 'entertainment'}


def _is_company_name(name: str) -> bool:
    """Heuristic: name is a company, not a person."""
    if not name:
        return False
    tokens = set(name.lower().replace(',', '').replace('.', '').split())
    return bool(tokens & _COMPANY_TOKENS) or '&' in name or name.endswith(('Inc', 'LLC', 'Corp', 'Ltd'))


def _ptip_name_to_last_first(name: str, is_agent: bool = False) -> str:
    """Convert 'First [Middle] Last' PTIP name → 'Last, First' workbook format.

    Company/agency names are returned unchanged.
    Single-letter middle initials are dropped; full middle names are kept.
    """
    if not name:
        return ''
    name = name.strip()
    if is_agent or _is_company_name(name) or ',' in name:
        return name
    parts = name.split()
    if len(parts) <= 1:
        return name
    last = parts[-1]
    first = parts[0]
    middle = [p for p in parts[1:-1] if len(p.rstrip('.')) > 1]
    if middle:
        return f"{last}, {first} {' '.join(middle)}"
    return f"{last}, {first}"


def _normalize_for_match(name: str) -> tuple[str, str]:
    """Return (full_norm, first_last_norm) for PDF↔PTIP name matching."""
    clean = re.sub(r"[.,']", '', name.lower()).strip()
    parts = clean.split()
    full = ' '.join(parts)
    first_last = f"{parts[0]} {parts[-1]}" if len(parts) >= 2 else full
    return full, first_last


def _ssn_last4(ssn: str) -> str:
    """Extract last 4 digits from a masked or full SSN string."""
    digits = re.sub(r'\D', '', str(ssn or ''))
    return digits[-4:] if len(digits) >= 4 else ''


# ── Address parsing ───────────────────────────────────────────────────────────

def _parse_ptip_address(addr_str) -> tuple[str, str, str, str]:
    """Parse multi-line PTIP address → (street, city, state, zip)."""
    if not addr_str:
        return '', '', '', ''
    lines = [ln.strip() for ln in str(addr_str).split('\n') if ln.strip()]
    if not lines:
        return '', '', '', ''

    city = state = zip_code = ''
    street_lines = lines

    m = re.match(r'^(.+?),\s+([A-Z]{2})\s+(\d{5})', lines[-1])
    if m:
        city, state, zip_code = m.group(1).strip(), m.group(2), m.group(3)
        street_lines = lines[:-1]

    street = street_lines[0] if street_lines else ''
    return street, city, state, zip_code


# Full street-type words → standard abbreviations (all-caps key → mixed-case value)
_STREET_TYPE_ABBREVS = {
    'AVENUE': 'Ave', 'STREET': 'St', 'DRIVE': 'Dr', 'ROAD': 'Rd',
    'BOULEVARD': 'Blvd', 'PLACE': 'Pl', 'COURT': 'Ct', 'LANE': 'Ln',
    'CIRCLE': 'Cir', 'HIGHWAY': 'Hwy', 'PARKWAY': 'Pkwy',
    'TERRACE': 'Ter', 'TRAIL': 'Trl',
}
# Two-letter compass directions that stay fully uppercase
_COMPOUND_DIRS = frozenset({'NE', 'NW', 'SE', 'SW'})
# Full-word directionals → single-letter abbreviation
_SINGLE_DIRS = {'NORTH': 'N', 'SOUTH': 'S', 'EAST': 'E', 'WEST': 'W'}


def _fmt_street_address(addr: str) -> str:
    if not addr:
        return ''
    # Strip unit numbers — comma-tolerant, no end-anchor so it catches mid-string units too
    addr = re.sub(r',?\s*#\s*\S+', '', addr.strip())
    addr = re.sub(r',?\s*(APT|APARTMENT|UNIT|SUITE|STE|FL|FLOOR|ROOM)\s+\S+',
                  '', addr.strip(), flags=re.IGNORECASE)
    result = []
    for word in addr.split():
        up = word.upper().rstrip('.')
        if up in _COMPOUND_DIRS:
            result.append(up)
        elif up in _SINGLE_DIRS:
            result.append(_SINGLE_DIRS[up])
        elif up in _STREET_TYPE_ABBREVS:
            result.append(_STREET_TYPE_ABBREVS[up])
        else:
            result.append(word.capitalize().rstrip('.'))
    return ' '.join(result).strip()


def _fmt_city(city: str) -> str:
    """Title-case a city name."""
    return city.title() if city else ''


# ── PTIP Excel parsing ────────────────────────────────────────────────────────

def _find_ptip_sheet(wb):
    """Find the worksheet that contains the PTIP talent data.

    Supports both the ER native format (Customer Name as first column, no '#')
    and the manually-enhanced format ('#' prepended as first column).
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_vals in ws.iter_rows(max_row=10, values_only=True):
            row_strs = [str(v).strip() if v is not None else '' for v in row_vals]
            if 'Invoice Number' in row_strs and 'Wages' in row_strs and 'Talent Name' in row_strs:
                return ws, row_strs
    return wb.active, []


def parse_ptip_xlsx(xlsx_bytes: bytes) -> tuple[list[dict], list[str], list[str]]:
    """
    Parse a PTIP Excel file.

    Returns (rows, issues, header_cols).
    Each row dict has keys matching the header column names.
    header_cols is the original ordered list of column names (for re-export).
    """
    rows: list[dict] = []
    issues: list[str] = []
    header_cols: list[str] = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws, header_cols = _find_ptip_sheet(wb)

        if not header_cols:
            issues.append("PTIP: could not locate header row with Invoice Number, Wages, Talent Name")
            return rows, issues, header_cols

        # Find the row number of the header so we can start data below it
        header_found_at = None
        for i, row_vals in enumerate(ws.iter_rows(values_only=True), start=1):
            row_strs = [str(v).strip() if v is not None else '' for v in row_vals]
            if 'Invoice Number' in row_strs and 'Wages' in row_strs and 'Talent Name' in row_strs:
                header_found_at = i
                break

        if header_found_at is None:
            issues.append("PTIP: header row not found")
            return rows, issues, header_cols

        inv_col_idx = next(
            (i for i, c in enumerate(header_cols) if 'Invoice' in c and 'Number' in c),
            -1,
        )

        for row_vals in ws.iter_rows(min_row=header_found_at + 1, values_only=True):
            if all(v is None for v in row_vals[:5]):
                break

            row = dict(zip(header_cols, row_vals))
            # Skip rows with no invoice number (totals / footer lines)
            if inv_col_idx >= 0:
                inv_val = str(row_vals[inv_col_idx] or '').strip()
                if not inv_val:
                    continue
            rows.append(row)

    except Exception as e:
        issues.append(f"PTIP parse error: {e}")

    return rows, issues, header_cols


# ── ER Invoice PDF parsing ───────────────────────────────────────────────────

_AMOUNT_RE = re.compile(r'^\d[\d,]*\.\d{2}$')
_DATE_YY_RE = re.compile(r'^\d{2}/\d{2}/\d{2}$')
# Cam Code / Cat Code tokens (e.g. "22NS") -- digits optionally followed by a
# few letters. Pure-digit cam codes match too, so this subsumes the old
# digit-only check.
_CODE_TOKEN_RE = re.compile(r'^\d+[A-Za-z]{0,4}$')


def _parse_talent_row(line: str) -> dict | None:
    """
    Parse one row from an Extreme Reach talent invoice table.

    Handles both union (5 amounts, 2 dates, cam code) and non-union (3 amounts,
    1 date, no cam code) formats.  Agent rows have 'Off' before their role token.

    Returns dict with keys: row_no, name, is_agent, work_date, gross_wages,
    misc_pmt, pah.  Returns None if the line is not a talent data row.
    """
    parts = line.split()
    if len(parts) < 7:
        return None
    if not (parts[0].isdigit() and parts[1].isdigit()):
        return None
    if parts[-1] != 'USD':
        return None

    row_no = int(parts[0])
    tail = parts[2:-1]  # strip row#, TID, USD

    # Pop trailing amount tokens
    amounts: list[float] = []
    while tail and _AMOUNT_RE.match(tail[-1]):
        amounts.insert(0, float(tail.pop().replace(',', '')))

    if len(amounts) not in (3, 5):
        return None

    # A one-off "Tag" token (e.g. an on-air spot ID like "SO'N37478ST") can
    # sit wedged between the two dates and the amounts on some rows --
    # discard it so the date pop below still lands on the real dates.
    if (tail and not _DATE_YY_RE.match(tail[-1]) and len(tail) >= 3
            and _DATE_YY_RE.match(tail[-2]) and _DATE_YY_RE.match(tail[-3])):
        tail.pop()

    # Pop trailing date tokens (MM/DD/YY)
    dates: list[str] = []
    while tail and _DATE_YY_RE.match(tail[-1]):
        dates.insert(0, tail.pop())

    if not dates:
        return None

    work_date = _expand_date_yy(dates[0])

    # Union: [Yr, GrossWages, MiscPmt, ApplyAmt, P&HSub]
    # Non-union: [Yr, GrossWages, MiscPmt]
    if len(amounts) == 5:
        gross_wages, misc_pmt, pah = amounts[1], amounts[2], amounts[4]
    else:
        gross_wages, misc_pmt, pah = amounts[1], amounts[2], 0.0

    # Remaining tail: [name...] [cam_code?] [On|Off] [role_or_agent_code] [OS1] [tags...]
    on_off_idx = next((i for i, t in enumerate(tail) if t.lower() in ('on', 'off')), None)
    if on_off_idx is None:
        return None

    on_off = tail[on_off_idx].lower()
    role_code = tail[on_off_idx + 1] if on_off_idx + 1 < len(tail) else ''
    is_agent = (on_off == 'off') or (role_code.lower() == 'agent')

    # Name is everything before On/Off, minus trailing Cam/Cat code token(s)
    # (e.g. "Dayanari Umana 22NS" -> "Dayanari Umana"; a loan-out row has no
    # name at all here, just the code(s), which strips down to empty).
    before = tail[:on_off_idx]
    while before and _CODE_TOKEN_RE.match(before[-1]):
        before = before[:-1]
    name = ' '.join(before)

    return {
        'row_no':     row_no,
        'name':       name,
        'is_agent':   is_agent,
        'work_date':  work_date,
        'gross_wages': round(gross_wages, 2),
        'misc_pmt':   round(misc_pmt, 2),
        'pah':        round(pah, 2),
    }


def parse_er_invoice_pdf(pdf_bytes: bytes) -> dict | None:
    """
    Parse an Extreme Reach talent invoice PDF.

    Returns dict with: invoice_no, invoice_date, cycle_dates, union_type,
    total_wages, total_misc, total_pah, total_er_tax, total_handling,
    talent_rows (list of _parse_talent_row dicts).
    Returns None if this does not look like an ER talent invoice.
    """
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = '\n'.join(pg.extract_text() or '' for pg in pdf.pages)
    except Exception:
        return None

    if 'Extreme Reach Talent, Inc' not in text:
        return None

    inv_no = ''
    m = re.search(r'Invoice\s*#\s+(\d+)', text)
    if m:
        inv_no = m.group(1).strip()

    invoice_date = ''
    m = re.search(r'Invoice Date\s+(\d{1,2}/\d{1,2}/\d{4})', text)
    if m:
        invoice_date = _fmt_date(m.group(1))

    cycle_dates = ''
    m = re.search(r'Cycle Dates\s+([\d/]+ - [\d/]+)', text)
    if m:
        cycle_dates = m.group(1).strip()
    else:
        m = re.search(r'Cycle Dates\s+([\d/]+)\s*-\s*$', text, re.MULTILINE)
        if m:
            cycle_dates = m.group(1).strip()

    union_type = ''
    m = re.search(r'Union\s+(SAG|Non-Union|AFTRA)', text, re.IGNORECASE)
    if m:
        raw = m.group(1)
        union_type = 'Non-union' if 'non' in raw.lower() else raw.upper()

    pay_type = ''
    m = re.search(r'Pay Type\s+(\S+)\s+Handling', text)
    if m:
        pay_type = m.group(1).strip()

    commercial_title = ''
    m = re.search(r'Primary\s+\S+\s+(.+)', text)
    if m:
        commercial_title = m.group(1).strip()

    total_wages    = _to_float(re.search(r'Wages\s+\$([\d,.]+)', text) and
                                re.search(r'Wages\s+\$([\d,.]+)', text).group(1))
    total_misc     = _to_float(re.search(r'Misc Payments\s+\$([\d,.]+)', text) and
                                re.search(r'Misc Payments\s+\$([\d,.]+)', text).group(1))
    total_pah      = _to_float(re.search(r'SAG Pension.*?\$([\d,.]+)', text) and
                                re.search(r'SAG Pension.*?\$([\d,.]+)', text).group(1))
    total_er_tax   = _to_float(re.search(r'Payroll Taxes\s+\$([\d,.]+)', text) and
                                re.search(r'Payroll Taxes\s+\$([\d,.]+)', text).group(1))
    total_handling = _to_float(re.search(r'Handling\s+\$([\d,.]+)', text) and
                                re.search(r'Handling\s+\$([\d,.]+)', text).group(1))

    talent_rows: list[dict] = []
    pdf_lines = text.split('\n')
    i = 0
    while i < len(pdf_lines):
        row = _parse_talent_row(pdf_lines[i].strip())
        if row:
            row['loan_out_company'] = ''
            if not row['name']:
                # Name wrapped onto the line above the data row. For a
                # loan-out entry ER prints the real performer's own name
                # above the row and their loan-out company below it
                # (confirmed real pattern: "Joshua E Vasquez" / data row /
                # "MaJeliv Productions Inc").
                prev_line = pdf_lines[i - 1].strip() if i > 0 else ''
                if prev_line and not _parse_talent_row(prev_line):
                    row['name'] = prev_line

                next_line = pdf_lines[i + 1].strip() if i + 1 < len(pdf_lines) else ''
                has_next = bool(next_line) and not _parse_talent_row(next_line)
                if row['name'] and has_next and _is_company_name(next_line):
                    row['loan_out_company'] = next_line
                    i += 1
                elif not row['name'] and has_next:
                    # No usable name line above -- fall back to the
                    # trailing line as this row's identity (old behavior).
                    row['name'] = next_line
                    if _is_company_name(next_line):
                        row['is_agent'] = True
                    i += 1
            talent_rows.append(row)
        i += 1

    return {
        'invoice_no':        inv_no,
        'invoice_date':      invoice_date,
        'cycle_dates':       cycle_dates,
        'union_type':        union_type,
        'pay_type':          pay_type,
        'commercial_title':  commercial_title,
        'total_wages':       total_wages,
        'total_misc':        total_misc,
        'total_pah':         total_pah,
        'total_er_tax':      total_er_tax,
        'total_handling':    total_handling,
        'talent_rows':       talent_rows,
    }


# ── Duplicate detection ───────────────────────────────────────────────────────

def _detect_duplicates(ptip_rows: list[dict], pdf_name_counts: dict) -> set[int]:
    """
    Return the set of ptip_rows indices that are duplicates.

    Duplicate rule: same (invoice_no, talent_name, wages, er_tax) in the PTIP,
    where the PDF shows fewer occurrences of that name on that invoice than the
    PTIP does.  Agent rows (cast_category == 'Agent') follow the same rule.

    pdf_name_counts: {(invoice_no, full_norm_name): count_in_pdf}
    """
    duplicates: set[int] = set()
    # Group PTIP rows by (invoice_no, norm_name, wages, er_tax)
    ptip_groups: dict[tuple, list[int]] = defaultdict(list)

    for idx, row in enumerate(ptip_rows):
        inv_no = str(row.get('Invoice Number', '')).strip()
        name_raw = str(row.get('Talent Name', '')).strip()
        wages = _to_float(row.get('Wages'))
        er_tax = _to_float(row.get('Employer Taxes'))
        full_norm, _ = _normalize_for_match(name_raw)
        key = (inv_no, full_norm, wages, er_tax)
        ptip_groups[key].append(idx)

    for (inv_no, full_norm, wages, er_tax), indices in ptip_groups.items():
        if len(indices) <= 1:
            continue
        pdf_count = pdf_name_counts.get((inv_no, full_norm), len(indices))
        for extra_idx in indices[pdf_count:]:
            duplicates.add(extra_idx)

    return duplicates


# ── Workbook row builder ──────────────────────────────────────────────────────

def _build_row(
    *,
    item_no: int,
    ptip: dict | None,
    pdf_talent: dict | None,
    pdf_invoice: dict | None,
    received_invoice: bool,
    is_duplicate: bool,
    first_tax_row: bool,
    scenario: str,  # 'A', 'B', or 'C'
    ptip_row_no: int | None,
    payment_entity: str = 'Extreme Reach Talent, Inc',
    pah_from_pdf: bool = False,    # Teams: P&H from PDF talent table, not PTIP
    workbook_type: str = '',       # 'ga' changes ptip_amount's total definition -- see below
    first_row_of_invoice: bool = False,  # Teams: this row absorbs the invoice's Other Fees
    loan_out: bool = False,
    loan_out_company: str = '',
    talent_name_override: str = '',  # use the real performer's name instead of a matched
                                      # loan-out corp's PTIP name (caller-determined)
) -> dict:
    """Assemble one workbook row from PTIP and/or PDF data."""

    is_agent = False
    cast_cat = ''
    if ptip:
        cast_cat = str(ptip.get('Cast Category', '') or '').strip()
        # Strip multiline cell artifacts (e.g. "Dream Team Talent\n(Dream Team Talent Agency LLC)")
        name_for_agent_check = str(ptip.get('Talent Name', '') or '').split('\n')[0].strip()
        # Some PTIP files label agent rows as "Model" instead of "Agent"
        # so use company-name heuristic as a fallback
        is_agent = cast_cat.lower() == 'agent' or _is_company_name(name_for_agent_check)
    elif pdf_talent:
        is_agent = pdf_talent.get('is_agent', False)

    # ER's PTIP sometimes writes a loan-out performer's name as
    # "First Last\n(Company Name)" in the same Talent Name cell (confirmed
    # real example: "Aspen K Wilson\n(Aspen Kennedy Inc)") -- an agent row's
    # own name is never itself the loan-out pattern, so skip those.
    ptip_loan_out_company = ''
    if ptip and not is_agent:
        full_ptip_name = str(ptip.get('Talent Name', '') or '').strip()
        m = re.search(r'\(([^()]+)\)\s*$', full_ptip_name)
        if m and '\n' in full_ptip_name:
            ptip_loan_out_company = m.group(1).strip()

    # ── Name ────────────────────────────────────────────────────────────────
    if talent_name_override:
        talent_name = talent_name_override
    elif ptip:
        name_raw = str(ptip.get('Talent Name', '') or '').split('\n')[0].strip()
        talent_name = _ptip_name_to_last_first(name_raw, is_agent)
    elif pdf_talent:
        name_raw = pdf_talent.get('name', '')
        talent_name = _ptip_name_to_last_first(name_raw, is_agent)
    else:
        talent_name = ''

    # ── Title ────────────────────────────────────────────────────────────────
    if is_agent:
        title = 'Agency fee'
    elif cast_cat:
        title = cast_cat
    elif pdf_talent and pdf_talent.get('title'):
        title = pdf_talent['title']
    else:
        title = ''

    # ── Address ──────────────────────────────────────────────────────────────
    street = city = state = zip_code = ''
    if ptip:
        street, city, state, zip_code = _parse_ptip_address(ptip.get('Talent Address'))
        street = _fmt_street_address(street)
        city   = _fmt_city(city)

    # ── Work state & qualify ─────────────────────────────────────────────────
    if ptip:
        work_state = str(ptip.get('Work State', '') or '').strip().upper()
    elif pdf_invoice:
        work_state = 'IL'  # default; frontend overrides from project
    else:
        work_state = ''

    if state and work_state and state.upper() != work_state:
        qualify = 'NO-OOS'
    else:
        qualify = ''

    # ── PTIP financial columns ───────────────────────────────────────────────
    if ptip:
        # Try explicit Total column (manually-enhanced PTIP); otherwise compute it.
        # ER native PTIP excludes State Tax Withheld from IL's own total (it's an
        # employee-side deduction there). GA's Total Amount column is a real Excel
        # formula (=SUM(J:S)) that DOES include State Tax Withheld -- ptip_amount
        # has to match whatever "Total" means on the tab it's feeding, or
        # PTIP_Check would show a false discrepancy on every single row, off by
        # exactly the state tax amount, regardless of workbook_type.
        _RAW_TOTAL = _to_float(ptip.get('Total') or ptip.get('TOTAL'))
        if _RAW_TOTAL:
            ptip_amount = _RAW_TOTAL
        else:
            _TOTAL_KEYS = [
                'Wages', 'P&H', 'Non Resident Corp Tax Amount',
                'Local Tax Withheld', 'State Disability Withheld',
                'Employer Taxes', 'Workers Compensation', 'Handling Fee', 'Signatory Fee',
            ]
            if _is_ga_workbook(workbook_type):
                _TOTAL_KEYS = _TOTAL_KEYS + ['State Tax Withheld']
            ptip_amount = round(sum(_to_float(ptip.get(k)) for k in _TOTAL_KEYS), 2)
        er_tax_ptip    = _to_float(ptip.get('Employer Taxes'))
        wc_ptip        = _to_float(ptip.get('Workers Compensation'))
        handling_ptip  = _to_float(ptip.get('Handling Fee'))
        sag_ptip       = _to_float(ptip.get('P&H'))
        signatory_ptip = _to_float(ptip.get('Signatory Fee'))
        state_tax_withheld       = _to_float(ptip.get('State Tax Withheld'))
        local_tax_withheld       = _to_float(ptip.get('Local Tax Withheld'))
        state_disability_withheld = _to_float(ptip.get('State Disability Withheld'))
        check_number      = str(ptip.get('Check Number', '') or '').strip()
        commercial_id     = str(ptip.get('Commercial Id', '') or '').strip()
        commercial_title  = str(ptip.get('Commercial Title', '') or '').strip()
        ssn_fein          = (str(ptip.get('SSN', '') or '').strip() or
                             str(ptip.get('FEIN', '') or '').strip())
        on_ptip           = True
    else:
        ptip_amount = er_tax_ptip = wc_ptip = handling_ptip = sag_ptip = signatory_ptip = None
        state_tax_withheld = local_tax_withheld = state_disability_withheld = 0.0
        check_number = commercial_id = commercial_title = ssn_fein = ''
        on_ptip = False

    # PDF title includes spot length (e.g. "Carried Away :30"); prefer it over PTIP
    if pdf_invoice and pdf_invoice.get('commercial_title'):
        commercial_title = pdf_invoice['commercial_title']

    # ── Other Fees (invoice-level, no dedicated column) ──────────────────────
    # Business Affairs Fee / Wire Fee / etc. have no column of their own on the
    # template regardless of whether this invoice has a PTIP report -- a PTIP
    # match only ever carries per-performer wage/tax data, never these. Attach
    # to exactly one row per invoice (the first PDF-matched row) so the money
    # isn't silently lost when PTIP exists but doesn't reflect it.
    other_fees = 0.0
    notes      = ''
    if first_row_of_invoice and pdf_invoice:
        other_fees = pdf_invoice.get('total_other_fees', 0.0)
        detail     = pdf_invoice.get('other_fees_detail') or []
        if detail:
            notes = ', '.join(f"{d['label']} (${d['amount']:,.2f})" for d in detail)

    # ── Wages and Misc Pmt ───────────────────────────────────────────────────
    if scenario in ('A',) and pdf_talent:
        # PDF-only: wages and misc from PDF talent row
        wages    = pdf_talent.get('gross_wages', 0.0)
        misc_pmt = pdf_talent.get('misc_pmt', 0.0)
        # Taxes: first talent row only (agent rows always get 0 taxes in PDF-only)
        if first_tax_row and not is_agent and pdf_invoice:
            er_tax        = pdf_invoice.get('total_er_tax', 0.0)
            sag           = pdf_invoice.get('total_pah', 0.0)
            handling      = pdf_invoice.get('total_handling', 0.0)
            wc            = pdf_invoice.get('total_wc', 0.0)
            # No PTIP for this invoice means the PDF is the only source of
            # truth -- take its numbers as-is. No synthetic percentage guess;
            # only use a signatory fee when the PDF prints a real "TS
            # Signatory Fee" footer line.
            signatory_fee = pdf_invoice.get('total_signatory_fee', 0.0)
        else:
            er_tax = sag = handling = wc = signatory_fee = 0.0
    elif scenario == 'B' and ptip:
        # PTIP-only: talent rows go to wages; agent rows go to misc (we know the cast)
        ptip_wages = _to_float(ptip.get('Wages'))
        if is_agent:
            wages    = 0.0
            misc_pmt = ptip_wages
        else:
            wages    = ptip_wages
            misc_pmt = 0.0
        er_tax   = er_tax_ptip or 0.0
        wc       = wc_ptip or 0.0
        handling = handling_ptip or 0.0
        sag      = sag_ptip or 0.0
        # No synthetic percentage guess -- a lot of payroll has no signatory
        # fee at all. Only record it when it's actually on the PTIP report.
        signatory_fee = signatory_ptip or 0.0
    elif scenario == 'C':
        # Both: PDF wins on wages/misc placement; PTIP for taxes
        if pdf_talent and received_invoice:
            wages    = pdf_talent.get('gross_wages', 0.0)
            misc_pmt = pdf_talent.get('misc_pmt', 0.0)
        else:
            # No PDF for this row (PTIP-only invoice in a mixed batch)
            ptip_wages = _to_float(ptip.get('Wages')) if ptip else 0.0
            if is_agent:
                wages    = 0.0
                misc_pmt = ptip_wages
            else:
                wages    = ptip_wages
                misc_pmt = 0.0
        er_tax   = er_tax_ptip or 0.0
        wc       = wc_ptip or 0.0
        handling = handling_ptip or 0.0
        if pah_from_pdf and pdf_talent and pdf_talent.get('pah', 0.0) > 0:
            sag = pdf_talent['pah']
        else:
            sag = sag_ptip or 0.0
        # No synthetic percentage guess -- only record a signatory fee when
        # it's actually on the PTIP report.
        signatory_fee = signatory_ptip or 0.0
    else:
        wages = misc_pmt = er_tax = wc = handling = sag = signatory_fee = 0.0

    # ── Dates ────────────────────────────────────────────────────────────────
    if pdf_talent:
        work_dates   = pdf_talent.get('work_date', '')
        work_days    = 1
    else:
        work_dates = ''
        work_days  = None

    if pdf_invoice:
        invoice_date = pdf_invoice.get('invoice_date', '')
    elif ptip:
        invoice_date = _fmt_date(ptip.get('Check Date') or '')
    else:
        invoice_date = ''

    # ── Invoice info ─────────────────────────────────────────────────────────
    if ptip:
        invoice_no = str(ptip.get('Invoice Number', '') or '').strip()
    elif pdf_invoice:
        invoice_no = pdf_invoice.get('invoice_no', '')
    else:
        invoice_no = ''

    # ── Pay type (Session / Reuse / Non-Union / etc.) ────────────────────────
    if pdf_invoice:
        pay_type = pdf_invoice.get('pay_type', '')
    else:
        pay_type = ''

    # ── Total (static) ───────────────────────────────────────────────────────
    if ptip_amount is not None:
        total = ptip_amount
    else:
        total = round(wages + misc_pmt + er_tax + wc + handling + sag + signatory_fee + other_fees, 2)

    if not loan_out_company and ptip_loan_out_company:
        loan_out         = True
        loan_out_company = ptip_loan_out_company

    if loan_out_company:
        loan_out_note = f"Loan-out company: {loan_out_company}"
        notes = f"{notes}; {loan_out_note}" if notes else loan_out_note

    return {
        'item_no':        item_no,
        'qualify':        qualify,
        'on_ptip':        on_ptip,
        'ptip_amount':    ptip_amount,
        'work_state':     work_state,
        'talent_name':    talent_name,
        'loan_out':       'YES' if loan_out else 'NO',
        'loan_out_company': loan_out_company,
        'title':          title,
        'work_days':      work_days,
        'work_dates':     work_dates,
        'invoice_no':     invoice_no,
        'invoice_date':   invoice_date,
        'wages':          round(wages, 2),
        'misc_pymt':      round(misc_pmt, 2),
        'er_tax':         round(er_tax, 2),
        'wc':             round(wc, 2),
        'handling':       round(handling, 2),
        'sag':            round(sag, 2),
        'signatory_fee':  round(signatory_fee, 2),
        'other_fees':     round(other_fees, 2),
        'state_tax_withheld':        round(state_tax_withheld, 2),
        'local_tax_withheld':        round(local_tax_withheld, 2),
        'state_disability_withheld': round(state_disability_withheld, 2),
        'total':          total,
        'check_number':   check_number,
        'received_invoice': received_invoice,
        'payment_entity': payment_entity,
        'type':           pay_type,
        'home_address':   street,
        'city':           city,
        'state':          state,
        'zip':            zip_code,
        'ssn_fein':          ssn_fein,
        'commercial_id':     commercial_id,
        'commercial_title':  commercial_title,
        'notes':             notes,
        'ptip_row_no':    ptip_row_no,
        'is_duplicate':   is_duplicate,
    }


# ── Organized PTIP builder ───────────────────────────────────────────────────

def build_organized_ptip_xlsx(
    ptip_rows: list[dict],
    header_cols: list[str],
    received_invoice_nos: set[str],
    duplicate_indices: set[int],
    xlsx_bytes: bytes,
    pdf_order_within_invoice: dict | None = None,
) -> bytes:
    """
    Return a new .xlsx that mirrors the original ER PTIP layout:
      Row 1  — empty merged row (matches original height)
      Row 2  — title text copied verbatim from the original (same font/fill)
      Row 3  — column headers (same teal header style)
      Row 4+ — data rows (same body style, sorted by invoice number)

    '#' column values:
      sequential int  — invoice received (resets per invoice group)
      blank           — no matching PDF received for this invoice
      'duplicate'     — row is a duplicate, excluded from workbook
    """
    from copy import copy as _copy
    from openpyxl.utils import get_column_letter

    invoice_col = 'Invoice Number'
    if invoice_col not in header_cols:
        invoice_col = next((c for c in header_cols if 'Invoice' in c), header_cols[0])

    sorted_with_orig = sorted(enumerate(ptip_rows),
                              key=lambda t: str(t[1].get(invoice_col, '') or '').zfill(20))

    # Reorder within each invoice group to match PDF talent order (by name)
    if pdf_order_within_invoice:
        from itertools import groupby as _groupby
        reordered: list[tuple[int, dict]] = []
        for inv_key, group_iter in _groupby(
            sorted_with_orig,
            key=lambda t: str(t[1].get(invoice_col, '') or '').strip(),
        ):
            group = list(group_iter)
            pdf_names = pdf_order_within_invoice.get(inv_key)
            if pdf_names:
                ordered, remaining = [], list(group)
                for pdf_norm in pdf_names:
                    for i, (orig_idx, row) in enumerate(remaining):
                        row_norm, _ = _normalize_for_match(
                            str(row.get('Talent Name', '') or '')
                        )
                        # Exact match first; substring fallback for agency names
                        # that differ between PDF (full legal) and PTIP (short).
                        if (row_norm == pdf_norm or
                                (row_norm and pdf_norm and
                                 (row_norm in pdf_norm or pdf_norm in row_norm))):
                            ordered.append(remaining.pop(i))
                            break
                ordered.extend(remaining)
                reordered.extend(ordered)
            else:
                reordered.extend(group)
        sorted_with_orig = reordered

    has_hash_col = bool(header_cols) and header_cols[0] == '#'
    out_header = list(header_cols) if has_hash_col else ['#'] + list(header_cols)
    n_cols = len(out_header)

    # ── Read layout + formatting from original source file ───────────────────
    wb_src = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws_src = None
    for sn in wb_src.sheetnames:
        ws_cand = wb_src[sn]
        for rv in ws_cand.iter_rows(max_row=10, values_only=True):
            rstr = [str(v).strip() if v is not None else '' for v in rv]
            if 'Invoice Number' in rstr and 'Wages' in rstr and 'Talent Name' in rstr:
                ws_src = ws_cand
                break
        if ws_src:
            break
    if ws_src is None:
        ws_src = wb_src.active

    # Locate header row in source to derive title row and data row
    src_hdr_row = 3
    for i, rv in enumerate(ws_src.iter_rows(max_row=10, values_only=True), start=1):
        rstr = [str(v).strip() if v is not None else '' for v in rv]
        if 'Invoice Number' in rstr and 'Wages' in rstr and 'Talent Name' in rstr:
            src_hdr_row = i
            break
    src_title_row = max(1, src_hdr_row - 1)
    src_data_row  = src_hdr_row + 1

    title_src  = ws_src.cell(row=src_title_row, column=1)
    hdr_src    = ws_src.cell(row=src_hdr_row,   column=1)
    data_src   = ws_src.cell(row=src_data_row,  column=1)

    r1_height   = ws_src.row_dimensions[1].height          or 28.9
    r2_height   = ws_src.row_dimensions[src_title_row].height or 129.6
    r3_height   = ws_src.row_dimensions[src_hdr_row].height   or 42.75
    data_height = ws_src.row_dimensions[src_data_row].height  or 38.25

    # ── Build output workbook ────────────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = 'Organized PTIP'

    # Row 1: empty merged row
    ws.row_dimensions[1].height = r1_height
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Row 2: A2 holds the logo; B2:end holds the title text
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.styles import Color as _OXLColor
    from openpyxl.drawing.image import Image as _OXLImage
    import zipfile as _zipfile

    ws.row_dimensions[2].height = r2_height

    # A2: logo copied from original (not merged into the text block)
    try:
        with _zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as _zf:
            _media = [n for n in _zf.namelist() if n.startswith('xl/media/')]
            if _media:
                _logo_bytes = _zf.read(_media[0])
                _logo = _OXLImage(io.BytesIO(_logo_bytes))
                _logo.anchor = 'A2'
                ws.add_image(_logo)
    except Exception:
        pass  # no image in source — skip silently

    # B2:end: title text, first line 18pt, remaining lines 12pt
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=n_cols)
    t = ws.cell(row=2, column=2)

    title_text = title_src.value or ''
    src_fname = (title_src.font.name or 'Tahoma') if title_src.font else 'Tahoma'
    try:
        src_fcolor = title_src.font.color.rgb if title_src.font.color.type == 'rgb' else 'FF000000'
    except Exception:
        src_fcolor = 'FF000000'

    f_large = InlineFont(rFont=src_fname, sz=18, color=_OXLColor(rgb=src_fcolor))
    f_small = InlineFont(rFont=src_fname, sz=12, color=_OXLColor(rgb=src_fcolor))
    title_parts = title_text.split('\n', 1)
    if len(title_parts) > 1:
        t.value = CellRichText(
            TextBlock(f_large, title_parts[0]),
            TextBlock(f_small, '\n' + title_parts[1]),
        )
    else:
        t.value = CellRichText(TextBlock(f_large, title_text))

    t.fill      = _copy(title_src.fill)
    t.alignment = _copy(title_src.alignment)
    t.border    = _copy(title_src.border)

    # Row 3: column headers
    ws.row_dimensions[3].height = r3_height
    for col_i, col_name in enumerate(out_header, start=1):
        cell = ws.cell(row=3, column=col_i, value=col_name)
        cell.font      = _copy(hdr_src.font)
        cell.fill      = _copy(hdr_src.fill)
        cell.alignment = _copy(hdr_src.alignment)
        cell.border    = _copy(hdr_src.border)

    # Column widths: # gets 6, remaining mirror original source columns
    if not has_hash_col:
        ws.column_dimensions['A'].width = 6
        for i in range(1, len(header_cols) + 1):
            w = ws_src.column_dimensions[get_column_letter(i)].width
            ws.column_dimensions[get_column_letter(i + 1)].width = w or 10
    else:
        for i in range(1, n_cols + 1):
            w = ws_src.column_dimensions[get_column_letter(i)].width
            ws.column_dimensions[get_column_letter(i)].width = w or 10

    # Data rows (row 4+)
    last_invoice = None
    seq = 0
    for out_row_i, (orig_idx, row) in enumerate(sorted_with_orig, start=4):
        ws.row_dimensions[out_row_i].height = data_height

        if has_hash_col:
            row_vals = [row.get(col) for col in header_cols]
        else:
            row_vals = [None] + [row.get(col) for col in header_cols]

        inv_no = str(row.get(invoice_col, '') or '').strip()
        if inv_no != last_invoice:
            last_invoice = inv_no
            seq = 0

        if orig_idx in duplicate_indices:
            hash_val = 'duplicate'
        elif inv_no not in received_invoice_nos:
            hash_val = ''
        else:
            seq += 1
            hash_val = seq

        row_vals[0] = hash_val

        for col_i, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=out_row_i, column=col_i, value=val)
            cell.font      = _copy(data_src.font)
            cell.fill      = _copy(data_src.fill)
            cell.alignment = _copy(data_src.alignment)
            cell.border    = _copy(data_src.border)

    buf = io.BytesIO()
    wb_out.save(buf)
    return buf.getvalue()


# ── Main extraction entry point ───────────────────────────────────────────────

def _llm_fuzzy_match_talent(
    invoice_names: list[str],
    ptip_names: list[str],
    openai_key: str,
) -> tuple[dict[str, str], dict[str, str]]:
    """LLM fallback: match invoice names to PTIP names that exact normalization missed.

    Handles name-order differences (First Last vs Last, First), suffixes (II/Jr/Sr),
    company DBA variations, and typos.

    Returns (matches, unmatched_reasons) where:
      matches:           {invoice_name: ptip_name} for confident matches
      unmatched_reasons: {invoice_name: reason} for names that couldn't be matched
    """
    if not invoice_names or not ptip_names or not openai_key:
        return {}, {}
    try:
        from openai import OpenAI
        client = OpenAI(api_key=openai_key)
        prompt = (
            "You are reconciling a payroll invoice against a PTIP (Production Tax Incentive) report.\n\n"
            "Invoice names (not yet matched):\n"
            f"{json.dumps(invoice_names, indent=2)}\n\n"
            "PTIP names (not yet matched):\n"
            f"{json.dumps(ptip_names, indent=2)}\n\n"
            "Identify which invoice name and PTIP name refer to the SAME person or company despite "
            "formatting differences. Common differences:\n"
            "- Name order: 'William Rose II' (invoice) = 'Rose, William' (PTIP)\n"
            "- Company DBA: 'Ashley Washington Holdings LLC dba Karen Stavins Enterprises' = "
            "'Karen Stavins Enterprises'\n"
            "- Typos: 'Bac Talent Managment Inc' = 'Bac Talent Management Inc'\n"
            "- A cast/category code glued directly onto the name with no separator, from imperfect "
            "PDF table extraction: 'Cedric Williams 22NS' (invoice) = 'Cedric Williams' (PTIP), "
            "'Rollie Smith 22NS' = 'Rollie Smith'\n"
            "- Loan-out payment: the invoice sometimes labels a performer's row with just their "
            "loan-out corporation's name (from a PDF line the performer's own name wrapped away "
            "from), while the PTIP entry lists the performer with their loan-out company in "
            "parentheses on a second line: 'Aspen Kennedy Inc' (invoice) = 'Aspen K Wilson\\n"
            "(Aspen Kennedy Inc)' (PTIP) -- match on the parenthetical company name appearing "
            "verbatim in the PTIP entry, even though the invoice name shares no words with the "
            "PTIP entry's own first line.\n\n"
            "Rules:\n"
            "- Only match if you are CONFIDENT they are the same person or entity.\n"
            "- Do NOT guess — if uncertain, put the name in 'unmatched' with a brief reason.\n"
            "- Each invoice name maps to at most one PTIP name and vice versa.\n"
            "- Every invoice name must appear in either 'matches' or 'unmatched'.\n\n"
            "Return ONLY a JSON object with exactly two keys:\n"
            '{"matches": {"invoice_name": "ptip_name", ...}, '
            '"unmatched": {"invoice_name": "brief reason why no confident match was found", ...}}'
        )
        resp = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=800,
            response_format={"type": "json_object"},
        )
        raw = json.loads(resp.choices[0].message.content)
        if not isinstance(raw, dict):
            return {}, {}
        inv_set  = set(invoice_names)
        ptip_set = set(ptip_names)
        matches = {
            str(k): str(v)
            for k, v in (raw.get("matches") or {}).items()
            if str(k) in inv_set and str(v) in ptip_set
        }
        reasons = {
            str(k): str(v)
            for k, v in (raw.get("unmatched") or {}).items()
            if str(k) in inv_set
        }
        return matches, reasons
    except Exception:
        return {}, {}


def _is_ga_workbook(workbook_type: str) -> bool:
    """workbook_type is an unvalidated free-text string from the frontend's
    own form field -- nothing in this backend defines or normalizes its
    exact casing/spelling, so match tolerantly rather than gating two real
    features (the ptip_amount total definition and AICP classification) on
    a single guessed literal like 'ga' that could silently never match."""
    return (workbook_type or "").strip().lower() in ("ga", "georgia")


# ── AICP classification (GA only -- IL's Talent & Extras has no such column) ──

def _classify_aicp_codes_talent(rows: list[dict], openai_key: str) -> None:
    """Assigns row['aicp_code'] (int 1-25, or None) to every row in place, via
    one batched GPT call. Same steering-not-restriction principle as Crew
    Payroll's _classify_aicp_codes: 23/24 (Georgia Cast/Extras Hires) will be
    correct for most rows here, but every category stays available since an
    odd case (e.g. an agent fee, or a per diem paid through Talent) is
    genuinely possible. Best-effort -- leaves aicp_code as None on any
    failure rather than raising, same as the Crew Payroll version."""
    for row in rows:
        row['aicp_code'] = None

    if not rows or not openai_key:
        return

    try:
        from openai import OpenAI
        client = OpenAI(api_key=openai_key)

        items = [
            {"index": i, "talent_name": r.get("talent_name", ""), "title": r.get("title", "")}
            for i, r in enumerate(rows)
        ]

        prompt = (
            "You are classifying rows on a Georgia film production's Talent Report "
            "by AICP category, for the state tax incentive submission.\n\n"
            "Pick the single best-matching AICP category number for each row below, "
            "using this list:\n"
            f"{_AICP_CATEGORIES_TEXT}\n\n"
            "On this tab, categories 23 and 24 (Georgia Cast Hires / Georgia Extras "
            "Hires) will be correct for the large majority of rows -- these are "
            "on-camera talent/performer payments. But this is steering, not a "
            "restriction: if a row's title/context clearly indicates something else "
            "(e.g. an agent fee, a per diem), use the correct category instead of "
            "defaulting to 23/24.\n\n"
            "Rows to classify:\n"
            f"{json.dumps(items, indent=2)}\n\n"
            "Return ONLY a JSON object mapping each row's index (as a string) to its "
            'AICP category number (an integer): {"0": 23, "1": 24, ...}\n'
            "Every index above must appear in your response. No explanation. No "
            "markdown. No code fences. JSON object only."
        )

        resp = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=4096,
            response_format={"type": "json_object"},
        )
        raw = json.loads(resp.choices[0].message.content)
        if not isinstance(raw, dict):
            return

        for key, val in raw.items():
            try:
                idx = int(key)
                code = int(val)
            except (TypeError, ValueError):
                continue
            if 0 <= idx < len(rows) and 1 <= code <= 25:
                rows[idx]['aicp_code'] = code
    except Exception:
        pass  # leave every aicp_code as None -- reviewable by hand, not fatal


def extract_talent(
    pdf_files: list[tuple[str, bytes]],  # [(filename, bytes), ...]
    ptip_bytes_list: list[bytes] | None,  # one Extreme Reach PTIP file per invoice, or a single consolidated file
    project_title: str,
    workbook_type: str,
    openai_key: str = "",
) -> dict:
    """
    Run the full talent extraction.

    GA productions get one Extreme Reach PTIP report PER INVOICE rather than
    IL's single consolidated file, so ptip_bytes_list is combined the same
    way parse_teams_ptip_xlsx already combines multiple Teams PTIP files --
    parse each independently (header row position/columns can vary file to
    file) and concatenate the row lists. Works unchanged for IL's
    single-file case too (a list of length 1).

    Returns the /extract-talent response dict:
    {rows, ptip_excel_b64, summary: {total_rows, invoices_with_pdf,
     invoices_ptip_only, duplicates_found, issues}}
    """
    issues: list[str] = []

    # ── Step 1: Parse PDFs ────────────────────────────────────────────────────
    pdf_invoices: dict[str, dict] = {}  # invoice_no → parse result

    for filename, data in pdf_files:
        result = parse_er_invoice_pdf(data)
        if result is None:
            issues.append(
                f"{filename}: not recognized as an Extreme Reach talent invoice "
                "(AI fallback not yet implemented for talent — review manually)"
            )
            continue
        inv_no = result['invoice_no']
        if not inv_no:
            issues.append(f"{filename}: could not extract invoice number")
            continue
        if inv_no in pdf_invoices:
            issues.append(f"{filename}: duplicate invoice number {inv_no} — skipped")
            continue
        pdf_invoices[inv_no] = result

    received_invoice_nos: set[str] = set(pdf_invoices.keys())

    # ── Step 2: Parse PTIP (one or more files, combined) ────────────────────────
    ptip_rows: list[dict] = []
    header_cols: list[str] = []
    # First file only, for organized-PTIP styling reference (logo, header
    # colors, fonts) -- build_organized_ptip_xlsx uses this purely cosmetically,
    # every file's row DATA still goes into the combined ptip_rows list below.
    original_ptip_bytes = ptip_bytes_list[0] if ptip_bytes_list else None

    for file_i, pb in enumerate(ptip_bytes_list or []):
        if not pb:
            continue
        file_rows, file_issues, file_header_cols = parse_ptip_xlsx(pb)
        if len(ptip_bytes_list) > 1:
            issues.extend(f"PTIP file {file_i + 1}: {msg}" for msg in file_issues)
        else:
            issues.extend(file_issues)
        ptip_rows.extend(file_rows)
        if not header_cols and file_header_cols:
            header_cols = file_header_cols

    # ── Step 3: Determine scenario ────────────────────────────────────────────
    has_pdf  = bool(pdf_invoices)
    has_ptip = bool(ptip_rows)

    if has_pdf and has_ptip:
        scenario = 'C'
    elif has_pdf:
        scenario = 'A'
    elif has_ptip:
        scenario = 'B'
    else:
        return {
            'rows': [],
            'ptip_excel_b64': None,
            'summary': {
                'total_rows': 0,
                'invoices_with_pdf': [],
                'invoices_ptip_only': [],
                'duplicates_found': 0,
                'issues': issues + ['No PDFs or PTIP file provided.'],
            },
        }

    # ── Step 4: Build PDF name count map (for duplicate detection) ────────────
    pdf_name_counts: dict[tuple, int] = defaultdict(int)
    if scenario == 'C':
        for inv_no, inv_data in pdf_invoices.items():
            for tr in inv_data['talent_rows']:
                full_norm, _ = _normalize_for_match(tr['name'])
                pdf_name_counts[(inv_no, full_norm)] += 1

    # ── Step 5: Detect duplicates in PTIP ────────────────────────────────────
    duplicate_indices: set[int] = set()
    if has_ptip:
        duplicate_indices = _detect_duplicates(ptip_rows, pdf_name_counts)

    # ── Step 6: Group PTIP rows by invoice for ordered assembly ─────────────
    ptip_by_invoice: dict[str, list[dict]] = defaultdict(list)
    ptip_orig_by_invoice: dict[str, list[int]] = defaultdict(list)
    for idx, pr in enumerate(ptip_rows):
        inv_no = str(pr.get('Invoice Number', '') or '').strip()
        ptip_by_invoice[inv_no].append(pr)
        ptip_orig_by_invoice[inv_no].append(idx)

    # ── Step 7: Assemble workbook rows (invoice-sorted, PDF order within each) ─
    workbook_rows: list[dict] = []
    item_no = 1
    all_llm_matches: dict[str, dict[str, str]] = {}  # {inv_no: {invoice_name: ptip_name}}

    all_invoice_nos = sorted(
        set(received_invoice_nos) | set(ptip_by_invoice.keys()),
        key=lambda x: x.zfill(20),
    )

    for inv_no in all_invoice_nos:
        inv_data = pdf_invoices.get(inv_no)
        inv_ptip_rows = ptip_by_invoice.get(inv_no, [])
        inv_ptip_orig = ptip_orig_by_invoice.get(inv_no, [])
        received = (inv_no in received_invoice_nos)
        inv_has_ptip = bool(inv_ptip_rows)

        if inv_data:
            # PDF exists — iterate talent rows in PDF order; match each to a PTIP row
            ptip_consumed: set[int] = set()
            first_talent_seen = False

            # Pass 1: exact normalized matching
            matched_pairs: list[tuple] = []  # (tr, ptip_match | None, ptip_match_local_i | None)
            for tr in inv_data['talent_rows']:
                ptip_match: dict | None = None
                ptip_match_local_i: int | None = None
                tr_full_norm, _ = _normalize_for_match(tr['name'])

                for local_i, pr in enumerate(inv_ptip_rows):
                    if local_i in ptip_consumed:
                        continue
                    if inv_ptip_orig[local_i] in duplicate_indices:
                        continue
                    pr_name = str(pr.get('Talent Name', '') or '').split('\n')[0].strip()
                    pr_full_norm, _ = _normalize_for_match(pr_name)
                    if pr_full_norm == tr_full_norm:
                        ptip_match = pr
                        ptip_match_local_i = local_i
                        ptip_consumed.add(local_i)
                        break

                matched_pairs.append((tr, ptip_match, ptip_match_local_i))

            # LLM fallback: attempt fuzzy matches for rows still unmatched after exact pass.
            # Used to require at least one exact match first (ptip_consumed non-empty) as
            # evidence the invoice/PTIP pair actually correspond -- but they're already
            # scoped to the same invoice number on both sides before we ever get here, which
            # is real structural evidence on its own. Gating on a prior exact match meant a
            # PDF layout that corrupts EVERY name in an invoice (e.g. a cast-category code
            # glued onto the name with no separator) got zero exact matches, so the fallback
            # never even ran -- exactly the invoice that most needs it.
            llm_reasons: dict[str, str] = {}
            if openai_key and inv_has_ptip:
                unmatched_inv = [
                    tr for tr, pm, _ in matched_pairs
                    if pm is None
                ]
                remaining_ptip_local = [
                    li for li in range(len(inv_ptip_rows))
                    if li not in ptip_consumed
                    and inv_ptip_orig[li] not in duplicate_indices
                ]
                if unmatched_inv and remaining_ptip_local:
                    inv_names = [tr['name'] for tr in unmatched_inv]
                    # Keep the full (possibly two-line) PTIP name intact here --
                    # a loan-out's PDF row is sometimes labeled with just the
                    # corp name (e.g. "Aspen Kennedy Inc"), which only the LLM
                    # can recognize if it can see that same corp name already
                    # sitting in the PTIP entry's own "First Last\n(Corp Name)"
                    # text. Truncating to the first line hides that clue.
                    remaining_ptip_names = [
                        str(inv_ptip_rows[li].get('Talent Name', '') or '').strip()
                        for li in remaining_ptip_local
                    ]
                    llm_map, llm_reasons = _llm_fuzzy_match_talent(inv_names, remaining_ptip_names, openai_key)
                    if llm_map:
                        all_llm_matches[inv_no] = llm_map
                        ptip_name_to_local = {
                            str(inv_ptip_rows[li].get('Talent Name', '') or '').strip(): li
                            for li in remaining_ptip_local
                        }
                        for idx, (tr, pm, pm_li) in enumerate(matched_pairs):
                            if pm is not None:
                                continue
                            matched_ptip_name = llm_map.get(tr['name'])
                            if matched_ptip_name:
                                li = ptip_name_to_local.get(matched_ptip_name)
                                if li is not None and li not in ptip_consumed:
                                    matched_pairs[idx] = (tr, inv_ptip_rows[li], li)
                                    ptip_consumed.add(li)

            # Pass 2: emit rows in PDF order
            for tr, ptip_match, ptip_match_local_i in matched_pairs:
                # Invoice-level tax stacking only when no PTIP exists for this invoice
                is_first_tax = (not inv_has_ptip) and (not tr['is_agent']) and (not first_talent_seen)
                if not tr['is_agent']:
                    first_talent_seen = True
                eff_scenario = 'A' if not inv_has_ptip else scenario

                pdf_loan_out_company = tr.get('loan_out_company', '')
                row = _build_row(
                    item_no=item_no,
                    ptip=ptip_match,
                    pdf_talent=tr,
                    pdf_invoice=inv_data,
                    received_invoice=True,
                    is_duplicate=False,
                    first_tax_row=is_first_tax,
                    scenario=eff_scenario,
                    ptip_row_no=(inv_ptip_orig[ptip_match_local_i] + 1) if ptip_match_local_i is not None else None,
                    workbook_type=workbook_type,
                    loan_out=bool(pdf_loan_out_company),
                    loan_out_company=pdf_loan_out_company,
                )
                if ptip_match is None:
                    reason = llm_reasons.get(tr['name'], '')
                    if reason:
                        llm_note = f"[LLM no match] {reason}"
                        row['notes'] = f"{row['notes']}; {llm_note}" if row.get('notes') else llm_note
                workbook_rows.append(row)
                item_no += 1

            # Append any PTIP rows for this invoice not matched to a PDF row
            for local_i, (pr, orig_idx) in enumerate(zip(inv_ptip_rows, inv_ptip_orig)):
                if local_i in ptip_consumed:
                    continue
                is_dup = orig_idx in duplicate_indices
                row = _build_row(
                    item_no=item_no if not is_dup else 0,
                    ptip=pr,
                    pdf_talent=None,
                    pdf_invoice=inv_data,
                    received_invoice=False,  # talent not found on received PDF
                    is_duplicate=is_dup,
                    first_tax_row=False,
                    scenario=scenario,
                    ptip_row_no=orig_idx + 1,
                    workbook_type=workbook_type,
                )
                workbook_rows.append(row)
                if not is_dup:
                    item_no += 1

        else:
            # PTIP-only invoice (no PDF received)
            for pr, orig_idx in zip(inv_ptip_rows, inv_ptip_orig):
                is_dup = orig_idx in duplicate_indices
                row = _build_row(
                    item_no=item_no if not is_dup else 0,
                    ptip=pr,
                    pdf_talent=None,
                    pdf_invoice=None,
                    received_invoice=False,
                    is_duplicate=is_dup,
                    first_tax_row=False,
                    scenario=scenario,
                    ptip_row_no=orig_idx + 1,
                    workbook_type=workbook_type,
                )
                workbook_rows.append(row)
                if not is_dup:
                    item_no += 1

    # ── Step 8: Note invoices with PDF but no PTIP match ─────────────────────
    if has_ptip and has_pdf:
        ptip_inv_nos = {str(r.get('Invoice Number', '') or '').strip() for r in ptip_rows}
        for inv_no in sorted(received_invoice_nos):
            if inv_no not in ptip_inv_nos:
                issues.append(
                    f"Invoice {inv_no}: PDF received, not in PTIP "
                    "— rows included with 'Included in PTIP report' = NO"
                )

    # ── Step 9: Build organized PTIP ─────────────────────────────────────────
    # Build PDF name-order map so rows sort to match PDF order within each invoice.
    er_pdf_order: dict[str, list[str]] = {}
    for inv_no, inv_data in pdf_invoices.items():
        # Sort by row_no if present; fall back to list index (also PDF order)
        # so future formats without explicit numbers still sort correctly.
        ordered_rows = sorted(
            enumerate(inv_data['talent_rows']),
            key=lambda t: t[1].get('row_no') or t[0],
        )
        # Include ALL rows (models and agents) so the PTIP mirrors the
        # alternating model/agent order that ER invoices use.
        er_pdf_order[inv_no] = [
            _normalize_for_match(tr['name'])[0]
            for _, tr in ordered_rows
        ]

    # Patch er_pdf_order with LLM matches: replace normalized invoice names with
    # normalized PTIP names so build_organized_ptip_xlsx can find exact matches.
    for inv_no, llm_inv_matches in all_llm_matches.items():
        if inv_no not in er_pdf_order:
            continue
        norm_inv_to_norm_ptip = {
            _normalize_for_match(inv_name)[0]: _normalize_for_match(ptip_name)[0]
            for inv_name, ptip_name in llm_inv_matches.items()
        }
        er_pdf_order[inv_no] = [
            norm_inv_to_norm_ptip.get(n, n)
            for n in er_pdf_order[inv_no]
        ]

    ptip_excel_b64 = None
    if has_ptip and original_ptip_bytes and header_cols:
        try:
            ptip_excel_bytes = build_organized_ptip_xlsx(
                ptip_rows,
                header_cols,
                received_invoice_nos,
                duplicate_indices,
                original_ptip_bytes,
                pdf_order_within_invoice=er_pdf_order,
            )
            ptip_excel_b64 = base64.b64encode(ptip_excel_bytes).decode()
        except Exception as e:
            issues.append(f"Organized PTIP build error: {e}")

    # ── Step 10: AICP classification (GA only) ──────────────────────────────────
    if _is_ga_workbook(workbook_type):
        _classify_aicp_codes_talent(workbook_rows, openai_key)

    # ── Step 11: Build summary ────────────────────────────────────────────────
    ptip_inv_nos = {str(r.get('Invoice Number', '') or '').strip() for r in ptip_rows}
    invoices_with_pdf  = sorted(received_invoice_nos & ptip_inv_nos) if has_ptip else sorted(received_invoice_nos)
    invoices_ptip_only = sorted(ptip_inv_nos - received_invoice_nos) if has_ptip else []
    duplicates_found   = len(duplicate_indices)
    non_dup_rows       = [r for r in workbook_rows if not r['is_duplicate']]

    return {
        'rows': workbook_rows,
        'ptip_excel_b64': ptip_excel_b64,
        'summary': {
            'total_rows':         len(non_dup_rows),
            'invoices_with_pdf':  invoices_with_pdf,
            'invoices_ptip_only': invoices_ptip_only,
            'duplicates_found':   duplicates_found,
            'issues':             issues,
        },
    }


# ── Teams Invoice PDF parsing ─────────────────────────────────────────────────

_TEAMS_AMOUNT_RE  = re.compile(r'^\d[\d,]*\.\d{2}$')
_TEAMS_DATE_RE    = re.compile(r'^\d{2}/\d{2}/\d{2}$')
_TEAMS_CAT_CODES  = {'P', 'EXB', 'PVO'}
_TEAMS_CAT_TITLES = {'P': 'Principal', 'EXB': 'Extra', 'PVO': 'PVO'}


_NAME_SUFFIXES = frozenset({'jr', 'sr', 'ii', 'iii', 'iv', 'v', '2nd', '3rd'})


def _fmt_teams_talent_name(name: str) -> str:
    """Format Teams talent name → 'Last, First' (title case).

    Handles both 'LAST [SUFFIX], FIRST [MIDDLE]' (PDF/PTIP comma format) and
    'FIRST [MIDDLE] LAST [SUFFIX]' (no-comma PTIP format).
    Strips middle initials (single letter ± period) and name suffixes (Jr/Sr/II/III...).
    Corp names in parentheses are title-cased and re-appended.
    """
    if not name:
        return ''
    corp = ''
    m = re.match(r'^(.*?)\s*\((.+)\)\s*$', name.strip())
    if m:
        base, corp = m.group(1).strip(), m.group(2).strip()
    else:
        base = name.strip()

    if ',' in base:
        last_raw, _, first_mid_raw = base.partition(',')
        # Strip suffix tokens from last name (e.g. "ROGERS JR" → "ROGERS")
        last_tokens = [t for t in last_raw.strip().split()
                       if t.lower().rstrip('.') not in _NAME_SUFFIXES]
        last = ' '.join(last_tokens).title()
        # Strip single-letter middle initials; keep all remaining tokens as first name
        first_tokens = first_mid_raw.strip().split()
        non_init = [t for t in first_tokens if not re.match(r'^[A-Za-z]\.?$', t)]
        first = ' '.join(t.title() for t in non_init) if non_init else (
            first_tokens[0].title() if first_tokens else ''
        )
        result = f"{last}, {first}".strip(', ') if first else last
    else:
        # No-comma format: FIRST [MIDDLE] LAST [SUFFIX]
        tokens = base.strip().split()
        while tokens and tokens[-1].lower().rstrip('.') in _NAME_SUFFIXES:
            tokens.pop()
        if not tokens:
            result = ''
        elif len(tokens) == 1:
            result = tokens[0].title()
        else:
            inner = [t for t in tokens[1:-1] if not re.match(r'^[A-Za-z]\.?$', t)]
            cleaned = [tokens[0]] + inner + [tokens[-1]]
            result = f"{cleaned[-1].title()}, {cleaned[0].title()}"

    if corp:
        result = f"{result} ({corp.title()})"
    return result


def _parse_teams_crp_row(line: str) -> str | None:
    """If line is a Teams CRP (loan-out corp) row return corp name, else None."""
    m = re.match(r'^\*+\d+\s+(.+?)\s+CRP\s*$', line.strip())
    return m.group(1).strip() if m else None


def _parse_teams_sag_talent_row(line: str, has_apply_col: bool = False) -> dict | None:
    """
    Parse one talent data row from a Teams SAG invoice table.

    Row format (after SSN): Name [OS1] [OS2] [Cat] Cam St [Agnt]
                            Date [Spot] [Yr] [Apply] [Applied] Amount... Amount [P&H]

    has_apply_col: True when the invoice header includes an Apply/Applied column.
      In that case amounts[0] is the Apply credit (skip it); amounts[1] is Gross Wages.
    Returns dict or None if line is not a talent row.
    """
    line = line.strip()
    if not re.match(r'^\*+\d', line):
        return None

    tokens = line.split()
    if len(tokens) < 6:
        return None

    ssn  = tokens[0]
    rest = tokens[1:]

    # Find date (MM/DD/YY)
    date_idx = next((i for i, t in enumerate(rest) if _TEAMS_DATE_RE.match(t)), None)
    if date_idx is None:
        return None

    before_date = rest[:date_idx]
    after_date  = list(rest[date_idx + 1:])

    # Pop amounts from end of after_date
    amounts: list[float] = []
    while after_date and _TEAMS_AMOUNT_RE.match(after_date[-1]):
        amounts.insert(0, _to_float(after_date.pop()))

    if not amounts:
        return None

    work_date = _expand_date_yy(rest[date_idx])

    # Find cam code (ON/OFF) in before_date — serves as the anchor
    cam_idx = next(
        (i for i, t in enumerate(before_date) if t.upper() in ('ON', 'OFF')), None
    )
    if cam_idx is None:
        return None
    cam = before_date[cam_idx].upper()

    # State: first 2-letter token after cam
    work_state = ''
    if cam_idx + 1 < len(before_date):
        cand = before_date[cam_idx + 1]
        if re.match(r'^[A-Z]{2}$', cand):
            work_state = cand

    # Cat: one of P/EXB/PVO immediately before cam (if present)
    cat = ''
    cat_idx = None
    if cam_idx > 0 and before_date[cam_idx - 1] in _TEAMS_CAT_CODES:
        cat     = before_date[cam_idx - 1]
        cat_idx = cam_idx - 1

    # Work backwards from cat (or cam) to find OS2 / OS1 boundary of name
    pre_pos = cat_idx if cat_idx is not None else cam_idx

    os2_idx = None
    if pre_pos > 0:
        t = before_date[pre_pos - 1]
        if re.match(r'^[A-Z]{2,4}$', t) and ',' not in t:
            os2_idx = pre_pos - 1

    os1_idx = None
    probe = (os2_idx if os2_idx is not None else pre_pos) - 1
    if probe >= 0 and before_date[probe].isdigit():
        os1_idx = probe

    name_end = min(x for x in [os1_idx, os2_idx, pre_pos] if x is not None)
    name = ' '.join(before_date[:name_end]).strip()

    if len(amounts) == 1:
        # Single trailing amount -- seen on SAG/AFTRA guarantee-installment rows,
        # which carry no separate Misc Payment breakdown. The Team Companies
        # books guarantee payments through the P&H line rather than Gross Wages,
        # so an OFF-camera lone amount is P&H; an ON-camera one is Gross Wages.
        # Confirmed against invoice 26225373's own footer totals (Wages & Misc.
        # Payments = 0.00, Pension & Health Cont = this row's one amount, exactly).
        gross = 0.0 if cam == 'OFF' else amounts[0]
        misc  = 0.0
        pah   = amounts[0] if cam == 'OFF' else 0.0
    elif has_apply_col:
        # amounts = [Apply(skip), GrossWages, P&H]  — or [Apply, Gross, Misc, P&H] if 4
        # 2-amount rows on an Apply invoice are treated as plain [Gross/Misc, P&H]
        if len(amounts) == 2:
            gross = 0.0 if cam == 'OFF' else amounts[0]
            misc  = amounts[0] if cam == 'OFF' else 0.0
            pah   = amounts[1]
        elif len(amounts) == 3:
            gross, misc, pah = amounts[1], 0.0, amounts[2]
        else:
            gross, misc, pah = amounts[1], amounts[2], amounts[3]
    else:
        if len(amounts) == 2:
            gross, misc, pah = (0.0, amounts[0], amounts[1]) if cam == 'OFF' else (amounts[0], 0.0, amounts[1])
        else:
            gross, misc, pah = amounts[0], amounts[1], amounts[2]

    return {
        'ssn':         ssn,
        'name':        name,
        'cat':         cat,
        'cam':         cam,
        'work_state':  work_state,
        'work_date':   work_date,
        'gross_wages': round(gross, 2),
        'misc_pmt':    round(misc, 2),
        'pah':         round(pah, 2),
        'title':       _TEAMS_CAT_TITLES.get(cat, ''),
        'corp_name':   '',
    }


def parse_teams_sag_invoice_pdf(pdf_bytes: bytes) -> dict | None:
    """Parse a Teams SAG union invoice PDF."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = '\n'.join(pg.extract_text() or '' for pg in pdf.pages)
    except Exception:
        return None

    if 'Soc. Sec # Performer Name' not in text:
        return None

    inv_no = ''
    # A stray colon from an overlapping PDF text run can land mid-digit-string
    # (observed as literal "Invoice #2:6225373" for invoice 26225373) -- allow
    # embedded colons in the capture and strip them, or a plain \d+ match would
    # silently truncate to the digits before the colon.
    m = re.search(r'Invoice\s*#[:\s]*([\d:]+)', text)
    if m:
        inv_no = m.group(1).replace(':', '').strip()

    invoice_date = ''
    # Invoice date always precedes "Page:" on the same line; use that to avoid
    # matching "First Air Date" or "Expiration Date"
    m = re.search(r'\bDate\s+(\d{2}/\d{2}/\d{2})\s+Page:', text)
    if m:
        invoice_date = _expand_date_yy(m.group(1))

    pay_type = ''
    m = re.search(r'Payment Type\s+(\S+)', text)
    if m:
        pay_type = m.group(1).strip()

    commercial_id = ''
    m = re.search(r'Comml ID#\s+(\S+)', text)
    if m:
        commercial_id = m.group(1).strip()

    # Footer invoice-level totals
    m_wages    = re.search(r'Wages\s*&\s*Misc\.\s*Payments\s+([\d,]+\.\d{2})', text)
    m_taxes    = re.search(r'Payroll Taxes\s+([\d,]+\.\d{2})', text)
    m_pah      = re.search(r'Pension\s*&\s*Health\s*Cont\s+([\d,]+\.\d{2})', text)
    m_handling = re.search(r'Handling\s+([\d,]+\.\d{2})', text)
    # TS Signatory Fee is a real, occasionally-present footer line -- when it's
    # there, it's the actual signatory fee for the invoice, not something to
    # estimate. Business Affairs Fee / Wire Fee - Domestic are one-off invoice
    # charges with no dedicated column on the GA template; they get summed into
    # "Other Fees" with a note describing what's in it.
    m_signatory = re.search(r'TS Signatory Fee\s+([\d,]+\.\d{2})', text)
    m_biz_aff   = re.search(r'Business Affairs Fee\s+([\d,]+\.\d{2})', text)
    m_wire      = re.search(r'Wire Fee\s*-?\s*Domestic\s+([\d,]+\.\d{2})', text)

    total_er_tax       = _to_float(m_taxes.group(1)     if m_taxes     else 0)
    total_pah          = _to_float(m_pah.group(1)       if m_pah       else 0)
    total_handling     = _to_float(m_handling.group(1)  if m_handling  else 0)
    total_signatory_fee = _to_float(m_signatory.group(1) if m_signatory else 0)

    other_fees_detail: list[dict] = []
    if m_biz_aff:
        other_fees_detail.append({'label': 'Business Affairs Fee', 'amount': _to_float(m_biz_aff.group(1))})
    if m_wire:
        other_fees_detail.append({'label': 'Wire Fee', 'amount': _to_float(m_wire.group(1))})
    total_other_fees = round(sum(d['amount'] for d in other_fees_detail), 2)

    # Auto-detect whether the first amount in each row is an Apply credit (skip it)
    # or Gross Wages.  Parse with both interpretations and keep the one whose
    # sum(gross+misc) is closest to the footer Wages & Misc. Payments total.
    _footer_wages = _to_float(m_wages.group(1) if m_wages else 0)

    def _build_talent_rows(lines, has_apply_col):
        rows: list[dict] = []
        for line in lines:
            s = line.strip()
            crp = _parse_teams_crp_row(s)
            if crp is not None:
                if rows:
                    rows[-1]['corp_name'] = crp
                    rows[-1]['name'] = f"{rows[-1]['name']} ({crp})"
            else:
                r = _parse_teams_sag_talent_row(s, has_apply_col=has_apply_col)
                if r:
                    rows.append(r)
        return rows

    _all_lines = text.split('\n')
    _rows_plain = _build_talent_rows(_all_lines, False)
    _rows_apply = _build_talent_rows(_all_lines, True)

    _sum_plain = sum(r['gross_wages'] + r['misc_pmt'] for r in _rows_plain)
    _sum_apply = sum(r['gross_wages'] + r['misc_pmt'] for r in _rows_apply)

    if (_footer_wages > 0 and
            abs(_sum_apply - _footer_wages) < abs(_sum_plain - _footer_wages)):
        talent_rows = _rows_apply
    else:
        talent_rows = _rows_plain

    return {
        'invoice_no':     inv_no,
        'invoice_date':   invoice_date,
        'pay_type':       pay_type,
        'commercial_id':  commercial_id,
        'total_wages_misc': _to_float(m_wages.group(1) if m_wages else 0),
        'total_er_tax':   total_er_tax,
        'total_pah':      total_pah,
        'total_handling': total_handling,
        'total_wc':       0.0,
        'total_signatory_fee': total_signatory_fee,
        'total_other_fees':    total_other_fees,
        'other_fees_detail':   other_fees_detail,
        'talent_rows':    talent_rows,
        'union_type':     'SAG',
        'format':         'teams_sag',
    }


def parse_teams_nonunion_invoice_pdf(pdf_bytes: bytes) -> dict | None:
    """Parse a Teams non-union employer-cost invoice PDF (2-page format)."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = '\n'.join(pg.extract_text() or '' for pg in pdf.pages)
    except Exception:
        return None

    if 'Taxable Wages' not in text or 'Employer of Record' not in text:
        return None
    if 'Soc. Sec # Performer Name' in text:
        return None  # SAG invoice, not non-union

    inv_no = ''
    # A stray colon from an overlapping PDF text run can land mid-digit-string
    # (observed as literal "Invoice #2:6225373" for invoice 26225373) -- allow
    # embedded colons in the capture and strip them, or a plain \d+ match would
    # silently truncate to the digits before the colon.
    m = re.search(r'Invoice\s*#[:\s]*([\d:]+)', text)
    if m:
        inv_no = m.group(1).replace(':', '').strip()

    invoice_date = ''
    m = re.search(r'Inv\.\s*Date[:\s]+(\d{1,2}/\d{1,2}/\d{2})', text)
    if m:
        invoice_date = _expand_date_yy(m.group(1))

    # Page 2 talent table: lines between "Payee Name … SSN/FED WS" header and "TOTALS:"
    talent_rows: list[dict] = []
    lines = text.split('\n')
    in_table = False
    for line in lines:
        stripped = line.strip()
        if 'Payee Name' in stripped and 'SSN/FED WS' in stripped:
            in_table = True
            continue
        if in_table and stripped.upper().startswith('TOTALS'):
            break
        if not in_table or not stripped:
            continue
        # Talent row pattern: LAST, FIRST *****NNNNSS amounts...
        m = re.match(r'^([A-Z][A-Z ,\.]+?)\s+(\*+\d+[A-Z]{2})\s+(.+)$', stripped)
        if not m:
            continue
        name     = m.group(1).strip().rstrip(',').strip()
        ssn_ws   = m.group(2)
        work_st  = ssn_ws[-2:]
        ssn      = ssn_ws[:-2]
        vals     = m.group(3).split()
        # cols: taxable corp_wages reimb FICA Medicare FUTA SUI SDI Other WC RS P&H Handling Deductions Total
        if len(vals) < 13:
            continue
        taxable  = _to_float(vals[0])
        fica     = _to_float(vals[3])
        medicare = _to_float(vals[4])
        futa     = _to_float(vals[5])
        sui      = _to_float(vals[6])
        sdi      = _to_float(vals[7])
        other    = _to_float(vals[8])
        wc       = _to_float(vals[9])
        pah      = _to_float(vals[11])
        handling = _to_float(vals[12])
        er_tax   = round(fica + medicare + futa + sui + sdi + other, 2)
        talent_rows.append({
            'ssn':         ssn,
            'name':        name,
            'work_state':  work_st,
            'gross_wages': taxable,
            'misc_pmt':    0.0,
            'pah':         pah,
            'er_tax':      er_tax,
            'wc':          wc,
            'handling':    handling,
            'title':       '',
            'cat':         '',
            'cam':         'ON',
            'work_date':   '',
            'corp_name':   '',
        })

    if not talent_rows:
        return None

    return {
        'invoice_no':     inv_no,
        'invoice_date':   invoice_date,
        'pay_type':       'Non-union',
        'commercial_id':  '',
        'total_wages_misc': sum(tr['gross_wages'] for tr in talent_rows),
        'total_er_tax':   round(sum(tr['er_tax']   for tr in talent_rows), 2),
        'total_pah':      round(sum(tr['pah']       for tr in talent_rows), 2),
        'total_handling': round(sum(tr['handling']  for tr in talent_rows), 2),
        'total_wc':       round(sum(tr['wc']        for tr in talent_rows), 2),
        'talent_rows':    talent_rows,
        'union_type':     'Non-union',
        'format':         'teams_nonunion',
    }


def parse_teams_invoice_pdf(pdf_bytes: bytes) -> dict | None:
    """Try SAG then non-union Teams parser; return None if neither matches."""
    result = parse_teams_sag_invoice_pdf(pdf_bytes)
    if result:
        return result
    return parse_teams_nonunion_invoice_pdf(pdf_bytes)


# ── Teams PTIP parsing ────────────────────────────────────────────────────────

def _normalize_teams_col(v) -> str:
    """Strip leading/embedded \\n from Teams PTIP column header cells."""
    if v is None:
        return ''
    return re.sub(r'\s*\n\s*', ' ', str(v)).strip()


def _normalize_teams_ptip_row(row: dict) -> dict:
    """Convert a Teams PTIP row dict (normalized headers) to ER-compatible keys for _build_row()."""
    fica     = _to_float(row.get('FICA', 0))
    medicare = _to_float(row.get('Medicare', 0))
    futa     = _to_float(row.get('FUTA', 0))
    sui      = _to_float(row.get('SUI', 0))
    sdi      = _to_float(row.get('SDI', 0))
    other    = _to_float(row.get('Other Taxes', 0))
    # SDI is reported separately below (State Disability Withheld) -- _build_row's
    # _TOTAL_KEYS sums Employer Taxes and State Disability Withheld independently,
    # so folding SDI into er_tax here would double-count it in ptip_amount.
    er_tax   = round(fica + medicare + futa + sui + other, 2)

    street    = str(row.get('Address Line 1', '') or '').strip()
    city      = str(row.get('City', '') or '').strip()
    res_state = str(row.get('Resident State', '') or '').strip()
    postal    = str(row.get('Postal Code', '') or '').strip()
    addr_parts = [street]
    if city and res_state and postal:
        addr_parts.append(f'{city}, {res_state} {postal}')
    full_address = '\n'.join(p for p in addr_parts if p)

    # Prefer Invoice Number column; fall back to Invoice
    inv_no = str(row.get('Invoice Number', '') or row.get('Invoice', '') or '').strip()

    occupation = str(row.get('Occupation', '') or '').strip().title()

    pah_raw = row.get('Pension / H&W') if row.get('Pension / H&W') is not None else row.get('Pension/H&W')
    wc_raw  = row.get('Workers Comp.') if row.get('Workers Comp.') is not None else row.get('Workers Comp')

    return {
        'Invoice Number':      inv_no,
        'Talent Name':         str(row.get('Name', '') or '').strip(),
        'Wages':               _to_float(row.get('Gross Wages', 0)),
        'Employer Taxes':      er_tax,
        'Workers Compensation': _to_float(wc_raw or 0),
        'Handling Fee':        _to_float(row.get('Handling', 0)),
        'P&H':                 _to_float(pah_raw or 0),
        'Signatory Fee':       0.0,
        'Cast Category':       occupation,
        'Work State':          str(row.get('Work State', '') or '').strip(),
        'Talent Address':      full_address,
        'Check Date':          row.get('Invoice Date'),
        # 'Job Number' is what GA's Teams PTIP export actually calls this column
        # (confirmed against a real GA file); 'Project ID' kept as a fallback in
        # case some other Teams export still uses that name.
        'Commercial Id':       str(row.get('Job Number', '') or row.get('Project ID', '') or '').strip(),
        'SSN':                 str(row.get('SSN', '') or '').strip(),
        'FEIN':                '',
        'Check Number':        '',
        # GA-only fields (blank/0 for IL Teams files that don't have these columns).
        # 'GA SIT' is confirmed as the GA state-income-tax-withheld column name;
        # the plain 'SIT' column seen alongside it is NOT mapped here -- its exact
        # meaning (resident-state SIT when different from GA?) isn't confirmed yet.
        'State Tax Withheld':        _to_float(row.get('GA SIT', 0)),
        'Local Tax Withheld':        0.0,
        'State Disability Withheld': sdi,
    }


def parse_teams_ptip_xlsx(
    xlsx_bytes_list: list[bytes],
) -> tuple[list[dict], list[dict], list[str], list[str]]:
    """
    Read one or more Teams PTIP Excel files and combine.

    Teams sends one file per commercial ID; rows with empty Name are
    continuation rows for the same talent on an additional invoice — the
    Name/address is back-filled from the immediately preceding full row.

    Returns (normalized_rows, raw_rows, raw_headers, issues).
    """
    normalized_rows: list[dict] = []
    raw_rows:        list[dict] = []
    raw_headers:     list[str]  = []
    issues:          list[str]  = []

    for file_i, xlsx_bytes in enumerate(xlsx_bytes_list):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            ws = wb.active

            # Find header row: contains 'Name' and a 'Gross' column within first 12 rows
            hdr_row_idx: int | None = None
            for r_idx, row_vals in enumerate(ws.iter_rows(max_row=12, values_only=True)):
                norms = [_normalize_teams_col(v) for v in row_vals]
                if 'Name' in norms and any('Gross' in n for n in norms):
                    hdr_row_idx = r_idx
                    break

            if hdr_row_idx is None:
                issues.append(f"PTIP file {file_i + 1}: could not find header row")
                continue

            file_headers_raw = list(ws.iter_rows(
                min_row=hdr_row_idx + 1,
                max_row=hdr_row_idx + 1,
                values_only=True,
            ))[0]
            file_headers = [_normalize_teams_col(v) for v in file_headers_raw]
            # Drop trailing empty header columns
            while file_headers and not file_headers[-1]:
                file_headers.pop()

            if file_i == 0:
                raw_headers = file_headers

            invoice_col = next(
                (h for h in file_headers if h == 'Invoice Number'),
                next((h for h in file_headers if h == 'Invoice'), None),
            )

            last_full: dict = {}

            for row_vals in ws.iter_rows(min_row=hdr_row_idx + 2, values_only=True):
                # Skip truly blank rows (all columns empty).
                # Do NOT use [:5] — some PTIP formats have continuation rows that
                # are blank in the first 5 (name/SSN/address) but have invoice
                # number and amounts in later columns.
                if not any(v is not None and str(v).strip() not in ('', 'None')
                           for v in row_vals):
                    continue

                row = dict(zip(file_headers, row_vals))

                # Skip disclaimer row
                name_val = str(row.get('Name', '') or '').strip()
                if name_val.startswith('This document'):
                    continue

                # Skip rows with no invoice number
                if invoice_col:
                    inv_val = str(row.get(invoice_col, '') or '').strip()
                    if not inv_val:
                        continue

                # Back-fill name/address from previous full row when Name is empty
                if not name_val and last_full:
                    for k in ('Name', 'SSN', 'Work State', 'Address Line 1',
                              'Address Line 2', 'City', 'Resident State',
                              'Postal Code', 'Occupation', 'EOR', 'Job Number'):
                        if k in last_full and not str(row.get(k, '') or '').strip():
                            row[k] = last_full[k]
                else:
                    last_full = dict(row)

                raw_rows.append(dict(row))
                normalized_rows.append(_normalize_teams_ptip_row(row))

        except Exception as e:
            issues.append(f"PTIP file {file_i + 1}: parse error — {e}")

    return normalized_rows, raw_rows, raw_headers, issues


# ── Teams Sorted PTIP builder ─────────────────────────────────────────────────

def build_teams_sorted_ptip_xlsx(
    ptip_rows_raw:             list[dict],
    raw_headers:               list[str],
    received_invoice_nos:      set[str],
    pdf_order_within_invoice:  dict[str, list[str]],  # inv_no → [norm_name, ...]
) -> bytes:
    """
    Build the Sorted/Combined PTIP Excel deliverable for Teams.

    Layout: '#' column prepended + original PTIP columns + 'Payroll Tax' appended.
    Sorted by Invoice Number; '#' resets per invoice and follows PDF talent order
    when PDF was received for that invoice.
    """
    from copy import copy as _copy
    from openpyxl.utils import get_column_letter

    invoice_key = 'Invoice Number' if 'Invoice Number' in raw_headers else (
        'Invoice' if 'Invoice' in raw_headers else (raw_headers[0] if raw_headers else '')
    )

    # Columns that sum into Payroll Tax
    _TAX_COLS = ('FICA', 'Medicare', 'FUTA', 'SUI', 'SDI', 'Other Taxes')

    out_headers = ['#'] + raw_headers + ['Payroll Tax']
    n_cols = len(out_headers)

    # Group raw rows by invoice
    inv_rows: dict[str, list[dict]] = defaultdict(list)
    for row in ptip_rows_raw:
        inv_no = str(row.get(invoice_key, '') or '').strip()
        inv_rows[inv_no].append(row)

    # Sort invoice numbers
    sorted_inv_nos = sorted(inv_rows.keys(), key=lambda x: x.zfill(20))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sorted PTIP'

    # Header row
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    hdr_font  = Font(bold=True, size=10)
    hdr_fill  = PatternFill('solid', fgColor='1F4E79')
    hdr_font_white = Font(bold=True, size=10, color='FFFFFF')
    thin_side = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)
    data_font = Font(size=10)
    data_align = Alignment(wrap_text=False, vertical='top')

    ws.row_dimensions[1].height = 30
    for col_i, col_name in enumerate(out_headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=col_name)
        cell.font   = hdr_font_white
        cell.fill   = hdr_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Column widths
    ws.column_dimensions['A'].width = 5  # #
    col_widths = {
        'Name': 22, 'SSN': 14, 'Work State': 8, 'Address Line 1': 22,
        'Address Line 2': 14, 'City': 14, 'Resident State': 8, 'Postal Code': 8,
        'Occupation': 16, 'EOR': 20, 'Job Number': 16, 'Invoice': 12,
        'Invoice Number': 12, 'Project ID': 14, 'Invoice Date': 12,
        'Gross Wages': 11, 'SIT': 9, 'FICA': 9, 'Medicare': 9,
        'FUTA': 9, 'SUI': 9, 'SDI': 9, 'Other Taxes': 9,
        'Handling': 9, 'Pension / H&W': 11, 'Workers Comp.': 11,
        'Total Hours': 9, 'Work Start': 12, 'Work End': 12, 'Payroll Tax': 11,
    }
    for col_i, col_name in enumerate(out_headers[1:], start=2):
        w = col_widths.get(col_name, 12)
        ws.column_dimensions[get_column_letter(col_i)].width = w

    # Data rows
    out_row = 2
    for inv_no in sorted_inv_nos:
        rows = inv_rows[inv_no]

        # Reorder within invoice to match PDF talent order (by SSN last-4)
        if inv_no in pdf_order_within_invoice and pdf_order_within_invoice[inv_no]:
            pdf_order = pdf_order_within_invoice[inv_no]
            ordered, remaining = [], list(rows)
            for pdf_s4 in pdf_order:
                if not pdf_s4:
                    continue
                for i, row in enumerate(remaining):
                    row_s4 = _ssn_last4(str(row.get('SSN', '') or ''))
                    if row_s4 and row_s4 == pdf_s4:
                        ordered.append(remaining.pop(i))
                        break
            ordered.extend(remaining)
            rows = ordered

        seq = 0
        for row in rows:
            received = inv_no in received_invoice_nos
            seq += 1 if received else 0
            hash_val = seq if received else ''

            payroll_tax = round(sum(_to_float(row.get(c, 0)) for c in _TAX_COLS), 2)

            row_vals = [hash_val] + [row.get(h) for h in raw_headers] + [payroll_tax or '']

            # Format date columns to MM/DD/YYYY
            date_cols = {'Invoice Date', 'Work Start', 'Work End'}
            date_col_indices = {
                col_i for col_i, h in enumerate(out_headers) if h in date_cols
            }

            ws.row_dimensions[out_row].height = 18
            for col_i, val in enumerate(row_vals, start=1):
                if col_i - 1 in date_col_indices:
                    val = _fmt_date(val) or val
                cell = ws.cell(row=out_row, column=col_i, value=val)
                cell.font      = data_font
                cell.border    = thin_border
                cell.alignment = data_align
            out_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Teams main extraction entry point ────────────────────────────────────────

def extract_teams_talent(
    pdf_files:       list[tuple[str, bytes]],  # [(filename, bytes), ...]
    ptip_bytes_list: list[bytes],
    project_title:   str,
    workbook_type:   str,
    openai_key:      str = '',
) -> dict:
    """
    Run the full Teams talent extraction.

    Returns the same shape as extract_talent():
    {rows, ptip_excel_b64, summary: {total_rows, invoices_with_pdf,
     invoices_ptip_only, duplicates_found, issues}}
    """
    issues: list[str] = []

    # ── Step 1: Parse PDFs ────────────────────────────────────────────────────
    pdf_invoices: dict[str, dict] = {}

    for filename, data in pdf_files:
        result = parse_teams_invoice_pdf(data)
        if result is None:
            issues.append(
                f"{filename}: not recognized as a Teams invoice "
                "(check for 'Soc. Sec # Performer Name' or 'Taxable Wages' in the PDF)"
            )
            continue
        inv_no = result['invoice_no']
        if not inv_no:
            issues.append(f"{filename}: could not extract invoice number")
            continue
        if inv_no in pdf_invoices:
            issues.append(f"{filename}: duplicate invoice number {inv_no} — skipped")
            continue
        pdf_invoices[inv_no] = result

    received_invoice_nos: set[str] = set(pdf_invoices.keys())

    # ── Step 2: Parse PTIP files ──────────────────────────────────────────────
    ptip_rows_normalized: list[dict] = []
    ptip_rows_raw:        list[dict] = []
    raw_headers:          list[str]  = []

    if ptip_bytes_list:
        ptip_rows_normalized, ptip_rows_raw, raw_headers, ptip_issues = \
            parse_teams_ptip_xlsx(ptip_bytes_list)
        issues.extend(ptip_issues)

    # ── Step 3: Determine scenario ────────────────────────────────────────────
    has_pdf  = bool(pdf_invoices)
    has_ptip = bool(ptip_rows_normalized)

    if has_pdf and has_ptip:
        scenario = 'C'
    elif has_pdf:
        scenario = 'A'
    elif has_ptip:
        scenario = 'B'
    else:
        return {
            'rows': [],
            'ptip_excel_b64': None,
            'summary': {
                'total_rows': 0,
                'invoices_with_pdf': [],
                'invoices_ptip_only': [],
                'duplicates_found': 0,
                'issues': issues + ['No PDFs or PTIP files provided.'],
            },
        }

    # ── Step 4: Group PTIP rows by invoice ───────────────────────────────────
    ptip_by_invoice: dict[str, list[tuple[dict, int]]] = defaultdict(list)
    for idx, pr in enumerate(ptip_rows_normalized):
        inv_no = str(pr.get('Invoice Number', '') or '').strip()
        ptip_by_invoice[inv_no].append((pr, idx))

    # ── Step 5: PDF talent order map (for Sorted PTIP ordering) ─────────────
    # Stored as SSN-last-4 lists so truncated PDF names don't break ordering.
    pdf_order_within_invoice: dict[str, list[str]] = {}
    for inv_no, inv_data in pdf_invoices.items():
        pdf_order_within_invoice[inv_no] = [
            _ssn_last4(tr.get('ssn', ''))
            for tr in inv_data['talent_rows']
        ]

    # ── Step 6: Assemble workbook rows ────────────────────────────────────────
    workbook_rows: list[dict] = []
    item_no = 1

    all_invoice_nos = sorted(
        set(received_invoice_nos) | set(ptip_by_invoice.keys()),
        key=lambda x: x.zfill(20),
    )

    for inv_no in all_invoice_nos:
        inv_data      = pdf_invoices.get(inv_no)
        inv_ptip_pairs = ptip_by_invoice.get(inv_no, [])
        inv_has_ptip  = bool(inv_ptip_pairs)

        if inv_data:
            ptip_consumed:   set[int] = set()
            first_talent_seen = False

            # Pass 1: exact normalized-name matching, falling back to SSN last-4
            matched_pairs: list[tuple] = []  # (tr, ptip_match | None, ptip_match_orig | None)
            for tr in inv_data['talent_rows']:
                tr_full_norm, _ = _normalize_for_match(tr['name'])

                ptip_match:       dict | None = None
                ptip_match_orig:  int | None  = None
                for local_i, (pr, orig_i) in enumerate(inv_ptip_pairs):
                    if local_i in ptip_consumed:
                        continue
                    pr_name = str(pr.get('Talent Name', '') or '').strip()
                    pr_norm, _ = _normalize_for_match(pr_name)
                    if pr_norm == tr_full_norm:
                        ptip_match      = pr
                        ptip_match_orig = orig_i
                        ptip_consumed.add(local_i)
                        break

                if ptip_match is None:
                    pdf_s4 = _ssn_last4(tr.get('ssn', ''))
                    if pdf_s4:
                        for local_i, (pr, orig_i) in enumerate(inv_ptip_pairs):
                            if local_i in ptip_consumed:
                                continue
                            if _ssn_last4(str(pr.get('SSN', '') or '')) == pdf_s4:
                                ptip_match      = pr
                                ptip_match_orig = orig_i
                                ptip_consumed.add(local_i)
                                break

                matched_pairs.append((tr, ptip_match, ptip_match_orig))

            # LLM fallback: a performer's PDF row (their own name) and their PTIP
            # row (often their loan-out corp's name, e.g. "Woods, Gloria" vs
            # "Glorilla Music LLC") don't share any normalized-name overlap, so
            # exact matching and SSN-last-4 both miss it -- same failure class
            # ER hit with a cast-category code glued onto the name.
            llm_reasons: dict[str, str] = {}
            # A fuzzy match that bridges a real person's PDF name to a
            # company-named PTIP entry is a loan-out payment, not a plain
            # name-formatting mismatch -- track which resolved indices this
            # applies to so Pass 2 can flag them and keep the real name.
            loan_out_idx: dict[int, str] = {}
            if openai_key and inv_has_ptip:
                unmatched_idx = [
                    i for i, (tr, pm, _) in enumerate(matched_pairs) if pm is None
                ]
                remaining_local = [
                    li for li in range(len(inv_ptip_pairs)) if li not in ptip_consumed
                ]
                if unmatched_idx and remaining_local:
                    inv_names = [matched_pairs[i][0]['name'] for i in unmatched_idx]
                    remaining_ptip_names = [
                        str(inv_ptip_pairs[li][0].get('Talent Name', '') or '').strip()
                        for li in remaining_local
                    ]
                    llm_map, llm_reasons = _llm_fuzzy_match_talent(inv_names, remaining_ptip_names, openai_key)
                    if llm_map:
                        ptip_name_to_local = {
                            str(inv_ptip_pairs[li][0].get('Talent Name', '') or '').strip(): li
                            for li in remaining_local
                        }
                        for i in unmatched_idx:
                            tr, pm, pm_orig = matched_pairs[i]
                            matched_name = llm_map.get(tr['name'])
                            if not matched_name:
                                continue
                            li = ptip_name_to_local.get(matched_name)
                            if li is not None and li not in ptip_consumed:
                                pr, orig_i = inv_ptip_pairs[li]
                                matched_pairs[i] = (tr, pr, orig_i)
                                ptip_consumed.add(li)
                                if _is_company_name(matched_name):
                                    loan_out_idx[i] = matched_name

            # Pass 2: emit rows in PDF order
            for pair_idx, (tr, ptip_match, ptip_match_orig) in enumerate(matched_pairs):
                is_first_tax = (not inv_has_ptip) and (not first_talent_seen)
                is_first_row = not first_talent_seen
                first_talent_seen = True
                eff_scenario = 'A' if not inv_has_ptip else scenario
                loan_out_company = loan_out_idx.get(pair_idx, '')

                row = _build_row(
                    item_no=item_no,
                    ptip=ptip_match,
                    pdf_talent=tr,
                    pdf_invoice=inv_data,
                    received_invoice=True,
                    is_duplicate=False,
                    first_tax_row=is_first_tax,
                    scenario=eff_scenario,
                    ptip_row_no=(ptip_match_orig + 1) if ptip_match_orig is not None else None,
                    payment_entity='The Team Companies',
                    pah_from_pdf=True,
                    workbook_type=workbook_type,
                    first_row_of_invoice=is_first_row,
                    loan_out=bool(loan_out_company),
                    loan_out_company=loan_out_company,
                    talent_name_override=tr['name'] if loan_out_company else '',
                )
                if ptip_match is None:
                    reason = llm_reasons.get(tr['name'], '')
                    if reason:
                        llm_note = f"[LLM no match] {reason}"
                        row['notes'] = f"{row['notes']}; {llm_note}" if row['notes'] else llm_note
                workbook_rows.append(row)
                item_no += 1

            # PTIP rows for this invoice not matched to any PDF talent row
            for local_i, (pr, orig_i) in enumerate(inv_ptip_pairs):
                if local_i in ptip_consumed:
                    continue
                row = _build_row(
                    item_no=item_no,
                    ptip=pr,
                    pdf_talent=None,
                    pdf_invoice=inv_data,
                    received_invoice=False,
                    is_duplicate=False,
                    first_tax_row=False,
                    scenario=scenario,
                    ptip_row_no=orig_i + 1,
                    payment_entity='The Team Companies',
                    pah_from_pdf=True,
                    workbook_type=workbook_type,
                )
                workbook_rows.append(row)
                item_no += 1

        else:
            # PTIP-only invoice (no PDF received)
            for pr, orig_i in inv_ptip_pairs:
                row = _build_row(
                    item_no=item_no,
                    ptip=pr,
                    pdf_talent=None,
                    pdf_invoice=None,
                    received_invoice=False,
                    is_duplicate=False,
                    first_tax_row=False,
                    scenario=scenario,
                    ptip_row_no=orig_i + 1,
                    payment_entity='The Team Companies',
                    pah_from_pdf=True,
                    workbook_type=workbook_type,
                )
                workbook_rows.append(row)
                item_no += 1

    # ── Step 7: Note PDF invoices absent from PTIP ────────────────────────────
    if has_ptip and has_pdf:
        ptip_inv_set = set(ptip_by_invoice.keys())
        for inv_no in sorted(received_invoice_nos):
            if inv_no not in ptip_inv_set:
                issues.append(
                    f"Invoice {inv_no}: PDF received, not in PTIP "
                    "— rows included with 'Included in PTIP report' = NO"
                )

    # ── Step 8: Build Sorted PTIP ─────────────────────────────────────────────
    ptip_excel_b64 = None
    if has_ptip and raw_headers:
        try:
            ptip_excel_bytes = build_teams_sorted_ptip_xlsx(
                ptip_rows_raw,
                raw_headers,
                received_invoice_nos,
                pdf_order_within_invoice,
            )
            ptip_excel_b64 = base64.b64encode(ptip_excel_bytes).decode()
        except Exception as e:
            issues.append(f"Sorted PTIP build error: {e}")

    # ── Step 9: Format talent names (title case, First Last order) ───────────
    for _r in workbook_rows:
        _r['talent_name'] = _fmt_teams_talent_name(_r['talent_name'])

    # ── Step 10: AICP classification (GA only) ───────────────────────────────
    if _is_ga_workbook(workbook_type):
        _classify_aicp_codes_talent(workbook_rows, openai_key)

    # ── Step 11: Summary ──────────────────────────────────────────────────────
    ptip_inv_nos        = set(ptip_by_invoice.keys())
    invoices_with_pdf   = sorted(received_invoice_nos & ptip_inv_nos) if has_ptip else sorted(received_invoice_nos)
    invoices_ptip_only  = sorted(ptip_inv_nos - received_invoice_nos) if has_ptip else []

    return {
        'rows': workbook_rows,
        'ptip_excel_b64': ptip_excel_b64,
        'summary': {
            'total_rows':         len(workbook_rows),
            'invoices_with_pdf':  invoices_with_pdf,
            'invoices_ptip_only': invoices_ptip_only,
            'duplicates_found':   0,
            'issues':             issues,
        },
    }


# ── Highland Talent invoice PDF parsing ─────────────────────────────────────
# Highland has no PTIP-style file at all -- instead a "Payroll Report" xlsx
# consolidates a PID's pay ACROSS every invoice in the job into one row
# (unlike ER/Teams, which key PTIP rows to a single invoice). Every performer
# and agency has a unique 6-char alphanumeric PID; there are no SSNs. Dollar
# figures on the report are authoritative and are never re-derived from the
# PDFs -- the PDFs exist here only to (a) support the PDF-only fallback when
# no report was provided, and (b) cross-check that every PID on a PDF is
# accounted for in the report and vice versa.

_HIGHLAND_PID_RE    = re.compile(r'^[A-Z0-9]{6}$')
_HIGHLAND_AMOUNT_RE = re.compile(r'^[\d,]+\.\d{2}$')
# The category code varies more than expected (seen: P, E, AGT, EXD, FE) --
# match any 1-4 letter code rather than a fixed enum, and rely on ON/OFF
# (consistently performer vs. agent across every example seen) to tell them
# apart rather than the category text itself.
_HIGHLAND_ROW_RE    = re.compile(
    r'^([A-Z0-9]{6})\s+(.+?)\s+(ON|OFF)\s+([A-Z]{1,4})\b(.*)$'
)
_HIGHLAND_CONT_RE   = re.compile(r'^([A-Z])\s+([\d,]+\.\d{2})\*?\s*$')
_HIGHLAND_TITLES    = {'P': 'Principal', 'E': 'Extra'}

# Confirmed by cross-referencing invoice line items against the payroll
# report's own Taxable Wages column across three real jobs: E (fitting) and
# N (night premium) are taxable; A (wardrobe), D (travel), R (reimburse), and
# B (agent fee) are not. An uncoded/base amount is always taxable.
_HIGHLAND_REIMB_CODES = frozenset({'A', 'D', 'R', 'B'})


def _parse_highland_amount_run(rest: str) -> tuple[list[tuple[str | None, float]], float]:
    """Parse the trailing '[LETTER] amount' tokens on a Highland row's own
    line. The LAST amount is always the row's own displayed running total,
    never a fee component -- everything before it is a real line item.

    Some invoices print a per-unit rate field (e.g. "19.25") ahead of the
    state code that reads exactly like a dollar amount but isn't a fee --
    real fee/total figures always come after the state code, so only scan
    for amounts in the text following it."""
    # Some invoices glue the state code directly onto the next field with no
    # space (e.g. "GA22J"), so a real \b word-boundary won't find it -- match
    # on a token start (not preceded by a non-space char) instead.
    state_m = None
    for state_m in re.finditer(r'(?<!\S)[A-Z]{2}(?=\d|\s|$)', rest):
        pass  # keep the LAST match
    scan_from = state_m.end() if state_m else 0
    pairs = re.findall(r'([A-Z])?\s*([\d,]+\.\d{2})\*?', rest[scan_from:])
    parsed = [(code or None, _to_float(amt)) for code, amt in pairs]
    if not parsed:
        return [], 0.0
    *components, (last_code, last_amt) = parsed
    # The trailing total is occasionally itself code-tagged (e.g. a lone 'B'
    # commission row) -- only treat it as a bare total when nothing else on
    # the line already claims that meaning.
    if last_code is not None and not components:
        components = [(last_code, last_amt)]
    return components, last_amt


def parse_highland_invoice_pdf(pdf_bytes: bytes) -> dict | None:
    """Parse a Highland Talent invoice PDF. Returns None if not recognized."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = '\n'.join(pg.extract_text() or '' for pg in pdf.pages)
    except Exception:
        return None

    if 'PAYROLL TAXES:' not in text or 'SERVICE CHARGE:' not in text:
        return None

    inv_no = ''
    invoice_date = ''
    # Date year is 4-digit on most invoices, but the older template uses a
    # 2-digit year (e.g. "08/29/24") -- accept both.
    m = re.search(r'USD STANDARD IV\s+(\d{1,2}/\d{1,2}/\d{2,4})\s+(\d+)', text)
    if m:
        invoice_date, inv_no = m.group(1), m.group(2)
        if len(invoice_date.split('/')[-1]) == 2:
            invoice_date = _expand_date_yy(invoice_date)

    work_state = 'GA'
    m = re.search(r'\b([A-Z][A-Za-z]+),\s+([A-Z]{2})\b', text)
    if m:
        work_state = m.group(2)

    m_taxes  = re.search(r'PAYROLL TAXES:\s*([\d,]+\.\d{2})', text)
    m_svc    = re.search(r'SERVICE CHARGE:\s*([\d,]+\.\d{2})', text)
    payroll_taxes  = _to_float(m_taxes.group(1) if m_taxes else 0)
    service_charge = _to_float(m_svc.group(1)   if m_svc   else 0)

    rows: list[dict] = []
    current: dict | None = None

    def _finalize(row):
        if row is None:
            return
        comp_sum = round(sum(a for _, a in row['_components']), 2)
        wages    = round(sum(a for c, a in row['_components'] if c not in _HIGHLAND_REIMB_CODES), 2)
        misc_pmt = round(comp_sum - wages, 2)
        row['wages']    = wages
        row['misc_pmt'] = misc_pmt
        del row['_components']
        rows.append(row)

    for raw_line in text.split('\n'):
        line = raw_line.strip()
        if not line or line == 'MINOR':
            continue

        m = _HIGHLAND_ROW_RE.match(line)
        if m:
            _finalize(current)
            pid, name, onoff, cat, rest = m.groups()
            is_agent = (onoff == 'OFF')
            name = re.sub(r'\s+\d{1,3}%.*$', '', name).strip()
            components, row_total = _parse_highland_amount_run(rest)
            current = {
                'pid':        pid,
                'name':       name,
                'is_agent':   is_agent,
                'title':      'Agency fee' if is_agent else _HIGHLAND_TITLES.get(cat, cat),
                'row_total':  row_total,
                '_components': components,
            }
            continue

        cm = _HIGHLAND_CONT_RE.match(line)
        if cm and current is not None:
            current['_components'].append((cm.group(1), _to_float(cm.group(2))))
            continue
        # Anything else (headers, legend text, MINOR, control numbers, footer
        # labels, page breaks) is ignored without disturbing `current` -- a
        # continuation line can legitimately follow a page-break header block.

    _finalize(current)

    if not rows:
        return None

    gross_wages = round(sum(r['row_total'] for r in rows), 2)

    return {
        'invoice_no':      inv_no,
        'invoice_date':    invoice_date,
        'work_state':      work_state,
        'gross_wages':     gross_wages,
        'payroll_taxes':   payroll_taxes,
        'service_charge':  service_charge,
        'invoice_total':   round(gross_wages + payroll_taxes + service_charge, 2),
        'talent_rows':     rows,
    }


# ── Highland Payroll Report parsing ─────────────────────────────────────────

def _normalize_highland_col(v) -> str:
    if v is None:
        return ''
    return re.sub(r'\s*\n\s*', ' ', str(v)).strip()


_HIGHLAND_REPORT_REQUIRED_COLS = ('PID', 'Taxable Wages', 'Invoice Total')


def _find_highland_report_sheet(wb):
    """Highland's report workbook is inconsistent -- sometimes one sheet,
    sometimes several (a GA-tax-only view, an audit/QA copy, a sheet named
    after a date). Find whichever sheet actually has the full financial
    columns, preferring a name that doesn't look like a QA/audit copy."""
    candidates = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_vals in ws.iter_rows(max_row=10, values_only=True):
            norms = [_normalize_highland_col(v) for v in row_vals]
            if all(col in norms for col in _HIGHLAND_REPORT_REQUIRED_COLS):
                candidates.append((sheet_name, ws, norms))
                break
    if not candidates:
        return None, []
    for sheet_name, ws, norms in candidates:
        if 'audit' not in sheet_name.lower():
            return ws, norms
    return candidates[0][1], candidates[0][2]


def parse_highland_report_xlsx(xlsx_bytes_list: list[bytes]) -> tuple[list[dict], list[str]]:
    """Read one or more Highland Payroll Report Excel files. Returns
    (rows, issues). Each row is one PID, aggregated across however many
    invoices the report itself grouped together for that PID."""
    rows:   list[dict] = []
    issues: list[str]  = []

    for file_i, xlsx_bytes in enumerate(xlsx_bytes_list):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            ws, headers = _find_highland_report_sheet(wb)
            if ws is None:
                issues.append(f"Payroll report {file_i + 1}: could not find a sheet with PID/Taxable Wages/Invoice Total columns")
                continue

            hdr_row_idx = None
            for r_idx, row_vals in enumerate(ws.iter_rows(max_row=10, values_only=True)):
                norms = [_normalize_highland_col(v) for v in row_vals]
                if all(col in norms for col in _HIGHLAND_REPORT_REQUIRED_COLS):
                    hdr_row_idx = r_idx
                    break
            if hdr_row_idx is None:
                continue

            file_headers = [_normalize_highland_col(v) for v in list(ws.iter_rows(
                min_row=hdr_row_idx + 1, max_row=hdr_row_idx + 1, values_only=True))[0]]

            for row_vals in ws.iter_rows(min_row=hdr_row_idx + 2, values_only=True):
                row = dict(zip(file_headers, row_vals))
                pid = str(row.get('PID', '') or '').strip()
                if not _HIGHLAND_PID_RE.match(pid):
                    continue  # skips totals rows, stray dates, blank rows

                name = str(row.get('Talent/Agency Name', '') or row.get('Talent Name', '') or '').strip()
                inv_no_raw = str(row.get('Invoice #', '') or '').strip()

                rows.append({
                    'pid':             pid,
                    'name':            name,
                    'loanout':         str(row.get('Loan-Out Company', '') or '').strip(),
                    'work_state':      str(row.get('Work State', '') or '').strip(),
                    'address':         str(row.get('Address', '') or '').strip(),
                    'city':            str(row.get('City', '') or '').strip(),
                    'state':           str(row.get('State', '') or '').strip(),
                    'zip':             str(row.get('Zip', '') or '').strip(),
                    'invoice_no_raw':  inv_no_raw,
                    'gross_payment':   _to_float(row.get('Gross Payment')),
                    'expense_reimb':   _to_float(row.get('Expense Reimb')),
                    'taxable_wages':   _to_float(row.get('Taxable Wages')),
                    'pension_health':  _to_float(row.get('Pension & Health')),
                    'employer_taxes':  _to_float(row.get("Employer's Payroll Taxes")),
                    'workers_comp':    _to_float(row.get("Workers' Comp Insurance")),
                    'service_charge':  _to_float(row.get('Service Charge')),
                    'invoice_total':   _to_float(row.get('Invoice Total')),
                })
        except Exception as e:
            issues.append(f"Payroll report {file_i + 1}: parse error — {e}")

    return rows, issues


# ── Highland row assembly ────────────────────────────────────────────────────

def _build_highland_row(
    *,
    item_no: int,
    report_row: dict | None,
    pdf_row: dict | None,          # PDF-only fallback: one raw invoice-line dict
    pdf_invoice: dict | None,      # the invoice this pdf_row came from (PDF-only mode)
    invoice_no: str,
    received_invoice: bool,
    first_row_of_invoice: bool,    # PDF-only mode: absorb this invoice's footer totals
    notes: str = '',
) -> dict:
    """Assemble one workbook row, mirroring _build_row()'s output shape so
    main.py's response contract is identical regardless of payroll company."""

    if report_row:
        name       = report_row['name']
        is_agent   = _is_company_name(name)
        title      = 'Agency fee' if is_agent else (pdf_row.get('title', '') if pdf_row else '')
        work_state = report_row['work_state'] or 'GA'
        home_addr  = report_row['address']
        city       = report_row['city']
        state      = report_row['state']
        zip_code   = report_row['zip']
        wages      = report_row['taxable_wages']
        misc_pmt   = report_row['expense_reimb']
        sag        = report_row['pension_health']
        er_tax     = report_row['employer_taxes']
        wc         = report_row['workers_comp']
        handling   = report_row['service_charge']
        total      = report_row['invoice_total']
        on_ptip    = True
        ptip_amount = total
        ssn_fein   = report_row['pid']
        loan_out_company = report_row['loanout']
    else:
        name       = pdf_row['name']
        is_agent   = pdf_row['is_agent']
        title      = pdf_row['title']
        work_state = pdf_invoice['work_state'] if pdf_invoice else 'GA'
        home_addr = city = state = zip_code = ''
        wages    = pdf_row['wages']
        misc_pmt = pdf_row['misc_pmt']
        sag = er_tax = wc = handling = 0.0
        if first_row_of_invoice and not is_agent and pdf_invoice:
            er_tax   = pdf_invoice['payroll_taxes']
            handling = pdf_invoice['service_charge']
        total    = round(wages + misc_pmt + sag + er_tax + wc + handling, 2)
        on_ptip  = False
        ptip_amount = None
        ssn_fein = pdf_row['pid']
        loan_out_company = ''  # not on the payroll report -- can't be identified from the PDF alone

    if loan_out_company:
        loan_out_note = f"Loan-out company: {loan_out_company}"
        notes = f"{notes}; {loan_out_note}" if notes else loan_out_note

    if state and work_state and state.upper() != work_state.upper():
        qualify = 'NO-OOS'
    else:
        qualify = ''

    return {
        'item_no':        item_no,
        'qualify':        qualify,
        'on_ptip':        on_ptip,
        'ptip_amount':    ptip_amount,
        'work_state':     work_state,
        'talent_name':    name,
        'loan_out':       'YES' if loan_out_company else 'NO',
        'loan_out_company': loan_out_company,
        'title':          title,
        'work_days':      None,
        'work_dates':     '',
        'invoice_no':     invoice_no,
        'invoice_date':   pdf_invoice.get('invoice_date', '') if pdf_invoice else '',
        'wages':          round(wages, 2),
        'misc_pymt':      round(misc_pmt, 2),
        'er_tax':         round(er_tax, 2),
        'wc':             round(wc, 2),
        'handling':       round(handling, 2),
        'sag':            round(sag, 2),
        'signatory_fee':  0.0,
        'other_fees':     0.0,
        'state_tax_withheld':        0.0,
        'local_tax_withheld':        0.0,
        'state_disability_withheld': 0.0,
        'total':          total,
        'check_number':   '',
        'received_invoice': received_invoice,
        'payment_entity': 'Highland Talent Payments, Inc',
        'type':           'Session',
        'home_address':   home_addr,
        'city':           city,
        'state':          state,
        'zip':            zip_code,
        'ssn_fein':          ssn_fein,
        'commercial_id':     '',
        'commercial_title':  '',
        'notes':             notes,
        'ptip_row_no':    None,
        'is_duplicate':   False,
    }


def extract_highland_talent(
    pdf_files:         list[tuple[str, bytes]],
    report_bytes_list:  list[bytes] | None,
    project_title:      str,
    workbook_type:      str,
    openai_key:         str = '',
) -> dict:
    """Run the full Highland Talent extraction. Returns the same shape as
    extract_talent()/extract_teams_talent()."""
    issues: list[str] = []

    # ── Step 1: Parse every PDF, regardless of whether a report exists -- we
    # always need per-PID/per-invoice PDF data for the completeness check and
    # the wage-mismatch note, and it's the only source at all in PDF-only mode.
    pdf_invoices: dict[str, dict] = {}
    for filename, data in pdf_files:
        result = parse_highland_invoice_pdf(data)
        if result is None:
            issues.append(f"{filename}: not recognized as a Highland Talent invoice")
            continue
        inv_no = result['invoice_no']
        if not inv_no:
            issues.append(f"{filename}: could not extract invoice number")
            continue
        if inv_no in pdf_invoices:
            issues.append(f"{filename}: duplicate invoice number {inv_no} — skipped")
            continue
        pdf_invoices[inv_no] = result

    received_invoice_nos = set(pdf_invoices.keys())

    # PID -> [(invoice_no, row_total), ...] across every received PDF, in the
    # order PDFs were processed -- used both for the completeness check and
    # as the invoice-number fallback when the report has no Invoice # column.
    pid_pdf_hits: dict[str, list[tuple[str, float]]] = defaultdict(list)
    pid_pdf_row:  dict[str, dict] = {}  # last-seen raw PDF row per PID, for PDF-only mode
    for inv_no, inv in pdf_invoices.items():
        for r in inv['talent_rows']:
            pid_pdf_hits[r['pid']].append((inv_no, r['row_total']))
            pid_pdf_row[r['pid']] = r

    # ── Step 2: Parse the payroll report, if provided ────────────────────────
    report_rows: list[dict] = []
    if report_bytes_list:
        report_rows, report_issues = parse_highland_report_xlsx(report_bytes_list)
        issues.extend(report_issues)

    workbook_rows: list[dict] = []
    item_no = 1

    if report_rows:
        # ── Report-driven mode: one row per PID, at whatever granularity the
        # report itself used. Never re-derive wages/taxes from the PDFs.
        report_pids = set()
        for rr in report_rows:
            report_pids.add(rr['pid'])
            hits = pid_pdf_hits.get(rr['pid'], [])
            pdf_sum = round(sum(total for _, total in hits), 2)
            received = bool(hits)

            if rr['invoice_no_raw']:
                # The report's own field is sometimes newline-separated
                # rather than comma-separated -- normalize either way.
                invoice_no = ', '.join(
                    tok.strip() for tok in re.split(r'[,\n]+', rr['invoice_no_raw']) if tok.strip()
                )
            else:
                invoice_no = ', '.join(sorted({inv for inv, _ in hits}, key=lambda x: x.zfill(20)))

            notes_parts = []
            if not received:
                notes_parts.append("Not found on any received PDF invoice")
            elif abs(pdf_sum - rr['gross_payment']) > 0:
                notes_parts.append(f"Production Report Gross Payment = ${rr['gross_payment']:,.2f} "
                                    f"but we only have ${pdf_sum:,.2f} in PDF wages")
            if len(hits) > 1:
                # Show how a multi-invoice PID's total actually breaks down,
                # since the workbook only ever shows one combined number.
                breakdown = ', '.join(
                    f"Invoice {inv} (${amt:,.2f})"
                    for inv, amt in sorted(hits, key=lambda h: h[0].zfill(20))
                )
                notes_parts.append(breakdown)
            notes = '; '.join(notes_parts)

            row = _build_highland_row(
                item_no=item_no,
                report_row=rr,
                pdf_row=pid_pdf_row.get(rr['pid']),
                pdf_invoice=None,
                invoice_no=invoice_no,
                received_invoice=received,
                first_row_of_invoice=False,
                notes=notes,
            )
            workbook_rows.append(row)
            item_no += 1

        # Completeness check: any PID seen on a PDF but absent from the report
        for pid, hits in pid_pdf_hits.items():
            if pid not in report_pids:
                inv_list = ', '.join(sorted({inv for inv, _ in hits}))
                name = pid_pdf_row.get(pid, {}).get('name', pid)
                issues.append(f"{name} ({pid}), invoice(s) {inv_list}: on the PDF but not found in the payroll report")

        # Completeness check: any invoice # the report references that we
        # never received a PDF for
        report_inv_nos = set()
        for rr in report_rows:
            for tok in re.split(r'[,\n]+', rr['invoice_no_raw']):
                tok = tok.strip()
                if tok:
                    report_inv_nos.add(tok)
        for inv_no in sorted(report_inv_nos - received_invoice_nos):
            issues.append(f"Invoice {inv_no}: referenced in the payroll report but no matching PDF was received")

    else:
        # ── PDF-only fallback: one row per invoice-line, in PDF order. The
        # invoice's own footer totals (Payroll Taxes, Service Charge) land on
        # the first non-agent row of that invoice, same as ER/Teams.
        for inv_no in sorted(pdf_invoices.keys(), key=lambda x: x.zfill(20)):
            inv = pdf_invoices[inv_no]
            first_seen = False
            for r in inv['talent_rows']:
                is_first = (not first_seen) and not r['is_agent']
                first_seen = first_seen or not r['is_agent']
                row = _build_highland_row(
                    item_no=item_no,
                    report_row=None,
                    pdf_row=r,
                    pdf_invoice=inv,
                    invoice_no=inv_no,
                    received_invoice=True,
                    first_row_of_invoice=is_first,
                )
                workbook_rows.append(row)
                item_no += 1

    if not report_rows and not pdf_invoices:
        return {
            'rows': [],
            'ptip_excel_b64': None,
            'summary': {
                'total_rows': 0,
                'invoices_with_pdf': [],
                'invoices_ptip_only': [],
                'duplicates_found': 0,
                'issues': issues + ['No PDFs or payroll report provided.'],
            },
        }

    if _is_ga_workbook(workbook_type):
        _classify_aicp_codes_talent(workbook_rows, openai_key)

    return {
        'rows': workbook_rows,
        'ptip_excel_b64': None,
        'summary': {
            'total_rows':         len(workbook_rows),
            'invoices_with_pdf':  sorted(received_invoice_nos),
            'invoices_ptip_only': [],
            'duplicates_found':   0,
            'issues':             issues,
        },
    }
