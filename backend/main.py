"""
TPC Extraction Service
──────────────────────
Two extractors the browser wizard POSTs documents to:

  • /extract-invoices          — GPT-4o vision + normalization for freelance invoices
  • /extract-payroll           — fringe report parser for Wrapbook and CAPS PDFs
                                 (auto-detects company; falls back to GPT-4o vision
                                  for image-only Wrapbook fringe PDFs)
  • /extract-billings          — GPT-4o vision for Agency, ProdCo, and Sub-ProdCo
                                 billing invoices; writes to the Billings tab
  • /extract-agency-subvendors — GPT-4o vision for sub-vendor invoices billed to the
                                 ad agency; writes to the Agency Sub-Vendors tab

It never builds the workbook — the wizard assembles the final .xlsx client-side.

Endpoints
  GET  /health                    → {"ok": true, "has_key": bool}
  POST /extract-invoices          → multipart: files[]=<pdf/png/jpg>, prodco_names="A, B"
                                    returns {"invoices": [...], "issues": [...]}
  POST /extract-payroll           → multipart: files[]=<pdf>
                                    returns {"rows": [...], "issues": [...]}
  POST /extract-billings          → multipart: files[]=<pdf>, vendor_type, vendor_name,
                                    vendor_address, vendor_city, vendor_state, vendor_zip,
                                    prodco_names, work_state
                                    returns {"rows": [...], "issues": [...]}
  POST /extract-agency-subvendors  → multipart: files[]=<pdf>, agency_name, agency_address
                                     returns {"rows": [...], "issues": [...]}
  POST /extract-prodco-subvendors  → multipart: files[]=<pdf>, prodco_name, prodco_address
                                     returns {"rows": [...], "issues": [...]}
  POST /extract-petty-cash         → multipart: files[]=<pdf>, prodco_name, work_state
                                     returns {"rows": [...], "issues": [...]}
  POST /extract-ga-ap              → multipart: files[]=<pdf> (one per call), prodco_name,
                                     prodco_address, agency_name, work_state, payer_entities
                                     (JSON array of {role,name,address})
                                     returns {"rows": [...], "issues": [...], "files": [...]}
  POST /extract-ga-petty-cash      → multipart: files[]=<pdf>, prodco_name, work_state
                                     reuses the IL petty cash engine (chunking, envelope-total
                                     reconciliation); GA-specific prompt/fields (FF1/FF2/AICP,
                                     dual proof-of-payment). Never auto-disqualifies a line --
                                     flags possible-DNQ candidates via notes only.
                                     returns {"rows": [...], "issues": [...], "files": [...]}
  POST /match-ap-positions         → JSON: {ap_names: [...], crew: [{name,positions[],dates[]}]}
                                     (crew comes from /extract-call-sheet). GPT matches each
                                     ap_name against the crew list and returns its position.
                                     returns {"mapping": {"First Last": "(Gaffer)"}, "issues": [...]}
  POST /build-ga-workbook          → multipart: files[]=<pdf>, template=<xlsx>, prodco_name,
                                     prodco_address, agency_name, work_state, payer_entities,
                                     project_title
                                     returns populated GA State Submission .xlsx download

Environment variables
  OPENAI_API_KEY     (required for invoices + image-based fringe) your OpenAI key
  APP_SHARED_SECRET  (optional) if set, callers must send header X-App-Secret
  ALLOWED_ORIGINS    (optional) comma-separated CORS origins; default "*"
"""

import os
import re
import io
import json
import time
import random
import base64
import asyncio
import functools

import openpyxl
from datetime import datetime as _dt

import fitz  # PyMuPDF
import pdfplumber
from openai import OpenAI, RateLimitError
from fastapi import FastAPI, UploadFile, File, Form, Header, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from parsers.base import FRINGE_FIELDS
from parsers.wrapbook.fringe_001 import enrich_from_register
from parsers.ai_fringe import extract_unknown, make_exec_parser
from parsers import registry
from notify import send_parser_alert, send_run_summary, ALERT_EMAIL
from talent_extractor import extract_talent, extract_teams_talent

# ── Config ──────────────────────────────────────────────────────────────────

OPENAI_API_KEY    = os.environ.get("OPENAI_API_KEY", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
APP_SHARED_SECRET = os.environ.get("APP_SHARED_SECRET", "")
ALLOWED_ORIGINS   = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "*").split(",") if o.strip()]

_PROMPT_PATH                    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "invoice_extraction_prompt.txt")
_CREW_FREELANCE_PROMPT_PATH     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "crew_freelance_prompt.txt")
_TALENT_FREELANCE_PROMPT_PATH   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "talent_freelance_prompt.txt")
_HOURS_LETTER_PROMPT_PATH       = os.path.join(os.path.dirname(os.path.abspath(__file__)), "hours_letter_prompt.txt")
_BILLING_PROMPT_PATH            = os.path.join(os.path.dirname(os.path.abspath(__file__)), "billing_extraction_prompt.txt")
_AGENCY_SUBVENDORS_PROMPT_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "agency_subvendors_extraction_prompt.txt")
_PRODCO_SUBVENDORS_PROMPT_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "prodco_subvendors_extraction_prompt.txt")
_AGENCY_HOURS_PROMPT_PATH       = os.path.join(os.path.dirname(os.path.abspath(__file__)), "agency_hours_prompt.txt")
_RESIDENCY_DOCS_PROMPT_PATH     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "residency_docs_prompt.txt")
_DIVERSITY_FORM_PROMPT_PATH     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "diversity_form_prompt.txt")
_CALL_SHEET_PROMPT_PATH         = os.path.join(os.path.dirname(os.path.abspath(__file__)), "call_sheet_prompt.txt")
_PETTY_CASH_PROMPT_PATH         = os.path.join(os.path.dirname(os.path.abspath(__file__)), "petty_cash_extraction_prompt.txt")
_GA_AP_PROMPT_PATH              = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ga_ap_extraction_prompt.txt")
_GA_PETTY_CASH_PROMPT_PATH      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ga_petty_cash_extraction_prompt.txt")

app = FastAPI(title="TPC Extraction Service")
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS or ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def _client():
    if not OPENAI_API_KEY:
        raise HTTPException(500, "Server is missing OPENAI_API_KEY.")
    return OpenAI(api_key=OPENAI_API_KEY)


def _anthropic_client():
    if not ANTHROPIC_API_KEY:
        raise HTTPException(500, "Server is missing ANTHROPIC_API_KEY.")
    import anthropic as ant
    return ant.Anthropic(api_key=ANTHROPIC_API_KEY)


def _load_prompt(prodco_names):
    with open(_PROMPT_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    return template.replace("{prodco_names}", ", ".join(prodco_names))


def _load_crew_freelance_prompt(prodco_names: list, prodco_addresses: list | None = None):
    with open(_CREW_FREELANCE_PROMPT_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    addrs = prodco_addresses or []
    if prodco_names:
        entries = []
        for i, name in enumerate(prodco_names):
            addr = addrs[i].strip() if i < len(addrs) else ""
            entries.append(f"{name} ({addr})" if addr else name)
        label = ", ".join(entries)
    else:
        label = "the production company"
    return template.replace("{prodco_names}", label)


def _load_talent_freelance_prompt(prodco_names):
    with open(_TALENT_FREELANCE_PROMPT_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    label = ", ".join(prodco_names) if prodco_names else "the production company"
    return template.replace("{prodco_names}", label)


def _load_hours_letter_prompt():
    with open(_HOURS_LETTER_PROMPT_PATH, "r", encoding="utf-8") as f:
        return f.read()


_VENDOR_TYPE_LABELS = {
    "sub_prodco": "Sub-ProdCo",
    "prodco":     "ProdCo",
    "agency":     "Agency",
}


def _load_billing_prompt(vendor_name: str, vendor_type: str, prodco_names: list) -> str:
    with open(_BILLING_PROMPT_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    type_label   = _VENDOR_TYPE_LABELS.get(vendor_type.lower(), vendor_type)
    prodco_label = ", ".join(prodco_names) if prodco_names else "the production company"
    return (
        template
        .replace("{vendor_name}", vendor_name)
        .replace("{vendor_type}", type_label)
        .replace("{prodco_names}", prodco_label)
    )


def _load_residency_docs_prompt() -> str:
    with open(_RESIDENCY_DOCS_PROMPT_PATH, "r", encoding="utf-8") as f:
        return f.read()


def _load_diversity_form_prompt() -> str:
    with open(_DIVERSITY_FORM_PROMPT_PATH, "r", encoding="utf-8") as f:
        return f.read()


def _load_call_sheet_prompt() -> str:
    with open(_CALL_SHEET_PROMPT_PATH, "r", encoding="utf-8") as f:
        return f.read()


def _load_agency_hours_prompt(agency_name: str = "") -> str:
    with open(_AGENCY_HOURS_PROMPT_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    if agency_name.strip():
        context = (
            f"AGENCY: {agency_name.strip()} — this is the company that wrote these letters. "
            f"Use this name for agency_name on each row."
        )
    else:
        context = (
            "Extract the agency name from the letter header or signature block. "
            "Include it as agency_name on each row."
        )
    return template.replace("{agency_context}", context)


def _load_agency_subvendors_prompt(agency_name: str, agency_address: str = "") -> str:
    with open(_AGENCY_SUBVENDORS_PROMPT_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    name_label = agency_name.strip() or "the ad agency"
    addr_label = f" ({agency_address.strip()})" if agency_address.strip() else ""
    return template.replace("{agency_name}", f"{name_label}{addr_label}")


def _load_prodco_subvendors_prompt(
    prodco_name: str,
    prodco_address: str = "",
    sub_prodco_name: str = "",
    sub_prodco_address: str = "",
) -> str:
    with open(_PRODCO_SUBVENDORS_PROMPT_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    name_label = prodco_name.strip() or "the production company"
    addr_label = f" ({prodco_address.strip()})" if prodco_address.strip() else ""
    if sub_prodco_name.strip():
        sub_addr = f" ({sub_prodco_address.strip()})" if sub_prodco_address.strip() else ""
        sub_line = f"SUB-CONTRACTED PRODUCTION COMPANY: {sub_prodco_name.strip()}{sub_addr}\n"
    else:
        sub_line = ""
    return (
        template
        .replace("{prodco_name}", f"{name_label}{addr_label}")
        .replace("{sub_prodco_line}", sub_line)
    )


def _load_petty_cash_prompt(prodco_name: str, work_state: str = "IL") -> str:
    with open(_PETTY_CASH_PROMPT_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    name_label  = prodco_name.strip() or "the production company"
    state_label = work_state.strip().upper() or "IL"
    return (
        template
        .replace("{prodco_name}", name_label)
        .replace("{work_state}", state_label)
    )


def _load_ga_petty_cash_prompt(prodco_name: str, work_state: str = "GA") -> str:
    with open(_GA_PETTY_CASH_PROMPT_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    name_label  = prodco_name.strip() or "the production company"
    state_label = work_state.strip().upper() or "GA"
    return (
        template
        .replace("{prodco_name}", name_label)
        .replace("{work_state}", state_label)
    )


def _load_ga_ap_prompt(payer_entities: list, work_state: str = "GA") -> str:
    with open(_GA_AP_PROMPT_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    lines = []
    for e in payer_entities:
        role = str(e.get("role", "")).upper()
        name = str(e.get("name", "")).strip()
        addr = str(e.get("address", "")).strip()
        if not name:
            continue
        line = f"  {role}: {name}"
        if addr:
            line += f" ({addr})"
        lines.append(line)
    entities_block = "\n".join(lines) if lines else "  (none provided)"
    return template.replace("{payer_entities_block}", entities_block)


# ── PDF / image → page images (base64 PNG) ────────────────────────────────────

def _file_to_images_b64(filename, data, dpi_scale=2.0, max_dim=None, max_pages=None):
    """Render pages of a PDF (or a single image file) to base64 PNGs.

    max_dim: if set, downsamples any page whose rendered width or height
    exceeds this pixel count (preserving aspect ratio).
    max_pages: if set, only renders the first N pages (prevents memory spikes
    from multi-page PDFs that contain extra non-document pages).
    """
    lower = filename.lower()
    if lower.endswith((".png", ".jpg", ".jpeg")):
        doc = fitz.open(stream=data, filetype="png" if lower.endswith(".png") else "jpg")
        data = doc.convert_to_pdf()
        doc.close()
    doc = fitz.open(stream=data, filetype="pdf")
    images = []
    for i, page in enumerate(doc):
        if max_pages and i >= max_pages:
            break
        pix = page.get_pixmap(matrix=fitz.Matrix(dpi_scale, dpi_scale))
        if max_dim and (pix.width > max_dim or pix.height > max_dim):
            factor = min(max_dim / pix.width, max_dim / pix.height)
            pix = page.get_pixmap(matrix=fitz.Matrix(dpi_scale * factor, dpi_scale * factor))
        images.append(base64.b64encode(pix.tobytes("png")).decode())
    doc.close()
    return images


# ── Rate-limit retry helper ──────────────────────────────────────────────────

_RETRY_AFTER_RE = re.compile(r"try again in ([\d.]+)\s*s", re.IGNORECASE)

def _rate_limit_wait_seconds(exc, attempt):
    """Parse the provider's suggested wait time out of a 429 error message
    (e.g. "Please try again in 5.988s"); fall back to exponential backoff
    with jitter if the message doesn't contain one."""
    m = _RETRY_AFTER_RE.search(str(exc))
    if m:
        try:
            return float(m.group(1)) + 0.5   # small buffer past the suggested wait
        except ValueError:
            pass
    return min(30.0, float(2 ** attempt)) + random.uniform(0, 1)


# ── GPT-4o vision call ────────────────────────────────────────────────────────

def _call_gpt(images_b64, system_prompt, client, user_text="Extract all invoices from these document pages.", max_tokens=4096, max_retries=5):
    content = [{"type": "text", "text": user_text}]
    for img in images_b64:
        content.append({
            "type": "image_url",
            "image_url": {"url": f"data:image/png;base64,{img}", "detail": "high"},
        })

    for attempt in range(max_retries + 1):
        try:
            resp = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": content},
                ],
                temperature=0,
                max_tokens=max_tokens,
            )
            break
        except RateLimitError as e:
            if attempt == max_retries:
                raise
            time.sleep(_rate_limit_wait_seconds(e, attempt))

    raw = resp.choices[0].message.content.strip()
    try:
        return json.loads(raw)
    except Exception:
        print(f"[_call_gpt] JSON parse failed. Raw response (first 500 chars): {raw[:500]!r}", flush=True)
        m = re.search(r"\[.*\]", raw, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                return []
    return []


def _call_gpt_text_json(user_prompt, client, max_tokens=2048, max_retries=5):
    """Text-only GPT call (no images) that expects a single JSON object back."""
    for attempt in range(max_retries + 1):
        try:
            resp = client.chat.completions.create(
                model="gpt-4o",
                messages=[{"role": "user", "content": user_prompt}],
                temperature=0,
                max_tokens=max_tokens,
            )
            break
        except RateLimitError as e:
            if attempt == max_retries:
                raise
            time.sleep(_rate_limit_wait_seconds(e, attempt))

    raw = resp.choices[0].message.content.strip()
    try:
        return json.loads(raw)
    except Exception:
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                return {}
    return {}


def _extract_petty_cash_from_file(filename, data, system_prompt, client, user_text=""):
    """Petty cash extractor: lower DPI to reduce input token count, higher
    max_tokens for large JSON responses, and page-chunking for big packets.
    Summary pages are always prepended to every chunk so the LLM retains
    line-number context across the entire packet."""
    images = _file_to_images_b64(filename, data, dpi_scale=1.5)
    if not images:
        return []

    MAX_TOKENS    = 16384
    SUMMARY_PAGES = 2   # first N pages are always summary; safe even if only 1
    CHUNK_SIZE    = 30  # receipt pages per chunk

    summary_imgs = images[:SUMMARY_PAGES]
    receipt_imgs = images[SUMMARY_PAGES:]

    if len(receipt_imgs) <= CHUNK_SIZE:
        return _call_gpt(summary_imgs + receipt_imgs, system_prompt, client,
                         user_text=user_text, max_tokens=MAX_TOKENS)

    # Large packet: chunk receipt pages, prepend summary to each chunk.
    # First chunk: return ALL summary lines (YES for visible receipts, NO for
    # the rest) so no-receipt rows are captured even if they never appear in a
    # later chunk.  Later chunks return only their own visible receipts; the
    # dedup loop below upgrades NO → YES when the receipt eventually shows up.
    results_by_key: dict[tuple, dict] = {}
    for i in range(0, len(receipt_imgs), CHUNK_SIZE):
        chunk = summary_imgs + receipt_imgs[i : i + CHUNK_SIZE]
        if i == 0:
            current_user_text = (
                user_text
                + " NOTE: You are seeing only the first batch of receipt pages. "
                "Return ALL lines from the summary page -- "
                "lines whose receipts appear in these pages (received_invoice=YES) AND "
                "lines whose receipts are NOT in these pages (received_invoice=NO, "
                "use vendor name and amount from the summary). "
                "Do not skip any summary line."
            )
        else:
            current_user_text = (
                user_text
                + " NOTE: You are seeing a later batch of receipt pages. "
                "Return rows ONLY for receipts physically visible on these pages. "
                "Do not re-return rows already covered by earlier batches."
            )
        for row in _call_gpt(chunk, system_prompt, client,
                             user_text=current_user_text, max_tokens=MAX_TOKENS):
            key = (row.get("env_number", 1), row.get("line_number", 0))
            ri  = str(row.get("received_invoice", "NO")).strip().upper()
            existing = results_by_key.get(key)
            # Keep this row if it's the first occurrence, or if it has a real
            # receipt and the existing entry does not.
            if existing is None or (
                ri == "YES"
                and str(existing.get("received_invoice", "NO")).strip().upper() != "YES"
            ):
                results_by_key[key] = row

    return sorted(
        results_by_key.values(),
        key=lambda r: (r.get("env_number", 1), r.get("line_number", 0)),
    )


# ── Petty cash: Claude fallback ───────────────────────────────────────────────
# GPT-4o has been observed to flatly decline a petty cash file for reasons
# that don't reproduce consistently and don't correspond to any real content
# problem (confirmed via a standalone debug script hitting the same prompt
# and images directly) -- the file extracts cleanly once a different model
# processes it. Rather than chase GPT-4o's exact refusal condition, this
# fallback retries any file that comes back empty using Claude instead,
# before giving up on it entirely.

def _call_claude_petty_cash(images_b64, system_prompt, client, user_text, max_tokens=20000, max_retries=5):
    """Streaming Claude vision call for the petty cash fallback path.
    Streaming is required (not optional) once max_tokens is set this high --
    the Anthropic SDK refuses a non-streaming call outright above its own
    long-request threshold. Returns (parsed_rows_or_None, stop_reason)."""
    import anthropic

    content = []
    for img in images_b64:
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": img},
        })
    content.append({"type": "text", "text": user_text})

    for attempt in range(max_retries + 1):
        try:
            with client.messages.stream(
                model="claude-sonnet-5",
                max_tokens=max_tokens,
                system=system_prompt,
                messages=[{"role": "user", "content": content}],
            ) as stream:
                for _ in stream.text_stream:
                    pass
                resp = stream.get_final_message()
            break
        except anthropic.RateLimitError as e:
            if attempt == max_retries:
                raise
            time.sleep(_rate_limit_wait_seconds(e, attempt))

    text_block = next((b for b in resp.content if b.type == "text"), None)
    raw = text_block.text.strip() if text_block else ""

    try:
        return json.loads(raw), resp.stop_reason
    except Exception:
        m = re.search(r"\[.*\]", raw, re.DOTALL)
        if m:
            try:
                return json.loads(m.group()), resp.stop_reason
            except Exception:
                pass
        return _repair_truncated_json(raw), resp.stop_reason


def _extract_petty_cash_from_file_claude(filename, data, system_prompt, client, user_text=""):
    """Claude fallback mirror of _extract_petty_cash_from_file -- same
    chunking (2 summary pages prepended to 30-page receipt chunks), same
    dedup-by-(env_number, line_number) merge, but a larger streaming
    max_tokens budget (20000) than GPT's 16384. Kept below 32000 on purpose:
    a full-size call at that budget can take long enough streaming out that
    it risks the whole request timing out client-side before Claude even
    finishes -- which throws away everything, vs. a lower budget that
    reliably finishes in time even if it occasionally truncates a large
    envelope. If Claude itself runs out of tokens partway through a chunk,
    it's asked to continue from exactly where it left off (up to
    MAX_CONTINUATIONS times) rather than accepting a partial result
    immediately -- only if it's STILL truncated after those retries does the
    last recovered row get a note flagging that later lines may be missing,
    so a reviewer knows to check the source file instead of assuming a
    partial result is the complete one."""
    images = _file_to_images_b64(filename, data, dpi_scale=1.5)
    if not images:
        return []

    MAX_TOKENS        = 20000
    SUMMARY_PAGES     = 2
    CHUNK_SIZE        = 30
    MAX_CONTINUATIONS = 3

    summary_imgs = images[:SUMMARY_PAGES]
    receipt_imgs = images[SUMMARY_PAGES:]

    def run_chunk(chunk_imgs, chunk_text):
        all_rows: list[dict] = []
        current_text = chunk_text
        for attempt in range(MAX_CONTINUATIONS + 1):
            parsed, stop_reason = _call_claude_petty_cash(
                chunk_imgs, system_prompt, client, current_text, max_tokens=MAX_TOKENS,
            )
            if not isinstance(parsed, list):
                parsed = []
            all_rows.extend(parsed)
            if stop_reason != "max_tokens" or not parsed:
                return all_rows
            # Ran out of tokens but produced rows -- send the same images
            # again, telling Claude exactly what it already returned so it
            # continues instead of restarting or repeating itself.
            covered = sorted({
                r.get("line_number", 0) for r in all_rows if isinstance(r, dict)
            })
            current_text = (
                chunk_text
                + f" You already returned {len(all_rows)} line item(s) in a previous response, "
                f"covering line numbers {covered}. Continue from the next line number onward "
                f"for the REMAINING items in this batch of pages -- do not repeat any line "
                f"number you already returned."
            )
        # Exhausted all continuation attempts and the last one was still
        # truncated -- keep what we have, but flag it.
        if all_rows:
            note = ("Claude fallback ran out of tokens after retrying -- later line items on "
                    "this envelope may still be missing, verify against the source PDF.")
            existing = all_rows[-1].get("notes", "")
            all_rows[-1]["notes"] = f"{existing}; {note}" if existing else note
        return all_rows

    if len(receipt_imgs) <= CHUNK_SIZE:
        return run_chunk(summary_imgs + receipt_imgs, user_text)

    results_by_key: dict[tuple, dict] = {}
    for i in range(0, len(receipt_imgs), CHUNK_SIZE):
        chunk = summary_imgs + receipt_imgs[i : i + CHUNK_SIZE]
        if i == 0:
            current_user_text = (
                user_text
                + " NOTE: You are seeing only the first batch of receipt pages. "
                "Return ALL lines from the summary page -- "
                "lines whose receipts appear in these pages (received_invoice=YES) AND "
                "lines whose receipts are NOT in these pages (received_invoice=NO, "
                "use vendor name and amount from the summary). "
                "Do not skip any summary line."
            )
        else:
            current_user_text = (
                user_text
                + " NOTE: You are seeing a later batch of receipt pages. "
                "Return rows ONLY for receipts physically visible on these pages. "
                "Do not re-return rows already covered by earlier batches."
            )
        for row in run_chunk(chunk, current_user_text):
            key = (row.get("env_number", 1), row.get("line_number", 0))
            ri  = str(row.get("received_invoice", "NO")).strip().upper()
            existing = results_by_key.get(key)
            if existing is None or (
                ri == "YES"
                and str(existing.get("received_invoice", "NO")).strip().upper() != "YES"
            ):
                results_by_key[key] = row

    return sorted(
        results_by_key.values(),
        key=lambda r: (r.get("env_number", 1), r.get("line_number", 0)),
    )


def _tag_claude_fallback_rows(raw_list: list) -> None:
    """Mutates raw_list in place: flags every row that came from the Claude
    fallback so a reviewer can tell OpenAI declined this file and Claude
    filled in instead. Worth flagging because OpenAI is the default/cheaper
    choice everywhere except Residency -- a workbook full of these notes is
    itself a signal something is going wrong upstream and worth a look."""
    for row in raw_list:
        if not isinstance(row, dict):
            continue
        existing = str(row.get("notes", "")).strip()
        flag = "Extracted by Claude (GPT declined)"
        row["notes"] = f"{existing}; {flag}" if existing else flag


def _petty_cash_unreadable_file_row(filename: str) -> dict:
    """Synthetic raw row used when a petty cash file comes back completely
    empty from both GPT-4o and the Claude fallback. Ensures the file still
    shows up in the output workbook -- as one loud, obviously-wrong row --
    instead of silently vanishing with 0 rows, so a reviewer is forced to
    notice it and go check the source PDF rather than assuming the file had
    nothing in it."""
    return {
        "custodian_name":   "",
        "vendor_name":      "UNREADABLE FILE - REVIEW MANUALLY",
        "notes":            (f"Could not extract any data from {filename} -- both OpenAI and "
                              "Claude failed on this file. Review the source PDF manually."),
        "received_invoice": "NO",
        "amount":           0,
    }


# ── Claude vision call ───────────────────────────────────────────────────────

def _call_claude(images_b64, system_prompt, client, user_text="Extract data from these document pages.", max_retries=5):
    import anthropic

    content = []
    for img in images_b64:
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": img},
        })
    content.append({"type": "text", "text": user_text})

    for attempt in range(max_retries + 1):
        try:
            resp = client.messages.create(
                model="claude-sonnet-5",
                max_tokens=2000,
                system=system_prompt,
                messages=[{"role": "user", "content": content}],
            )
            break
        except anthropic.RateLimitError as e:
            if attempt == max_retries:
                raise
            time.sleep(_rate_limit_wait_seconds(e, attempt))

    text_block = next((b for b in resp.content if b.type == "text"), None)
    if not text_block:
        return []
    raw = text_block.text.strip()
    try:
        return json.loads(raw)
    except Exception:
        m = re.search(r"\[.*\]", raw, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                return []
    return []


def _extract_from_file_claude(filename, data, system_prompt, client, user_text="Extract data from these document pages.", dpi_scale=2.0, max_dim=None, max_pages=None):
    images = _file_to_images_b64(filename, data, dpi_scale=dpi_scale, max_dim=max_dim, max_pages=max_pages)
    MAX_BYTES = 40 * 1024 * 1024
    batches, cur, cur_size = [], [], 0
    for img in images:
        approx = len(img) * 3 // 4
        if cur and cur_size + approx > MAX_BYTES:
            batches.append(cur); cur, cur_size = [img], approx
        else:
            cur.append(img); cur_size += approx
    if cur:
        batches.append(cur)
    results = []
    for batch in batches:
        results.extend(_call_claude(batch, system_prompt, client, user_text=user_text))
    return results


def _is_handwritten(filename: str, data: bytes, client) -> bool:
    """Render first page only at low res and ask GPT if the form is handwritten."""
    try:
        doc = fitz.open(stream=data, filetype="pdf")
        pix = doc[0].get_pixmap(matrix=fitz.Matrix(1.0, 1.0))
        img_b64 = base64.b64encode(pix.tobytes("png")).decode()
        doc.close()
    except Exception:
        return False
    resp = client.chat.completions.create(
        model="gpt-4o",
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": "Look at the data entry fields on this form (name, address, dates, etc.) — NOT the signature lines. Are those data fields filled in by hand (handwritten) or typed/digital? Reply with exactly one word: 'handwritten' or 'typed'."},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}", "detail": "low"}},
            ],
        }],
        max_tokens=10,
        temperature=0,
    )
    return "handwritten" in resp.choices[0].message.content.strip().lower()


def _extract_from_file(filename, data, system_prompt, client, user_text="Extract all invoices from these document pages.", dpi_scale=2.0):
    images = _file_to_images_b64(filename, data, dpi_scale=dpi_scale)
    MAX_BYTES = 45 * 1024 * 1024
    batches, cur, cur_size = [], [], 0
    for img in images:
        approx = len(img) * 3 // 4
        if cur and cur_size + approx > MAX_BYTES:
            batches.append(cur); cur, cur_size = [img], approx
        else:
            cur.append(img); cur_size += approx
    if cur:
        batches.append(cur)

    invoices = []
    for batch in batches:
        invoices.extend(_call_gpt(batch, system_prompt, client, user_text=user_text))
    return invoices


# ── Normalization (invoice extractor) ─────────────────────────────────────────

_STATE_ABBR = {
    "alabama":"AL","alaska":"AK","arizona":"AZ","arkansas":"AR","california":"CA",
    "colorado":"CO","connecticut":"CT","delaware":"DE","florida":"FL","georgia":"GA",
    "hawaii":"HI","idaho":"ID","illinois":"IL","indiana":"IN","iowa":"IA","kansas":"KS",
    "kentucky":"KY","louisiana":"LA","maine":"ME","maryland":"MD","massachusetts":"MA",
    "michigan":"MI","minnesota":"MN","mississippi":"MS","missouri":"MO","montana":"MT",
    "nebraska":"NE","nevada":"NV","new hampshire":"NH","new jersey":"NJ","new mexico":"NM",
    "new york":"NY","north carolina":"NC","north dakota":"ND","ohio":"OH","oklahoma":"OK",
    "oregon":"OR","pennsylvania":"PA","rhode island":"RI","south carolina":"SC",
    "south dakota":"SD","tennessee":"TN","texas":"TX","utah":"UT","vermont":"VT",
    "virginia":"VA","washington":"WA","west virginia":"WV","wisconsin":"WI",
    "wyoming":"WY","district of columbia":"DC",
}

_STATE_SET = frozenset(_STATE_ABBR.values())

_PYMT_MAP = {
    "check":"Check","cheque":"Check","p-card":"P-Card","pcard":"P-Card","p card":"P-Card",
    "purchasing card":"P-Card","credit card":"Credit Card","credit":"Credit Card",
    "visa":"Credit Card","mastercard":"Credit Card","amex":"Credit Card","cash":"Cash",
    "eft/wire":"EFT/WIRE","eft":"EFT/WIRE","wire":"EFT/WIRE","wire transfer":"EFT/WIRE",
    "ach":"EFT/WIRE","direct deposit":"EFT/WIRE","payroll company":"Payroll Company",
    "payroll":"Payroll Company","internal":"Internal","zero balance":"Zero Balance",
}

_CARD_ABBR = {
    "american express":"AMEX","amex":"AMEX","visa":"VISA",
    "mastercard":"MC","master card":"MC","mc":"MC","discover":"DISC",
}


def normalize_date(val: str) -> str:
    """Normalize dates to MM/DD/YYYY.
    Handles: YYYY-MM-DD, M/D/YY, M/D/YYYY, MM/DD/YY, MM/DD/YYYY."""
    if not val:
        return ""
    s = str(val).strip()
    # ISO format: YYYY-MM-DD
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', s)
    if m:
        return f"{m.group(2)}/{m.group(3)}/{m.group(1)}"
    # M/D/YY or MM/DD/YY — expand two-digit year
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2})$', s)
    if m:
        mo, dy, yr = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        year = f"20{yr}" if int(yr) <= 50 else f"19{yr}"
        return f"{mo}/{dy}/{year}"
    # M/D/YYYY or MM/DD/YYYY — zero-pad if needed
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        return f"{m.group(1).zfill(2)}/{m.group(2).zfill(2)}/{m.group(3)}"
    return s


def normalize_amount(val):
    if not val:
        return 0
    s = str(val).replace("$", "").replace(",", "").strip()
    try:
        return round(float(s), 2)
    except Exception:
        return 0


def clean_state(val):
    if not val:
        return ""
    s = str(val).strip()
    return _STATE_ABBR.get(s.lower(), s.upper()[:2])


def clean_zip(val):
    return str(val).strip()[:5] if val else ""


def normalize_pymt_method(val):
    return _PYMT_MAP.get(str(val).strip().lower(), "") if val else ""


def normalize_pymt_number(method, val):
    if not val:
        return ""
    s = str(val).strip()
    if method == "EFT/WIRE":
        return s if s.lower().startswith("on ") else "On " + s
    if method in ("Credit Card", "P-Card"):
        if "*" in s:
            return s.upper()
        if s.isdigit() and len(s) == 4:
            return s
        lower = s.lower()
        for name, abbr in _CARD_ABBR.items():
            if name in lower:
                digits = re.search(r"\d{4}", s)
                return f"{abbr}*{digits.group()}" if digits else abbr
        return s
    return s


def clean_name(val):
    if not val:
        return ""
    s = str(val).replace(".", "").strip()
    return " ".join(s.split()).title().rstrip(",").strip()


_ADDR_STREET_TYPES = {
    "avenue":"Ave","boulevard":"Blvd","circle":"Cir","court":"Ct","drive":"Dr",
    "highway":"Hwy","lane":"Ln","parkway":"Pkwy","place":"Pl","road":"Rd",
    "street":"St","terrace":"Ter","trail":"Trl",
}
_ADDR_SINGLE_DIRS   = {"north":"N","south":"S","east":"E","west":"W"}
_ADDR_COMPOUND_DIRS = frozenset({"ne","nw","se","sw"})

def clean_address(val):
    if not val:
        return ""
    s = str(val).strip()
    s = re.sub(r",?\s*#\s*\S+", "", s)
    s = re.sub(r",?\s*(Apt|Apartment|Suite|Ste|Unit|Room|Fl|Floor)(\s+\S+)?", "", s, flags=re.IGNORECASE)
    result = []
    for word in s.split():
        lw = word.lower().rstrip(".")
        if lw in _ADDR_COMPOUND_DIRS:
            result.append(lw.upper())
        elif lw in _ADDR_SINGLE_DIRS:
            result.append(_ADDR_SINGLE_DIRS[lw])
        elif lw in _ADDR_STREET_TYPES:
            result.append(_ADDR_STREET_TYPES[lw])
        else:
            result.append(word.capitalize().rstrip("."))
    return " ".join(result).rstrip(",").strip()


def _parse_vendor_address(full_address: str) -> dict:
    """Split 'Street, City, ST Zip' into components working right-to-left."""
    if not full_address:
        return {"address": "", "city": "", "state": "", "zip": ""}
    rest = full_address.strip()

    zip_ = ""
    m = re.search(r'\b(\d{5}(?:-\d{4})?)\s*$', rest)
    if m:
        zip_ = m.group(1)
        rest = rest[:m.start()].strip().rstrip(",").strip()

    state = ""
    m = re.search(r'\b([A-Za-z]{2})\s*$', rest)
    if m and m.group(1).upper() in _STATE_SET:
        state = m.group(1).upper()
        rest = rest[:m.start()].strip().rstrip(",").strip()

    city = ""
    if "," in rest:
        idx  = rest.rfind(",")
        city = rest[idx + 1:].strip()
        rest = rest[:idx].strip()

    return {
        "address": clean_address(rest),
        "city":    clean_name(city),
        "state":   state,
        "zip":     zip_,
    }


def normalize_invoice(inv):
    method = normalize_pymt_method(inv.get("pymt_method", ""))
    return {
        "po_number":      inv.get("po_number", ""),
        "vendor_name":    clean_name(inv.get("vendor_name", "")),
        "description":    clean_name(inv.get("description", "")),
        "address":        clean_address(inv.get("address", "")),
        "city":           clean_name(inv.get("city", "")),
        "state":          clean_state(inv.get("state", "")),
        "zip":            clean_zip(inv.get("zip", "")),
        "invoice_date":   str(inv.get("invoice_date", "")).strip(),
        "invoice_number": inv.get("invoice_number", ""),
        "invoice_amount": normalize_amount(inv.get("invoice_amount", "")),
        "pymt_method":    method,
        "pymt_number":    normalize_pymt_number(method, inv.get("pymt_number", "")),
        "notes":          inv.get("notes", ""),
    }


def normalize_billing_row(
    raw: dict,
    vendor_type: str,
    vendor_name: str,
    vendor_address: str,
    vendor_city: str,
    vendor_state: str,
    vendor_zip: str,
    work_state: str,
    filename: str,
) -> dict:
    # If the wizard sent split fields, use them; otherwise parse the full address string.
    if vendor_city or vendor_state or vendor_zip:
        addr  = clean_address(vendor_address or raw.get("address", ""))
        city  = clean_name(vendor_city or raw.get("city", ""))
        state = clean_state(vendor_state or raw.get("state", ""))
        zip_  = clean_zip(vendor_zip or raw.get("zip", ""))
    else:
        parsed = _parse_vendor_address(vendor_address)
        addr  = parsed["address"] or clean_address(raw.get("address", ""))
        city  = parsed["city"]    or clean_name(raw.get("city", ""))
        state = parsed["state"]   or clean_state(raw.get("state", ""))
        zip_  = parsed["zip"]     or clean_zip(raw.get("zip", ""))

    local     = bool(state) and state.upper() == work_state.upper()
    geo_state = "Local" if local else "OOS"

    proj_fee_raw = raw.get("project_fee")
    proj_fee     = normalize_amount(proj_fee_raw) if proj_fee_raw not in (None, "", 0, 0.0) else None

    return {
        "qualify":         "",
        "vendorName":      vendor_name or clean_name(raw.get("vendor_name", "")),
        "vendorType":      _VENDOR_TYPE_LABELS.get(vendor_type.lower(), vendor_type),
        "jobType":         str(raw.get("job_type", "")).strip(),
        "state":           geo_state,
        "description":     str(raw.get("description", "")).strip(),
        "details":         str(raw.get("details", "")).strip(),
        "invoiceDate":     normalize_date(raw.get("invoice_date", "")),
        "invoiceNo":       str(raw.get("invoice_number", "")).strip(),
        "eligibleTotal":   normalize_amount(raw.get("invoice_amount", 0)),
        "projectFee":      proj_fee,
        "address":         addr,
        "city":            city,
        "vendorState":     state,
        "zip":             zip_,
        "receivedInvoice": "Yes",
        "pop":             "Yes" if raw.get("pop") else None,
        "paymentDate":     normalize_date(raw.get("payment_date", "")),
        "paymentMethod":   str(raw.get("payment_method", "")).strip(),
        "jobNumber":       str(raw.get("job_number", "")).strip(),
        "notes":           str(raw.get("notes", "")).strip(),
        "sourceFile":      filename,
    }


def normalize_agency_subvendor_row(raw: dict, agency_name: str, filename: str) -> dict:
    method = normalize_pymt_method(raw.get("payment_method", ""))

    w9_status = str(raw.get("w9_status", "not_present")).strip().lower()
    w9_date   = normalize_date(raw.get("w9_date", ""))
    if w9_status == "signed_dated" and w9_date:
        w9_value = w9_date
    elif w9_status == "unsigned":
        w9_value = "Unsigned"
    elif w9_status == "undated":
        w9_value = "Not Dated"
    else:
        w9_value = "No"

    return {
        "qualify":          "",
        "vendorName":       clean_name(raw.get("vendor_name", "")),
        "description":      str(raw.get("description", "")).strip(),
        "address":          clean_address(raw.get("address", "")),
        "city":             clean_name(raw.get("city", "")),
        "zip":              clean_zip(raw.get("zip", "")),
        "vendorState":      clean_state(raw.get("state", "")),
        "invoiceDate":      normalize_date(raw.get("invoice_date", "")),
        "invoiceNo":        str(raw.get("invoice_number", "")).strip(),
        "invoiceAmount":    normalize_amount(raw.get("invoice_amount", 0)),
        "receivedInvoice":  "Yes",
        "w9ValidDate":      w9_value,
        "pop":              "Yes" if raw.get("pop") else "No",
        "paymentMethod":    method,
        "paymentNo":        normalize_pymt_number(method, raw.get("payment_number", "")),
        "paymentEntity":    agency_name.strip(),
        "jobNo":            str(raw.get("job_number", "")).strip(),
        "clientBillingNo":  str(raw.get("client_billing_number", "")).strip(),
        "notes":            str(raw.get("notes", "")).strip(),
        "sourceFile":       filename,
    }


def normalize_prodco_subvendor_row(raw: dict, prodco_name: str, filename: str) -> dict:
    method = normalize_pymt_method(raw.get("pymt_method", ""))

    w9_status = str(raw.get("w9_status", "not_present")).strip().lower()
    w9_date   = normalize_date(raw.get("w9_date", ""))
    if w9_status == "signed_dated" and w9_date:
        w9_value = w9_date
    elif w9_status == "unsigned":
        w9_value = "Unsigned"
    elif w9_status == "undated":
        w9_value = "Not Dated"
    else:
        w9_value = "No"

    return {
        "poNumber":         str(raw.get("po_number", "")).strip(),
        "vendorName":       clean_name(raw.get("vendor_name", "")),
        "description":      str(raw.get("description", "")).strip(),
        "address":          clean_address(raw.get("address", "")),
        "city":             clean_name(raw.get("city", "")),
        "vendorState":      clean_state(raw.get("state", "")),
        "zip":              clean_zip(raw.get("zip", "")),
        "invoiceDate":      normalize_date(raw.get("invoice_date", "")),
        "invoiceNo":        str(raw.get("invoice_number", "")).strip(),
        "invoiceAmount":    normalize_amount(raw.get("invoice_amount", 0)),
        "receivedInvoice":  "Yes",
        "w9ValidDate":      w9_value,
        "pop":              "Yes" if raw.get("pop") else "No",
        "paymentMethod":    method,
        "paymentNo":        normalize_pymt_number(method, raw.get("pymt_number", "")),
        "paymentEntity":    prodco_name.strip(),
        "notes":            str(raw.get("notes", "")).strip(),
        "sourceFile":       filename,
    }


# ── Petty cash multi-part-file custodian backfill (shared: IL + GA) ──────────
# A large custodian packet is sometimes split across multiple PDF files by
# page count (e.g. "Alantra, Wilson - Petty Cash (1 of 3).pdf", "(2 of 3)",
# "(3 of 3)"). Only the first part's cover sheet may show a real custodian
# name -- later parts' own cover/continuation pages sometimes have a broken
# or placeholder Employee Name field, which can make the model either return
# nothing for that file, or (worse) latch onto unrelated nearby text (e.g. a
# job code) as a fake name. Since each file is extracted independently with
# no memory of the others, this can't be fixed at extraction time -- it's
# fixed here, after the fact, using the filename itself as a reliable
# grouping signal.

_PETTY_CASH_PACKET_KEY_RE = re.compile(r"^(.*?)\s*-\s*petty\s*cash\b", re.IGNORECASE)


def _petty_cash_packet_key(filename: str) -> str:
    """Extract the '<Name/Dept>' prefix before ' - Petty Cash' in a filename
    (e.g. 'alantra wilson' from 'Alantra, Wilson - Petty Cash (1 of 3).pdf').
    Used only to GROUP multi-part files belonging to the same packet -- never
    to fabricate a last/first split, since filenames aren't reliably in any
    particular name order."""
    base = os.path.splitext(filename)[0]
    m = _PETTY_CASH_PACKET_KEY_RE.match(base)
    key = m.group(1) if m else base
    return re.sub(r"[^a-z0-9]+", " ", key.lower()).strip()


def _custodian_name_trustworthy(name: str, packet_key: str) -> bool:
    """A custodian_name is trustworthy if it shares at least one word with
    the filename's packet key -- catches both blank/placeholder values
    ("0", "N/A") and wrong guesses (e.g. a job code like "Sxmas") that share
    no overlap with the real name/department in the filename."""
    name_words = set(re.findall(r"[a-z0-9]+", (name or "").lower()))
    key_words = set(packet_key.split())
    if not name_words or not key_words:
        return False
    return bool(name_words & key_words)


def _split_custodian_name(name: str) -> tuple[str, str]:
    """Shared LAST/FIRST split used both at normalization time (GA petty
    cash rows carry separate last_name/first_name fields) and again here
    whenever the backfill below patches custodian_name -- so the split
    fields stay in sync with whatever name actually ends up on the row,
    instead of being permanently locked to whatever custodian_name was at
    the moment the row was first normalized."""
    name = name.strip()
    if "," in name:
        last, _, first = name.partition(",")
        return last.strip(), first.strip()
    parts = name.split()
    if len(parts) >= 2:
        return parts[0], " ".join(parts[1:])
    return name, ""


def _backfill_petty_cash_custodian_names(file_records: list[dict]) -> None:
    """Mutates file_records' rows in place. Groups files by the name/dept
    parsed from their filename and, whenever one file in a group has a
    trustworthy custodian_name, backfills it into sibling files whose own
    extracted name is blank or clearly wrong. Must run BEFORE envelope-number
    offset bucketing, since that bucketing is keyed by custodian_name.

    Also re-derives last_name/first_name from the backfilled name on any row
    that has those keys (GA rows only -- IL rows have no separate split
    fields). Those were split out of custodian_name back when the row was
    normalized, which is BEFORE this backfill runs, so they'd otherwise stay
    frozen at whatever they were computed from originally (usually blank)
    even after custodian_name itself gets corrected here.

    Finally, as a last resort, any row whose custodian_name is STILL blank
    after the above (no trustworthy sibling found, or no siblings at all)
    gets the source filename instead -- so the row is always traceable to
    where it came from rather than silently blank with no way to tell which
    file it belongs to. Deliberately does not attempt to split a filename
    into last_name/first_name (that would produce garbage); those stay
    blank, only custodian_name gets the fallback."""
    groups: dict[str, list[dict]] = {}
    for rec in file_records:
        if not rec["file_rows"]:
            continue
        key = _petty_cash_packet_key(rec["filename"])
        groups.setdefault(key, []).append(rec)

    for key, recs in groups.items():
        if len(recs) < 2:
            continue
        canonical = None
        for rec in recs:
            name = rec["file_rows"][0].get("custodian_name", "")
            if _custodian_name_trustworthy(name, key):
                canonical = name
                break
        if not canonical:
            continue
        canonical_last, canonical_first = _split_custodian_name(canonical)
        for rec in recs:
            name = rec["file_rows"][0].get("custodian_name", "")
            if not _custodian_name_trustworthy(name, key):
                for r in rec["file_rows"]:
                    r["custodian_name"] = canonical
                    if "last_name" in r:
                        r["last_name"] = canonical_last
                    if "first_name" in r:
                        r["first_name"] = canonical_first

    for rec in file_records:
        label = os.path.splitext(rec["filename"])[0]
        for r in rec["file_rows"]:
            if not str(r.get("custodian_name", "")).strip():
                r["custodian_name"] = label


def normalize_petty_cash_row(raw: dict, work_state: str, filename: str) -> dict:
    method = normalize_pymt_method(str(raw.get("pymt_method", "")).strip()) or "Cash"

    state_val = clean_state(raw.get("state", ""))
    if not state_val and work_state:
        state_val = clean_state(work_state) or work_state.strip().upper()[:2]

    env_num = raw.get("env_number", 1)
    try:
        env_num = max(1, int(env_num))
    except (TypeError, ValueError):
        env_num = 1

    line_num = raw.get("line_number", 0)
    try:
        line_num = max(0, int(line_num))
    except (TypeError, ValueError):
        line_num = 0

    received = str(raw.get("received_invoice", "YES")).strip().upper()
    if received not in ("YES", "NO"):
        received = "YES"

    return {
        "custodian_name":   clean_name(raw.get("custodian_name", "")),
        "env_number":       env_num,
        "line_number":      line_num,
        "vendor_name":      clean_name(raw.get("vendor_name", "")),
        "receipt_type":     str(raw.get("receipt_type", "Receipt")).strip() or "Receipt",
        "receipt_date":     normalize_date(str(raw.get("receipt_date", "")).strip()),
        "amount":           normalize_amount(raw.get("amount", 0)),
        "description":      str(raw.get("description", "")).strip(),
        "state":            state_val,
        "pymt_method":      method,
        "received_invoice": received,
        "notes":            str(raw.get("notes", "")).strip(),
        "sourceFile":       filename,
    }


def normalize_agency_hours_row(raw: dict, agency_name: str, filename: str) -> dict:
    hours = raw.get("agency_hours")
    try:
        hours = round(float(hours), 2) if hours is not None else None
    except (ValueError, TypeError):
        hours = None

    name = agency_name.strip() or clean_name(raw.get("agency_name", ""))

    return {
        "hoursLetter":        "Hours Letter",
        "invoiceDate":        normalize_date(raw.get("invoice_date", "")),
        "qualify":            "",
        "crewName":           clean_name(raw.get("crew_name", "")),
        "jobDescription":     str(raw.get("job_description", "")).strip(),
        "positionCategory":   str(raw.get("position_category", "")).strip(),
        "address":            clean_address(raw.get("address", "")),
        "city":               clean_name(raw.get("city", "")),
        "zip":                clean_zip(raw.get("zip", "")),
        "state":              clean_state(raw.get("state", "")),
        "agencyHours":        hours,
        "agencyHoursAmount":  normalize_amount(raw.get("agency_hours_amount", 0)),
        "hoursLetterType":    "Agency Hours Letter",
        "datesWorked":        str(raw.get("dates_worked", "")).strip(),
        "agencyName":         name,
        "sourceFile":         filename,
    }


def normalize_retainer_billing_row(raw: dict, agency_name: str, filename: str) -> dict:
    return {
        "invoiceNo":   str(raw.get("invoice_number", "")).strip(),
        "invoiceDate": normalize_date(raw.get("invoice_date", "")),
        "amount":      normalize_amount(raw.get("invoice_amount", 0)),
        "vendorName":  agency_name.strip() or clean_name(raw.get("vendor_name", "")),
        "sourceFile":  filename,
    }


def _i9_notes(raw: dict, handwritten: bool, shoot_date: str) -> str:
    notes = []

    if handwritten:
        notes.append("Handwritten - verify accuracy")

    if str(raw.get("document_type", "")).strip() != "I9":
        return "; ".join(notes)

    if raw.get("employee_signed") is False:
        notes.append("Missing employee signature")

    sig_date = normalize_date(str(raw.get("employee_signature_date") or ""))
    if sig_date and shoot_date:
        try:
            from datetime import datetime
            diff = abs((datetime.strptime(sig_date, "%m/%d/%Y") - datetime.strptime(shoot_date, "%m/%d/%Y")).days)
            if diff > 90:
                notes.append(f"Signature date ({sig_date}) is not close to shoot date ({shoot_date})")
        except ValueError:
            pass

    exp_date = normalize_date(str(raw.get("expiration_date") or ""))
    dob      = normalize_date(str(raw.get("date_of_birth") or ""))
    if exp_date and dob:
        try:
            from datetime import datetime
            exp_dt = datetime.strptime(exp_date, "%m/%d/%Y")
            dob_dt = datetime.strptime(dob,      "%m/%d/%Y")
            if exp_dt.month != dob_dt.month or exp_dt.day != dob_dt.day:
                notes.append(f"Work auth expiration ({exp_date}) month/day does not match birth date ({dob})")
        except ValueError:
            pass

    missing = []
    if raw.get("employer_signed") is False:
        missing.append("employer signature")
    if not str(raw.get("employer_signature_date") or "").strip():
        missing.append("employer date")
    if not str(raw.get("first_day_of_employment") or "").strip():
        missing.append("first day of employment")
    if not str(raw.get("employer_name") or "").strip():
        missing.append("employer name")
    if not str(raw.get("employer_address") or "").strip():
        missing.append("employer address")
    if missing:
        notes.append("Missing employer section: " + ", ".join(missing))

    return "; ".join(notes)


def normalize_diversity_row(raw: dict, filename: str) -> dict:
    sex = str(raw.get("sex", "")).strip().upper()
    diversity = str(raw.get("diversity", "")).strip().upper()
    if sex not in ("MALE", "FEMALE"):
        sex = ""
    if diversity not in ("AA", "ASIAN", "WHITE", "HISP", "NA"):
        diversity = ""
    return {
        "name":       str(raw.get("name", "")).strip().title(),
        "sex":        sex,
        "diversity":  diversity,
        "sourceFile": filename,
    }


def _last_name_key(name: str) -> str:
    last = name.split(",")[0] if "," in name else name
    return re.sub(r"[^a-z]", "", last.lower())


def _merge_residency_diversity(residency_rows: list, diversity_rows: list) -> list:
    div_map: dict = {}
    for d in diversity_rows:
        key = _last_name_key(d["name"])
        div_map.setdefault(key, []).append(d)

    matched_keys: set = set()
    merged = []

    for row in residency_rows:
        key = _last_name_key(row.get("documentName", ""))
        divs = div_map.get(key, [])
        if divs:
            matched_keys.add(key)
            row["sex"]       = divs[0].get("sex", "")
            row["diversity"] = divs[0].get("diversity", "")
            if len(divs) > 1:
                note = "Multiple diversity forms - verify manually"
                existing = row.get("notes", "")
                row["notes"] = f"{existing}; {note}".strip("; ") if existing else note
        else:
            row.setdefault("sex", "")
            row.setdefault("diversity", "")
        merged.append(row)

    for key, divs in div_map.items():
        if key in matched_keys:
            continue
        note = "Multiple diversity forms - verify manually" if len(divs) > 1 else ""
        merged.append({
            "documentName":   divs[0]["name"],
            "documentType":   "",
            "issueDate":      "",
            "expirationDate": "",
            "address":        "",
            "city":           "",
            "zip":            "",
            "state":          "",
            "notes":          note,
            "sex":            divs[0].get("sex", ""),
            "diversity":      divs[0].get("diversity", ""),
            "sourceFile":     divs[0].get("sourceFile", ""),
        })

    return merged


def normalize_residency_row(raw: dict, filename: str, handwritten: bool = False, shoot_date: str = "") -> dict:
    return {
        "documentName":   str(raw.get("document_name", "")).strip().title(),
        "documentType":   str(raw.get("document_type", "")).strip(),
        "issueDate":      normalize_date(raw.get("issue_date", "")),
        "expirationDate": normalize_date(raw.get("expiration_date", "")),
        "address":        clean_address(raw.get("address", "")),
        "city":           clean_name(raw.get("city", "")),
        "zip":            clean_zip(raw.get("zip", "")),
        "state":          clean_state(raw.get("state", "")),
        "notes":          _i9_notes(raw, handwritten, shoot_date),
        "sex":            "",
        "diversity":      "",
        "sourceFile":     filename,
    }


# ── Payroll company detection ─────────────────────────────────────────────────

def _is_wrapbook_register_only(pdf_bytes: bytes) -> bool:
    """Return True if this PDF is a standalone Wrapbook Payroll Register (NIS 007 style).

    A standalone register has 'Payroll Register' pages but no 'Fringe Report' pages.
    These are uploaded alongside fringe PDFs so their IL withholding data can enrich
    the project-level fringe rows (which have no invoice number).
    """
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            has_register = has_fringe = False
            for pg in pdf.pages[:10]:
                text = (pg.extract_text() or "").lower()
                if "payroll register" in text and "xxx-xx-" in text:
                    has_register = True
                if "fringe report" in text or "fringe recap" in text:
                    has_fringe = True
                    break
            return has_register and not has_fringe
    except Exception:
        return False


# ── Routes ────────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"ok": True, "has_key": bool(OPENAI_API_KEY)}


@app.post("/extract-invoices")
async def extract_invoices(
    files: list[UploadFile] = File(...),
    prodco_names: str = Form(""),
    x_app_secret: str = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client = _client()
    names = [n.strip() for n in prodco_names.split(",") if n.strip()]
    system_prompt = _load_prompt(names)

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        async with sem:
            try:
                raw = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_from_file, filename, data, system_prompt, client),
                )
                return filename, raw, None
            except Exception as e:
                return filename, [], str(e)

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    invoices, issues = [], []
    for filename, raw, err in extraction_results:
        if err:
            issues.append(f"{filename}: {err}")
        if not raw:
            invoices.append(normalize_invoice({
                "vendor_name": os.path.splitext(filename)[0],
                "notes": "Invoice could not be read - please review manually",
            }))
            issues.append(f"{filename}: no invoice data extracted")
            continue
        for inv in raw:
            invoices.append(normalize_invoice(inv))

    return {"invoices": invoices, "issues": issues}


async def _run_extract(files, x_app_secret, payroll_hints: str = ""):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    # Read all files up front so we can classify before extracting
    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))
    loaded.sort(key=lambda x: (x[0] or "").lower())

    # Classify each file: fringe PDF vs standalone Wrapbook register
    fringe_files   = []   # (filename, bytes)
    register_files = []   # bytes of standalone Wrapbook register PDFs

    for filename, data in loaded:
        if _is_wrapbook_register_only(data):
            register_files.append(data)
        else:
            fringe_files.append((filename, data))

    rows, issues, file_summaries = [], [], []
    wb_rows: list[dict] = []   # Wrapbook fringe rows that may need register enrichment

    # Track project-level Wrapbook fringe sources (invoiceNo always blank — one fringe
    # report consolidates many invoices). Stored as (filename, row_list) so we can
    # assign "Fringe Report 001 / 002 / …" labels after enrichment is complete.
    wb_project_sources: list[tuple[str, list[dict]]] = []

    # Ordered list of email alerts — one entry per generation event.
    # [{company_name, code, file_count, row_count, is_update}]
    alert_queue: list[dict] = []

    def _latest_alert(company: str) -> dict | None:
        for entry in reversed(alert_queue):
            if entry["company_name"] == company:
                return entry
        return None

    for filename, data in fringe_files:
        # ── Find all matching parsers (static + runtime) sorted by priority ────
        candidates = registry.find_parsers(data)

        extracted: list[dict] = []
        errs:      list[str]  = []
        company:   str | None = None

        # Try each candidate in order; stop on first that returns rows
        for candidate in candidates:
            try:
                extracted, errs = candidate.extract(data, openai_key=OPENAI_API_KEY)
            except Exception as e:
                extracted, errs = [], [f"{candidate.COMPANY} parser error: {e}"]
            if extracted:
                company = candidate.COMPANY
                break

        # ── Wrapbook-specific post-processing ─────────────────────────────────
        if company == "wrapbook":
            wb_rows.extend(extracted)
            if extracted and all(r.get("invoiceNo", "") == "" for r in extracted):
                wb_project_sources.append((filename, extracted))

        # ── Update alert counts for successful runtime-parser hits ─────────────
        if extracted and company:
            entry = _latest_alert(company)
            if entry:
                entry["file_count"] += 1
                entry["row_count"]  += len(extracted)

        # ── Phase 3: all known layout versions failed → generate new variant ──
        if not extracted:
            known_company = candidates[0].COMPANY if candidates else None
            hints = payroll_hints
            if known_company:
                hints = (
                    f"Company: {known_company} — all {len(candidates)} known layout(s) failed, likely a new variant\n"
                    + hints
                )

            ai_rows, ai_errs, ai_company, markers, code = extract_unknown(
                data, OPENAI_API_KEY, hints=hints, anthropic_key=ANTHROPIC_API_KEY
            )
            extracted, errs = ai_rows, ai_errs
            if ai_company:
                exec_fn = make_exec_parser(code)
                if exec_fn:
                    registry.register_parser(ai_company, markers, exec_fn)
                existing = _latest_alert(ai_company)
                if existing:
                    existing["file_count"] += 1
                    existing["row_count"]  += len(extracted)
                else:
                    alert_queue.append({
                        "company_name": ai_company,
                        "code":         code,
                        "file_count":   1,
                        "row_count":    len(extracted),
                        "is_update":    bool(known_company),
                    })
                company = ai_company
            else:
                errs.append(f"Could not identify payroll company in {filename}")
                company = known_company or "unknown"

        for r in extracted:
            r["sourceFile"] = filename

        rows.extend(extracted)
        for e in errs:
            issues.append(f"{filename}: {e}")

        file_summaries.append({
            "filename": filename,
            "company":  company or "unknown",
            "rows":     len(extracted),
            "issues":   errs,
        })

    # Enrich Wrapbook fringe rows with any standalone register PDFs (NIS 007 style).
    # Must happen BEFORE "Fringe Report" label assignment — enrichment joins on invoiceNo == "".
    for register_bytes in register_files:
        enrich_from_register(wb_rows, register_bytes)

    if register_files and not wb_rows:
        issues.append(
            f"{len(register_files)} Wrapbook register file(s) uploaded but no Wrapbook fringe "
            "rows found to enrich. Upload the fringe PDF alongside the register."
        )

    # Assign "Fringe Report" invoice labels for project-level fringe rows.
    # Single source → "Fringe Report"; multiple sources → "Fringe Report 001", "002", …
    if len(wb_project_sources) == 1:
        for r in wb_project_sources[0][1]:
            r["invoiceNo"] = "Fringe Report"
    elif len(wb_project_sources) > 1:
        for idx, (_, src_rows) in enumerate(wb_project_sources, 1):
            label = f"Fringe Report {idx:03d}"
            for r in src_rows:
                r["invoiceNo"] = label

    # Send one email per generation event (initial discovery or retry)
    for info in alert_queue:
        company_name = info["company_name"]
        sent = send_parser_alert(
            company_name,
            info["file_count"],
            info["row_count"],
            info["code"],
            is_update=info.get("is_update", False),
        )
        verb      = "Updated parser" if info.get("is_update") else "New parser"
        alert_msg = (
            f"{verb} for '{company_name}' — "
            f"{info['file_count']} file(s) processed with AI extraction"
        )
        alert_msg += (
            f". Parser code emailed to {ALERT_EMAIL} for review." if sent
            else ". (Email alert failed — check SENDGRID_API_KEY)"
        )
        issues.append(alert_msg)

    return {"rows": rows, "issues": issues, "columns": FRINGE_FIELDS, "files": file_summaries}


@app.post("/extract-fringe")
async def extract_fringe(
    files: list[UploadFile] = File(...),
    payroll_hints: str = Form(default=""),
    x_app_secret: str = Header(default=""),
):
    return await _run_extract(files, x_app_secret, payroll_hints=payroll_hints)


@app.post("/extract-payroll")
async def extract_payroll(
    files: list[UploadFile] = File(...),
    payroll_hints: str = Form(default=""),
    x_app_secret: str = Header(default=""),
):
    return await _run_extract(files, x_app_secret, payroll_hints=payroll_hints)


# ── Crew freelance invoice extractor ─────────────────────────────────────────

_CREW_REQUIRED = {"worker", "invoiceNo", "invoiceDate", "wages"}

def normalize_crew_freelance_row(raw: dict, filename: str) -> dict:
    worker       = clean_name(raw.get("worker", "")) or "[missing information]"
    invoice_no   = str(raw.get("invoiceNo", "")).strip() or "[missing information]"
    invoice_date = str(raw.get("invoiceDate", "")).strip() or "[missing information]"
    wages        = normalize_amount(raw.get("wages", 0))
    if wages == 0:
        wages = "[missing information]"

    method = normalize_pymt_method(raw.get("pymtMethod", ""))

    days = raw.get("daysWorked")
    if isinstance(days, float):
        days = int(days)
    elif not isinstance(days, int):
        days = None

    return {
        "worker":        worker,
        "jobTitle":      clean_name(raw.get("jobTitle", "")),
        "invoiceNo":     invoice_no,
        "invoiceDate":   invoice_date,
        "workDates":     str(raw.get("workDates", "")).strip(),
        "daysWorked":    days,
        "wages":         wages,
        "kitRental":     normalize_amount(raw.get("kitRental", 0)) or None,
        "mileage":       normalize_amount(raw.get("mileage", 0)) or None,
        "reimbursement": normalize_amount(raw.get("reimbursement", 0)) or None,
        "other":         normalize_amount(raw.get("other", 0)) or None,
        "invoiceTotal":  normalize_amount(raw.get("invoiceTotal", 0)) or None,
        "poNo":          str(raw.get("poNo", "")).strip(),
        "pymtMethod":    method,
        "pymtNo":        normalize_pymt_number(method, raw.get("pymtNo", "")),
        "street":        clean_address(raw.get("street", "")),
        "city":          clean_name(raw.get("city", "")),
        "state":         clean_state(raw.get("state", "")),
        "zip":           clean_zip(raw.get("zip", "")),
        "sourceFile":    filename,
    }


@app.post("/extract-crew-freelance")
async def extract_crew_freelance(
    files:            list[UploadFile] = File(...),
    prodco_names:     str              = Form(""),
    prodco_addresses: str              = Form(""),
    x_app_secret:     str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client  = _client()
    names   = [n.strip() for n in prodco_names.split(",") if n.strip()]
    addrs   = [a.strip() for a in prodco_addresses.split(",")]
    system_prompt = _load_crew_freelance_prompt(names, addrs)
    user_text     = "Extract crew freelance invoice data from these document pages."

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                return filename, raw_list, None
            except Exception as e:
                return filename, [], str(e)

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []

    for filename, raw_list, err in extraction_results:
        errs: list[str] = []
        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        if not raw_list:
            errs.append("no crew freelance data extracted — review manually")
            issues.append(f"{filename}: no crew freelance data extracted")
        else:
            for raw in raw_list:
                try:
                    file_rows.append(normalize_crew_freelance_row(raw, filename))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        rows.extend(file_rows)
        worker_label = file_rows[0]["worker"] if file_rows else "unknown"
        file_summaries.append({
            "filename": filename,
            "company":  worker_label,
            "rows":     len(file_rows),
            "issues":   errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── Talent freelance invoice extractor ───────────────────────────────────────

def normalize_talent_freelance_invoice(raw: dict, filename: str) -> list[dict]:
    talent_name    = clean_name(raw.get("talentName", "")) or "[missing information]"
    agency_name    = str(raw.get("agencyName", "")).strip()
    invoice_no     = str(raw.get("invoiceNo", "")).strip() or "[missing information]"
    invoice_date   = str(raw.get("invoiceDate", "")).strip() or "[missing information]"
    work_dates     = str(raw.get("workDates", "")).strip()
    payment_entity = str(raw.get("paymentEntity", "")).strip()

    days = raw.get("daysWorked")
    if isinstance(days, float):
        days = int(days)
    elif not isinstance(days, int):
        days = None

    talent_wages    = normalize_amount(raw.get("talentWages", 0))
    agency_fee      = normalize_amount(raw.get("agencyFee", 0))
    agency_expenses = normalize_amount(raw.get("agencyExpenses", 0))
    misc_pymt       = round(agency_fee + agency_expenses, 2)
    work_state      = clean_state(raw.get("workState", ""))

    method = normalize_pymt_method(raw.get("pymtMethod", ""))
    pymt_no = normalize_pymt_number(method, raw.get("pymtNo", ""))

    talent_row = {
        "talentName":      talent_name,
        "title":           "Talent",
        "rowType":         "talent",
        "invoiceNo":       invoice_no,
        "invoiceDate":     invoice_date,
        "workDates":       work_dates,
        "daysWorked":      days,
        "wages":           talent_wages if talent_wages else "[missing information]",
        "miscPymt":        0,
        "qualify":         "",
        "includedOnPtip":  "NO",
        "workState":       work_state,
        "receivedInvoice": "YES",
        "paymentEntity":   payment_entity,
        "pymtMethod":      method,
        "pymtNo":          pymt_no,
        "street":          clean_address(raw.get("talentStreet", "")),
        "city":            clean_name(raw.get("talentCity", "")),
        "state":           clean_state(raw.get("talentState", "")),
        "zip":             clean_zip(raw.get("talentZip", "")),
        "sourceFile":      filename,
    }

    rows = [talent_row]

    if agency_name:
        agency_row = {
            "talentName":      clean_name(agency_name),
            "title":           "Agency Fee",
            "rowType":         "agency",
            "invoiceNo":       invoice_no,
            "invoiceDate":     invoice_date,
            "workDates":       work_dates,
            "daysWorked":      days,
            "wages":           0,
            "miscPymt":        misc_pymt if misc_pymt else "[missing information]",
            "qualify":         "",
            "includedOnPtip":  "NO",
            "workState":       "",
            "receivedInvoice": "YES",
            "paymentEntity":   payment_entity,
            "pymtMethod":      method,
            "pymtNo":          pymt_no,
            "street":          "",
            "city":            "",
            "state":           "",
            "zip":             "",
            "sourceFile":      filename,
        }
        rows.append(agency_row)

    return rows


@app.post("/extract-talent-freelance")
async def extract_talent_freelance(
    files:        list[UploadFile] = File(...),
    prodco_names: str              = Form(""),
    x_app_secret: str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client        = _client()
    names         = [n.strip() for n in prodco_names.split(",") if n.strip()]
    system_prompt = _load_talent_freelance_prompt(names)
    user_text     = "Extract talent freelance invoice data from these document pages."

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                return filename, raw_list, None
            except Exception as e:
                return filename, [], str(e)

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []

    for filename, raw_list, err in extraction_results:
        errs: list[str] = []
        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        if not raw_list:
            errs.append("no talent freelance data extracted — review manually")
            issues.append(f"{filename}: no talent freelance data extracted")
        else:
            for raw in raw_list:
                try:
                    file_rows.extend(normalize_talent_freelance_invoice(raw, filename))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        rows.extend(file_rows)
        talent_label = file_rows[0]["talentName"] if file_rows else "unknown"
        file_summaries.append({
            "filename": filename,
            "talent":   talent_label,
            "rows":     len(file_rows),
            "issues":   errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── Hours letter extractor ────────────────────────────────────────────────────

def normalize_hours_letter_row(raw: dict, filename: str) -> dict:
    worker = clean_name(raw.get("worker", "")) or "[missing information]"
    wages  = normalize_amount(raw.get("wages", 0))
    if wages == 0:
        wages = "[missing information]"

    days = raw.get("daysWorked")
    if isinstance(days, float):
        days = int(days)
    elif not isinstance(days, int):
        days = None

    return {
        "worker":       worker,
        "jobTitle":     clean_name(raw.get("jobTitle", "")),
        "daysWorked":   days,
        "wages":        wages,
        "invoiceTotal": wages,
        "invoiceDate":  str(raw.get("invoiceDate", "")).strip(),
        "company":      str(raw.get("company", "")).strip(),
        "sourceFile":   filename,
    }


@app.post("/extract-hours-letters")
async def extract_hours_letters(
    files: list[UploadFile] = File(...),
    x_app_secret: str = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client        = _client()
    system_prompt = _load_hours_letter_prompt()
    user_text     = "Extract crew hours and billable amounts from this hours confirmation letter."

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                return filename, raw_list, None
            except Exception as e:
                return filename, [], str(e)

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []
    sources: list[tuple[str, list[dict]]] = []   # (filename, file_rows) for labeling

    for filename, raw_list, err in extraction_results:
        errs: list[str] = []
        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        if not raw_list:
            errs.append("no hours letter data extracted — review manually")
            issues.append(f"{filename}: no hours letter data extracted")
        else:
            for raw in raw_list:
                try:
                    file_rows.append(normalize_hours_letter_row(raw, filename))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        rows.extend(file_rows)
        if file_rows:
            sources.append((filename, file_rows))
        company_label = file_rows[0]["company"] if file_rows and file_rows[0]["company"] else "unknown"
        file_summaries.append({
            "filename": filename,
            "company":  company_label,
            "rows":     len(file_rows),
            "issues":   errs,
        })

    # Assign invoice labels — mirrors Wrapbook "Fringe Report" pattern
    if len(sources) == 1:
        for row in sources[0][1]:
            row["invoiceNo"] = "Hours Letter"
    elif len(sources) > 1:
        for idx, (_, src_rows) in enumerate(sources, 1):
            for row in src_rows:
                row["invoiceNo"] = f"Hours Letter {idx:03d}"

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── Billings extractor ───────────────────────────────────────────────────────

@app.post("/extract-billings")
async def extract_billings(
    files:          list[UploadFile] = File(...),
    vendor_type:    str              = Form(""),
    vendor_name:    str              = Form(""),
    vendor_address: str              = Form(""),
    vendor_city:    str              = Form(""),
    vendor_state:   str              = Form(""),
    vendor_zip:     str              = Form(""),
    prodco_names:   str              = Form(""),
    work_state:     str              = Form("IL"),
    x_app_secret:   str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client        = _client()
    names         = [n.strip() for n in prodco_names.split(",") if n.strip()]
    system_prompt = _load_billing_prompt(vendor_name, vendor_type, names)
    user_text     = "Extract billing invoice data from these document pages."

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                return filename, raw_list, None
            except Exception as e:
                return filename, [], str(e)

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []

    for filename, raw_list, err in extraction_results:
        errs: list[str] = []
        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        if not raw_list:
            errs.append("no billing data extracted — review manually")
            issues.append(f"{filename}: no billing data extracted")
        else:
            for raw in raw_list:
                try:
                    file_rows.append(normalize_billing_row(
                        raw,
                        vendor_type=vendor_type,
                        vendor_name=vendor_name,
                        vendor_address=vendor_address,
                        vendor_city=vendor_city,
                        vendor_state=vendor_state,
                        vendor_zip=vendor_zip,
                        work_state=work_state,
                        filename=filename,
                    ))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        rows.extend(file_rows)
        file_summaries.append({
            "filename": filename,
            "company":  vendor_name or "unknown",
            "rows":     len(file_rows),
            "issues":   errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── Agency Sub-Vendors extractor ─────────────────────────────────────────────

@app.post("/extract-agency-subvendors")
async def extract_agency_subvendors(
    files:          list[UploadFile] = File(...),
    agency_name:    str              = Form(""),
    agency_address: str              = Form(""),
    x_app_secret:   str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client        = _client()
    system_prompt = _load_agency_subvendors_prompt(agency_name, agency_address)
    user_text     = "Extract invoice data from these document pages."

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                return filename, raw_list, None
            except Exception as e:
                return filename, [], str(e)

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []

    for filename, raw_list, err in extraction_results:
        errs: list[str] = []
        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        if not raw_list:
            errs.append("no agency sub-vendor data extracted — review manually")
            issues.append(f"{filename}: no agency sub-vendor data extracted")
        else:
            for raw in raw_list:
                try:
                    file_rows.append(normalize_agency_subvendor_row(raw, agency_name, filename))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        rows.extend(file_rows)
        vendor_label = file_rows[0]["vendorName"] if file_rows else "unknown"
        file_summaries.append({
            "filename": filename,
            "company":  vendor_label,
            "rows":     len(file_rows),
            "issues":   errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── ProdCo Sub-Vendors extractor ─────────────────────────────────────────────

@app.post("/extract-prodco-subvendors")
async def extract_prodco_subvendors(
    files:              list[UploadFile] = File(...),
    prodco_name:        str              = Form(""),
    prodco_address:     str              = Form(""),
    sub_prodco_name:    str              = Form(""),
    sub_prodco_address: str              = Form(""),
    x_app_secret:       str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client        = _client()
    system_prompt = _load_prodco_subvendors_prompt(prodco_name, prodco_address, sub_prodco_name, sub_prodco_address)
    user_text     = "Extract invoice data from these document pages."

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                return filename, raw_list, None
            except Exception as e:
                return filename, [], str(e)

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []

    for filename, raw_list, err in extraction_results:
        errs: list[str] = []
        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        if not raw_list:
            errs.append("no prodco sub-vendor data extracted — review manually")
            issues.append(f"{filename}: no prodco sub-vendor data extracted")
        else:
            for raw in raw_list:
                try:
                    file_rows.append(normalize_prodco_subvendor_row(raw, prodco_name, filename))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        rows.extend(file_rows)
        vendor_label = file_rows[0]["vendorName"] if file_rows else "unknown"
        file_summaries.append({
            "filename": filename,
            "company":  vendor_label,
            "rows":     len(file_rows),
            "issues":   errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── Petty Cash extractor ──────────────────────────────────────────────────────

@app.post("/extract-petty-cash")
async def extract_petty_cash(
    files:        list[UploadFile] = File(...),
    prodco_name:  str              = Form(""),
    work_state:   str              = Form("IL"),
    x_app_secret: str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client        = _client()
    system_prompt = _load_petty_cash_prompt(prodco_name, work_state)
    user_text     = (
        "Extract all petty cash line items from this custodian's packet. "
        "Read the summary sheet first for the canonical ordered list, "
        "then check which receipts are present in the PDF."
    )

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        file_mb = len(data) / (1024 * 1024)
        if file_mb > 40:
            msg = (
                f"File too large to process ({file_mb:.0f} MB) -- "
                "compress to under 40 MB and resubmit"
            )
            return filename, [], msg, True

        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_petty_cash_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                # GPT-4o has been observed to come back empty on a file for
                # reasons that don't correspond to any real content problem
                # (confirmed via direct testing) -- retry with Claude before
                # giving up on the file entirely.
                if not raw_list:
                    print(f"[extract_petty_cash] {filename}: GPT-4o returned nothing, retrying with Claude", flush=True)
                    try:
                        anthropic_client = _anthropic_client()
                        raw_list = await loop.run_in_executor(
                            None,
                            functools.partial(_extract_petty_cash_from_file_claude, filename, data, system_prompt, anthropic_client, user_text=user_text),
                        )
                        _tag_claude_fallback_rows(raw_list)
                    except Exception as e:
                        print(f"[extract_petty_cash] {filename}: Claude fallback also failed: {e}", flush=True)
                return filename, raw_list, None, False
            except Exception as e:
                return filename, [], str(e), False

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []
    custodian_env_max: dict[str, int] = {}
    file_records: list[dict] = []

    for filename, raw_list, err, too_large in extraction_results:
        errs: list[str] = []

        if too_large:
            errs.append(err)
            issues.append(f"{filename}: {err}")
            file_summaries.append({
                "filename": filename,
                "company":  "unknown",
                "rows":     0,
                "issues":   errs,
            })
            continue

        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        env_totals: dict[int, float] = {}
        if not raw_list:
            errs.append("no petty cash data extracted - review manually")
            issues.append(f"{filename}: no petty cash data extracted")
            try:
                file_rows.append(normalize_petty_cash_row(
                    _petty_cash_unreadable_file_row(filename), work_state, filename,
                ))
            except Exception as e:
                issues.append(f"{filename}: placeholder row error: {e}")
        else:
            # Collect per-envelope totals from raw rows before normalization
            for raw in raw_list:
                en = raw.get("env_number", 1)
                try:
                    en = max(1, int(en))
                except (TypeError, ValueError):
                    en = 1
                et = normalize_amount(raw.get("envelope_total", 0))
                if et and en not in env_totals:
                    env_totals[en] = et
            for raw in raw_list:
                try:
                    file_rows.append(normalize_petty_cash_row(raw, work_state, filename))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        file_records.append({
            "filename":   filename,
            "file_rows":  file_rows,
            "env_totals": env_totals,
            "errs":       errs,
        })

    # Multi-part packets (e.g. "(1 of 3)", "(2 of 3)") sometimes have a
    # broken/placeholder custodian name on later parts' own cover sheet --
    # backfill from a sibling part that read it correctly BEFORE bucketing
    # envelope numbers below, since that bucketing is keyed by custodian_name.
    _backfill_petty_cash_custodian_names(file_records)

    for rec in file_records:
        filename   = rec["filename"]
        file_rows  = rec["file_rows"]
        env_totals = rec["env_totals"]
        errs       = rec["errs"]

        if file_rows:
            custodian = file_rows[0]["custodian_name"]
            env_offset = custodian_env_max.get(custodian, 0)
            if env_offset > 0:
                for r in file_rows:
                    r["env_number"] = r["env_number"] + env_offset
                env_totals = {k + env_offset: v for k, v in env_totals.items()}
            new_max = max(r["env_number"] for r in file_rows)
            custodian_env_max[custodian] = max(custodian_env_max.get(custodian, 0), new_max)

        if file_rows and env_totals:
            env_groups: dict[int, list] = {}
            for r in file_rows:
                env_groups.setdefault(r["env_number"], []).append(r)
            for env_num, group_rows in env_groups.items():
                et = env_totals.get(env_num, 0.0)
                if not et:
                    continue
                extracted_total = round(sum(r["amount"] for r in group_rows), 2)
                if abs(extracted_total - et) > 0.01:
                    diff = round(extracted_total - et, 2)
                    mismatch_note = (
                        f"Envelope total ${et:,.2f} | "
                        f"Extracted total ${extracted_total:,.2f} | "
                        f"Difference ${diff:,.2f}"
                    )
                    existing = group_rows[0]["notes"]
                    group_rows[0]["notes"] = (
                        f"{mismatch_note} | {existing}" if existing else mismatch_note
                    )

        rows.extend(file_rows)
        custodian_label = file_rows[0]["custodian_name"] if file_rows else "unknown"
        file_summaries.append({
            "filename": filename,
            "company":  custodian_label,
            "rows":     len(file_rows),
            "issues":   errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── GA Petty Cash extractor ───────────────────────────────────────────────────
# Reuses the IL petty cash engine (_extract_petty_cash_from_file: chunking,
# envelope-total reconciliation) unchanged -- the document structure (a
# custodian's cover sheet + receipts) is state-agnostic. Only the prompt and
# the normalized output fields differ (GA's FF1/FF2/AICP tax codes, dual
# proof-of-payment fields, etc.).

@app.post("/extract-ga-petty-cash")
async def extract_ga_petty_cash(
    files:        list[UploadFile] = File(...),
    prodco_name:  str              = Form(""),
    work_state:   str              = Form("GA"),
    x_app_secret: str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client        = _client()
    system_prompt = _load_ga_petty_cash_prompt(prodco_name, work_state)
    user_text     = (
        "Extract all petty cash line items from this custodian's packet. "
        "Read the cover sheet first for the canonical ordered list, "
        "then check which receipts are present in the PDF."
    )

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        file_mb = len(data) / (1024 * 1024)
        if file_mb > 40:
            msg = (
                f"File too large to process ({file_mb:.0f} MB) -- "
                "compress to under 40 MB and resubmit"
            )
            return filename, [], msg, True

        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_petty_cash_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                # GPT-4o has been observed to come back empty on a file for
                # reasons that don't correspond to any real content problem
                # (confirmed via direct testing) -- retry with Claude before
                # giving up on the file entirely.
                if not raw_list:
                    print(f"[extract_ga_petty_cash] {filename}: GPT-4o returned nothing, retrying with Claude", flush=True)
                    try:
                        anthropic_client = _anthropic_client()
                        raw_list = await loop.run_in_executor(
                            None,
                            functools.partial(_extract_petty_cash_from_file_claude, filename, data, system_prompt, anthropic_client, user_text=user_text),
                        )
                        _tag_claude_fallback_rows(raw_list)
                    except Exception as e:
                        print(f"[extract_ga_petty_cash] {filename}: Claude fallback also failed: {e}", flush=True)
                return filename, raw_list, None, False
            except Exception as e:
                return filename, [], str(e), False

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []
    custodian_env_max: dict[str, int] = {}
    file_records: list[dict] = []

    for filename, raw_list, err, too_large in extraction_results:
        errs: list[str] = []

        if too_large:
            errs.append(err)
            issues.append(f"{filename}: {err}")
            file_summaries.append({
                "filename": filename,
                "company":  "unknown",
                "rows":     0,
                "issues":   errs,
            })
            continue

        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        env_totals: dict[int, float] = {}
        if not raw_list:
            errs.append("no petty cash data extracted - review manually")
            issues.append(f"{filename}: no petty cash data extracted")
            try:
                file_rows.append(normalize_ga_petty_cash_row(
                    _petty_cash_unreadable_file_row(filename), work_state, filename,
                ))
            except Exception as e:
                issues.append(f"{filename}: placeholder row error: {e}")
        else:
            # Collect per-envelope totals from raw rows before normalization
            for raw in raw_list:
                en = raw.get("env_number", 1)
                try:
                    en = max(1, int(en))
                except (TypeError, ValueError):
                    en = 1
                et = normalize_amount(raw.get("envelope_total", 0))
                if et and en not in env_totals:
                    env_totals[en] = et
            for raw in raw_list:
                try:
                    file_rows.append(normalize_ga_petty_cash_row(raw, work_state, filename))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        file_records.append({
            "filename":   filename,
            "file_rows":  file_rows,
            "env_totals": env_totals,
            "errs":       errs,
        })

    # Multi-part packets (e.g. "(1 of 3)", "(2 of 3)") sometimes have a
    # broken/placeholder custodian name on later parts' own cover sheet --
    # backfill from a sibling part that read it correctly BEFORE bucketing
    # envelope numbers below, since that bucketing is keyed by custodian_name.
    _backfill_petty_cash_custodian_names(file_records)

    for rec in file_records:
        filename   = rec["filename"]
        file_rows  = rec["file_rows"]
        env_totals = rec["env_totals"]
        errs       = rec["errs"]

        if file_rows:
            custodian = file_rows[0]["custodian_name"]
            env_offset = custodian_env_max.get(custodian, 0)
            if env_offset > 0:
                for r in file_rows:
                    r["env_number"] = r["env_number"] + env_offset
                env_totals = {k + env_offset: v for k, v in env_totals.items()}
            new_max = max(r["env_number"] for r in file_rows)
            custodian_env_max[custodian] = max(custodian_env_max.get(custodian, 0), new_max)

        if file_rows and env_totals:
            env_groups: dict[int, list] = {}
            for r in file_rows:
                env_groups.setdefault(r["env_number"], []).append(r)
            for env_num, group_rows in env_groups.items():
                et = env_totals.get(env_num, 0.0)
                if not et:
                    continue
                extracted_total = round(sum(r["amount"] for r in group_rows), 2)
                if abs(extracted_total - et) > 0.01:
                    diff = round(extracted_total - et, 2)
                    mismatch_note = (
                        f"Envelope total ${et:,.2f} | "
                        f"Extracted total ${extracted_total:,.2f} | "
                        f"Difference ${diff:,.2f}"
                    )
                    existing = group_rows[0]["notes"]
                    group_rows[0]["notes"] = (
                        f"{mismatch_note} | {existing}" if existing else mismatch_note
                    )

        rows.extend(file_rows)
        custodian_label = file_rows[0]["custodian_name"] if file_rows else "unknown"
        file_summaries.append({
            "filename": filename,
            "company":  custodian_label,
            "rows":     len(file_rows),
            "issues":   errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── Agency Hours extractor ───────────────────────────────────────────────────

@app.post("/extract-agency-hours")
async def extract_agency_hours(
    files:        list[UploadFile] = File(...),
    agency_name:  str              = Form(""),
    x_app_secret: str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client        = _client()
    system_prompt = _load_agency_hours_prompt(agency_name)
    user_text     = "Extract crew hours data from these agency hours letter pages."

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                return filename, raw_list, None
            except Exception as e:
                return filename, [], str(e)

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []

    for filename, raw_list, err in extraction_results:
        errs: list[str] = []
        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        if not raw_list:
            errs.append("no agency hours data extracted — review manually")
            issues.append(f"{filename}: no agency hours data extracted")
        else:
            for raw in raw_list:
                try:
                    file_rows.append(normalize_agency_hours_row(raw, agency_name, filename))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        rows.extend(file_rows)
        agency_label = agency_name or (file_rows[0]["agencyName"] if file_rows else "unknown")
        file_summaries.append({
            "filename": filename,
            "company":  agency_label,
            "rows":     len(file_rows),
            "issues":   errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── Retainer Billings extractor ───────────────────────────────────────────────

@app.post("/extract-retainer-billings")
async def extract_retainer_billings(
    files:        list[UploadFile] = File(...),
    agency_name:  str              = Form(""),
    prodco_names: str              = Form(""),
    x_app_secret: str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client        = _client()
    names         = [n.strip() for n in prodco_names.split(",") if n.strip()]
    system_prompt = _load_billing_prompt(agency_name, "agency", names)
    user_text     = "Extract retainer billing invoice data from these document pages."

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                return filename, raw_list, None
            except Exception as e:
                return filename, [], str(e)

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []

    for filename, raw_list, err in extraction_results:
        errs: list[str] = []
        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        if not raw_list:
            errs.append("no retainer billing data extracted — review manually")
            issues.append(f"{filename}: no retainer billing data extracted")
        else:
            for raw in raw_list:
                try:
                    file_rows.append(normalize_retainer_billing_row(raw, agency_name, filename))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        rows.extend(file_rows)
        file_summaries.append({
            "filename": filename,
            "company":  agency_name or "unknown",
            "rows":     len(file_rows),
            "issues":   errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── Residency documents extractor ────────────────────────────────────────────

@app.post("/extract-residency-docs")
async def extract_residency_docs(
    files:        list[UploadFile] = File(...),
    shoot_date:   str              = Form(default=""),
    x_app_secret: str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    claude_client = _anthropic_client()
    system_prompt = _load_residency_docs_prompt()
    res_user_text = "Extract personal information from these residency verification documents."

    loaded_res = []
    for uf in sorted(files, key=lambda f: (f.filename or "").lower()):
        data = await uf.read()
        loaded_res.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def process_one_residency(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(
                        _extract_from_file_claude,
                        filename, data, system_prompt, claude_client,
                        user_text=res_user_text, dpi_scale=2.0, max_dim=2000, max_pages=4,
                    )
                )
            except Exception as e:
                return filename, False, [], [str(e)]
            return filename, False, raw_list, []

    res_results = await asyncio.gather(*[process_one_residency(fn, d) for fn, d in loaded_res])

    rows, issues, file_summaries = [], [], []

    for filename, handwritten, raw_list, errs in res_results:
        for e in errs:
            issues.append(f"{filename}: {e}")
        file_rows = []
        file_errs = list(errs)
        if not raw_list:
            file_errs.append("no residency document data extracted — review manually")
            issues.append(f"{filename}: no residency document data extracted")
        else:
            for raw in raw_list:
                try:
                    file_rows.append(normalize_residency_row(raw, filename, handwritten=handwritten, shoot_date=shoot_date))
                except Exception as e:
                    msg = f"row normalization error: {e}"
                    issues.append(f"{filename}: {msg}")
                    file_errs.append(msg)
        rows.extend(file_rows)
        name_label = file_rows[0]["documentName"] if file_rows else "unknown"
        file_summaries.append({
            "filename": filename,
            "company":  name_label,
            "rows":     len(file_rows),
            "issues":   file_errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


@app.post("/extract-diversity-docs")
async def extract_diversity_docs(
    files:        list[UploadFile] = File(...),
    x_app_secret: str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    claude_client = _anthropic_client()
    div_prompt    = _load_diversity_form_prompt()

    loaded = []
    for uf in sorted(files, key=lambda f: (f.filename or "").lower()):
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def process_one_diversity(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(
                        _extract_from_file_claude,
                        filename, data, div_prompt, claude_client,
                        user_text="Extract diversity information from this form.",
                        dpi_scale=2.0, max_dim=2000, max_pages=2,
                    )
                )
            except Exception as e:
                return filename, [], [str(e)]
            return filename, raw_list, []

    results = await asyncio.gather(*[process_one_diversity(fn, d) for fn, d in loaded])

    rows, issues = [], []
    for filename, raw_list, errs in results:
        for e in errs:
            issues.append(f"{filename}: {e}")
        for raw in raw_list:
            try:
                rows.append(normalize_diversity_row(raw, filename))
            except Exception as e:
                issues.append(f"{filename}: diversity row error: {e}")

    return {"rows": rows, "issues": issues}


class _MatchNamesRequest(BaseModel):
    residency_names: list[str] = []
    diversity_names: list[str] = []


@app.post("/match-names")
async def match_names(
    payload:      _MatchNamesRequest,
    x_app_secret: str = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    if not payload.residency_names or not payload.diversity_names:
        return {"mapping": {}}

    claude_client = _anthropic_client()

    res_list = "\n".join(f"  {i+1}. {n}" for i, n in enumerate(payload.residency_names))
    div_list = "\n".join(f"  {i+1}. {n}" for i, n in enumerate(payload.diversity_names))

    user_prompt = (
        "Match crew member names across two lists. The second list may use a different "
        "name format or contain OCR/transcription errors.\n\n"
        f"LIST A — map FROM (these become the JSON keys):\n{res_list}\n\n"
        f"LIST B — match TO (these become the JSON values):\n{div_list}\n\n"
        "For each name in List A, find the best matching name in List B using fuzzy logic:\n"
        "- Names may be in different formats: 'First Last' and 'Last, First' or "
        "'Last, First Middle' refer to the same person when name components match — "
        "match them regardless of order (e.g. 'Damian Huck' matches 'Huck, Damian Michael')\n"
        "- Ignore middle names: 'Damian Huck' matches 'Huck, Damian Michael'\n"
        "- Correct OCR errors (e.g. \"Cowley\" vs \"Conley\", \"Lecy\" vs \"Levy\", \"Pawch\" vs \"Pawela\")\n"
        "- Handle first-name abbreviations (\"Josh\" matches \"Joshua\", \"Matt\" matches \"Matthew\")\n"
        "- When multiple people share the same last name, use first names to distinguish them\n"
        "- Map to null only if no reasonable match exists in List B\n\n"
        "Return ONLY a JSON object. Keys must be EXACTLY the List A names as given:\n"
        "{\"List A Name 1\": \"Matched List B Name or null\", ...}"
    )

    loop = asyncio.get_running_loop()
    try:
        resp = await loop.run_in_executor(
            None,
            functools.partial(
                claude_client.messages.create,
                model="claude-sonnet-5",
                max_tokens=4096,
                messages=[{"role": "user", "content": user_prompt}],
            )
        )
    except Exception as e:
        return {"mapping": {}, "error": str(e)}

    text_block = next((b for b in resp.content if b.type == "text"), None)
    if not text_block:
        return {"mapping": {}}

    raw = text_block.text.strip()
    try:
        mapping = json.loads(raw)
    except Exception:
        m = re.search(r'\{.*\}', raw, re.DOTALL)
        if m:
            try:
                mapping = json.loads(m.group())
            except Exception:
                mapping = {}
        else:
            mapping = {}

    return {"mapping": mapping}


# ── Call sheet extractor ──────────────────────────────────────────────────────

def _repair_truncated_json(raw: str):
    """Best-effort recovery for a JSON value that got cut off mid-stream
    (e.g. the response hit max_tokens before finishing). Scans for the last
    point where a complete object/array had just closed at any nesting
    depth, drops everything after that point (the final incomplete
    element), and closes out whatever brackets were still open. Returns the
    parsed value, or None if nothing could be salvaged.

    E.g. for a large call sheet where Claude's crew list for day 1 alone
    exceeds max_tokens, this recovers every complete person entry already
    written instead of discarding the entire response over one truncated
    trailing entry.
    """
    start = None
    for ch in ("[", "{"):
        idx = raw.find(ch)
        if idx != -1 and (start is None or idx < start):
            start = idx
    if start is None:
        return None
    raw = raw[start:]

    stack: list[str] = []
    in_string = False
    escape = False
    last_clean_end = None
    last_clean_stack = None

    for i, ch in enumerate(raw):
        if in_string:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch in "[{":
            stack.append(ch)
        elif ch in "]}":
            if not stack:
                break
            stack.pop()
            last_clean_end = i + 1
            last_clean_stack = list(stack)

    if last_clean_end is None:
        return None

    closers = {"[": "]", "{": "}"}
    candidate = raw[:last_clean_end] + "".join(closers[c] for c in reversed(last_clean_stack))
    candidate = re.sub(r",\s*([\]}])", r"\1", candidate)
    try:
        return json.loads(candidate)
    except Exception:
        return None


def _call_claude_call_sheet(images_b64: list, system_prompt: str, client) -> list:
    """Send a batch of call sheet page images to Claude.

    Returns a list of {date, crew} dicts — one entry per page in the batch.
    Sending all pages from one file at once lets Claude infer missing years
    from surrounding context (e.g., Day 3 lacks a year but Day 1 shows 2023).
    """
    content = []
    for img in images_b64:
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": img},
        })
    content.append({
        "type": "text",
        "text": (
            f"This document has {len(images_b64)} page(s). "
            "Extract the shoot date and crew/talent list from each page. "
            "Return a JSON array with exactly one object per page, in page order."
        ),
    })
    resp = client.messages.create(
        model="claude-sonnet-5",
        # 4096 was not enough for a real multi-day call sheet with 60+ crew
        # per day -- confirmed via a standalone repro script that the YUMB
        # 008 call sheet (3 days, 190 total crew) needs ~8900 output tokens
        # to finish without truncating. 16000 leaves comfortable headroom.
        max_tokens=16000,
        system=system_prompt,
        messages=[{"role": "user", "content": content}],
    )
    text_block = next((b for b in resp.content if b.type == "text"), None)
    if not text_block:
        print(f"[_call_claude_call_sheet] No text block in Claude's response. stop_reason={resp.stop_reason!r}", flush=True)
        return []
    raw = text_block.text.strip()
    try:
        return json.loads(raw)
    except Exception:
        print(f"[_call_claude_call_sheet] JSON parse failed (stop_reason={resp.stop_reason!r}). "
              f"Raw response (first 1000 chars): {raw[:1000]!r}", flush=True)
        m = re.search(r'\[.*\]', raw, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                pass
        repaired = _repair_truncated_json(raw)
        if repaired is not None:
            print(f"[_call_claude_call_sheet] Recovered partial data from a truncated response "
                  f"instead of discarding everything ({len(repaired)} page object(s)).", flush=True)
            return repaired
        print("[_call_claude_call_sheet] Could not recover any usable data from this response.", flush=True)
        return []


def _normalize_cs_date(val: str) -> str:
    """Normalize various date formats to MM/DD/YYYY.

    Claude may return dates in many formats: "01/30/2023", "1/30/23",
    "2023-01-30", "January 30, 2023", etc.  We handle the common patterns
    and pass through anything we can't parse so issues are surfaced.
    """
    if not val:
        return ""
    s = str(val).strip()
    # Already MM/DD/YYYY
    if re.match(r'^\d{2}/\d{2}/\d{4}$', s):
        return s
    # M/D/YYYY or MM/DD/YYYY with single-digit parts
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
    # MM/DD/YY
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2})$', s)
    if m:
        yr = m.group(3)
        year = f"20{yr}" if int(yr) < 50 else f"19{yr}"
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{year}"
    # YYYY-MM-DD
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', s)
    if m:
        return f"{m.group(2)}/{m.group(3)}/{m.group(1)}"
    return s


def _cs_date_sort_key(d: str):
    try:
        parts = d.split("/")
        return (int(parts[2]), int(parts[0]), int(parts[1]))
    except Exception:
        return (9999, 0, 0)


@app.post("/extract-call-sheet")
async def extract_call_sheet(
    files:        list[UploadFile] = File(...),
    x_app_secret: str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    if not files:
        return {"crew": [], "shoot_dates": [], "issues": []}

    claude_client = _anthropic_client()
    system_prompt = _load_call_sheet_prompt()

    loaded = []
    for uf in sorted(files, key=lambda f: (f.filename or "").lower()):
        data = await uf.read()
        if data:
            loaded.append((uf.filename, data))

    if not loaded:
        return {"crew": [], "shoot_dates": [], "issues": []}

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def process_file(filename, data):
        async with sem:
            try:
                # Call sheets pack small multi-column text (crew/client/agency/talent
                # tables side by side) -- denser than residency/diversity docs, so this
                # needs at least their resolution (2.0/2000), with a bit more headroom.
                images = _file_to_images_b64(filename, data, dpi_scale=2.0, max_dim=2400)
            except Exception as e:
                return filename, [], [f"{filename}: {e}"]

            # Batch pages to stay under Claude's 40 MB image limit per call
            MAX_BYTES = 40 * 1024 * 1024
            batches, cur, cur_size = [], [], 0
            for img in images:
                approx = len(img) * 3 // 4
                if cur and cur_size + approx > MAX_BYTES:
                    batches.append(cur)
                    cur, cur_size = [img], approx
                else:
                    cur.append(img)
                    cur_size += approx
            if cur:
                batches.append(cur)

            page_results, file_issues = [], []
            for batch in batches:
                try:
                    result = await loop.run_in_executor(
                        None,
                        functools.partial(
                            _call_claude_call_sheet,
                            batch, system_prompt, claude_client,
                        )
                    )
                    if isinstance(result, list):
                        page_results.extend(result)
                except Exception as e:
                    file_issues.append(f"{filename}: {e}")

            return filename, page_results, file_issues

    results = await asyncio.gather(*[process_file(fn, d) for fn, d in loaded])

    all_page_data, issues = [], []
    for filename, page_results, file_issues in results:
        issues.extend(file_issues)
        for page in page_results:
            if isinstance(page, dict):
                all_page_data.append(page)

    # Aggregate: deduplicate crew across all pages/files
    shoot_dates_set: set = set()
    crew_by_key: dict = {}

    for page in all_page_data:
        raw_date = page.get("date") or ""
        date_str = _normalize_cs_date(str(raw_date)) if raw_date else ""
        if date_str:
            shoot_dates_set.add(date_str)

        for person in page.get("crew") or []:
            if not isinstance(person, dict):
                continue
            name     = (person.get("name") or "").strip()
            position = (person.get("position") or "").strip()
            if not name:
                continue
            key = name.lower().strip()
            if key not in crew_by_key:
                crew_by_key[key] = {"name": name, "positions": [], "dates": []}
            entry = crew_by_key[key]
            if position and position not in entry["positions"]:
                entry["positions"].append(position)
            if date_str and date_str not in entry["dates"]:
                entry["dates"].append(date_str)

    shoot_dates = sorted(list(shoot_dates_set), key=_cs_date_sort_key)

    crew = []
    for entry in crew_by_key.values():
        entry["dates"] = sorted(entry["dates"], key=_cs_date_sort_key)
        crew.append(entry)
    crew.sort(key=lambda e: e["name"].lower())

    return {"crew": crew, "shoot_dates": shoot_dates, "issues": issues}


# ── Talent & Extras extractor ────────────────────────────────────────────────

@app.post("/extract-talent")
async def extract_talent_endpoint(
    pdf_files:      list[UploadFile] = File(default=[]),
    ptip_file:      UploadFile       = File(default=None),   # ER: single PTIP file
    ptip_files:     list[UploadFile] = File(default=[]),     # Teams: multiple PTIP files
    project_title:  str              = Form(default=""),
    workbook_type:  str              = Form(default=""),
    payroll_company: str             = Form(default="er"),   # "er" or "teams"
    x_app_secret:   str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    pdf_bytes_list: list[tuple[str, bytes]] = []
    for uf in sorted(pdf_files or [], key=lambda f: (f.filename or "").lower()):
        data = await uf.read()
        if data:
            pdf_bytes_list.append((uf.filename, data))

    if payroll_company.lower() == 'teams':
        ptip_bytes_list: list[bytes] = []
        for uf in (ptip_files or []):
            data = await uf.read()
            if data:
                ptip_bytes_list.append(data)
        # Also accept single ptip_file upload for Teams if ptip_files is empty
        if not ptip_bytes_list and ptip_file:
            data = await ptip_file.read()
            if data:
                ptip_bytes_list.append(data)

        if not pdf_bytes_list and not ptip_bytes_list:
            raise HTTPException(400, "Provide at least one PDF or PTIP file.")

        return extract_teams_talent(
            pdf_files=pdf_bytes_list,
            ptip_bytes_list=ptip_bytes_list,
            project_title=project_title,
            workbook_type=workbook_type,
        )

    # Default: Extreme Reach
    ptip_bytes: bytes | None = None
    if ptip_file:
        ptip_bytes = await ptip_file.read()
        if not ptip_bytes:
            ptip_bytes = None

    if not pdf_bytes_list and not ptip_bytes:
        raise HTTPException(400, "Provide at least one PDF or a PTIP file.")

    return extract_talent(
        pdf_files=pdf_bytes_list,
        ptip_bytes=ptip_bytes,
        project_title=project_title,
        workbook_type=workbook_type,
        openai_key=OPENAI_API_KEY,
    )


# ── Consolidated run summary email ───────────────────────────────────────────

class _FileSummaryIn(BaseModel):
    filename: str = ""
    company:  str = ""
    rows:     int | None = 0
    issues:   list[str] = []

class _RunDataIn(BaseModel):
    endpoint: str = ""
    files:    list[_FileSummaryIn] = []
    issues:   list[str] = []

class _RunSummaryRequest(BaseModel):
    project_title: str = ""
    workbook_type: str = ""
    runs:          list[_RunDataIn] = []


@app.post("/send-run-summary")
async def send_run_summary_endpoint(
    payload: _RunSummaryRequest,
    x_app_secret: str = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    runs_dicts = [
        {
            "endpoint": r.endpoint,
            "files":    [f.dict() for f in r.files],
            "issues":   r.issues,
        }
        for r in payload.runs
    ]
    send_run_summary(
        project_title=payload.project_title,
        workbook_type=payload.workbook_type,
        runs=runs_dicts,
    )
    return {"ok": True}


# ── GA AP extractor ──────────────────────────────────────────────────────────

_GA_AP_FF1_VALID  = {"HT","BX","CR","AF","GX","HF","FA1","FA2","PD1","PD2","LO","NQ",""}
_GA_AP_FF2_VALID  = {"GS","GL","DL","NQ",""}
_GA_AP_QUAL_VALID = {"YES","NO-GA","NO-OOS"}


def normalize_date_iso(val: str) -> str:
    """Return YYYY-MM-DD. Accepts YYYY-MM-DD, M/D/YY, M/D/YYYY, MM/DD/YY, MM/DD/YYYY."""
    if not val:
        return ""
    s = str(val).strip()
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', s)
    if m:
        return s
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})$', s)
    if m:
        mo, dy, yr = m.group(1), m.group(2), m.group(3)
        if len(yr) == 2:
            yr = ("20" if int(yr) <= 50 else "19") + yr
        return f"{yr}-{mo.zfill(2)}-{dy.zfill(2)}"
    return s


def normalize_ga_ap_row(raw: dict) -> dict:
    ff1 = str(raw.get("ff1", "")).strip().upper()
    if ff1 not in _GA_AP_FF1_VALID:
        ff1 = ""

    ff2 = str(raw.get("ff2", "")).strip().upper()
    if ff2 not in _GA_AP_FF2_VALID:
        ff2 = ""

    qual = str(raw.get("qualified", "")).strip().upper()
    if qual not in _GA_AP_QUAL_VALID:
        qual = "NO-GA"

    aicp = raw.get("aicp_code")
    try:
        aicp = int(aicp)
        if not 1 <= aicp <= 25:
            aicp = None
    except (TypeError, ValueError):
        aicp = None

    nights = raw.get("nights")
    try:
        nights = int(nights) if nights is not None else None
    except (TypeError, ValueError):
        nights = None

    def yn(val):
        return "Yes" if str(val or "").strip().lower() in ("yes", "true", "1") else "No"

    return {
        "check_number":             str(raw.get("check_number", "")).strip(),
        "invoice_number":           str(raw.get("invoice_number", "")).strip(),
        "invoice_date":             normalize_date_iso(str(raw.get("invoice_date", ""))),
        "po_number":                str(raw.get("po_number", "")).strip(),
        "vendor_name":              clean_name(raw.get("vendor_name", "")),
        "payee_type":               "person" if str(raw.get("payee_type","")).strip().lower() == "person" else "company",
        "ff1":                      ff1,
        "ff2":                      ff2,
        "je_number":                str(raw.get("je_number", "")).strip(),
        "distribution_description": str(raw.get("distribution_description", "")).strip(),
        "last_name":                str(raw.get("last_name", "")).strip(),
        "first_name":               str(raw.get("first_name", "")).strip(),
        "home_state":               clean_state(raw.get("home_state", "")),
        "episode":                  str(raw.get("episode", "")).strip(),
        "nights":                   nights,
        "address":                  clean_address(raw.get("address", "")),
        "city":                     clean_name(raw.get("city", "")),
        "state":                    clean_state(raw.get("state", "")),
        "zip":                      clean_zip(raw.get("zip", "")),
        "amount":                   normalize_amount(raw.get("amount", 0)),
        "non_qualified":            normalize_amount(raw.get("non_qualified", 0)),
        "payment_method":           str(raw.get("payment_method", "")).strip(),
        "proof_of_payment":         yn(raw.get("proof_of_payment")),
        "payment_entity":           str(raw.get("payment_entity", "")).strip(),
        "received_invoice":         yn(raw.get("received_invoice")),
        "loan_out":                 yn(raw.get("loan_out")),
        "loan_out_individual_name": str(raw.get("loan_out_individual_name", "")).strip(),
        "sales_tax_on_invoice":     yn(raw.get("sales_tax_on_invoice")),
        "active_sales_tax_account": yn(raw.get("active_sales_tax_account")),
        "withholding":              str(raw.get("withholding", "")).strip(),
        "w9":                       yn(raw.get("w9")),
        "business_license":         yn(raw.get("business_license")),
        "aicp_code":                aicp,
        "qualified":                qual,
        "notes":                    str(raw.get("notes", "")).strip(),
        "website_address":          str(raw.get("website_address", "")).strip(),
        # Not written to AP tab — used downstream for tab routing (Payroll Roster, GL, etc.)
        "labor":                    bool(raw.get("labor", False)),
    }


def normalize_ga_petty_cash_row(raw: dict, work_state: str, filename: str) -> dict:
    def yn(val):
        s = str(val or "").strip().lower()
        return "Yes" if s in ("yes", "true", "1") else "No"

    ff1 = str(raw.get("ff1", "")).strip().upper()
    if ff1 not in _GA_AP_FF1_VALID:
        ff1 = ""

    ff2 = str(raw.get("ff2", "")).strip().upper()
    if ff2 not in _GA_AP_FF2_VALID:
        ff2 = "GS"

    aicp = raw.get("aicp_code")
    try:
        aicp = int(aicp)
        if not 1 <= aicp <= 25:
            aicp = None
    except (TypeError, ValueError):
        aicp = None

    custodian_name = clean_name(raw.get("custodian_name", ""))
    last_name, first_name = _split_custodian_name(custodian_name)

    nights = raw.get("nights")
    try:
        nights = int(nights) if nights is not None else None
    except (TypeError, ValueError):
        nights = None

    env_num = raw.get("env_number", 1)
    try:
        env_num = max(1, int(env_num))
    except (TypeError, ValueError):
        env_num = 1

    line_num = raw.get("line_number", 0)
    try:
        line_num = max(0, int(line_num))
    except (TypeError, ValueError):
        line_num = 0

    received = str(raw.get("received_invoice", "NO")).strip().upper()
    if received not in ("YES", "NO"):
        received = "NO"

    proof_vendor = str(raw.get("proof_of_payment_vendor", "NO")).strip().upper()
    if proof_vendor not in ("YES", "NO"):
        proof_vendor = "NO"

    # Uses the literal "MISSING" value (not "No") to match the production's
    # own established convention for this field -- see the real Coke 012
    # example workbook, where "MISSING" + a "Missing PC sign out sheet" note
    # is used whenever the custodian's own cover sheet is unsigned/absent.
    proof_remit = str(raw.get("proof_of_pc_remittance_crew", "MISSING")).strip().upper()
    if proof_remit not in ("YES", "MISSING"):
        proof_remit = "MISSING"

    return {
        "custodian_name":              custodian_name,
        "env_number":                  env_num,
        "line_number":                 line_num,
        "vendor_name":                 clean_name(raw.get("vendor_name", "")),
        "invoice_number":              str(raw.get("invoice_number", "")).strip() or "Receipt",
        "invoice_date":                normalize_date_iso(str(raw.get("invoice_date", ""))),
        "amount":                      normalize_amount(raw.get("amount", 0)),
        "distribution_description":    str(raw.get("distribution_description", "")).strip(),
        "ff1":                         ff1,
        "ff2":                         ff2,
        "je_number":                   str(raw.get("je_number", "")).strip(),
        "check_number":                str(raw.get("check_number", "")).strip(),
        "po_number":                   str(raw.get("po_number", "")).strip(),
        "last_name":                   last_name,
        "first_name":                  first_name,
        "home_state":                  clean_state(raw.get("home_state", "")),
        "episode":                     str(raw.get("episode", "")).strip(),
        "nights":                      nights,
        "address":                     clean_address(raw.get("address", "")),
        "city":                        clean_name(raw.get("city", "")),
        "state":                       clean_state(raw.get("state", "")),
        "zip":                         clean_zip(raw.get("zip", "")),
        "proof_of_payment_vendor":     proof_vendor,
        "proof_of_pc_remittance_crew": proof_remit,
        "received_invoice":            received,
        "sales_tax_on_invoice":        yn(raw.get("sales_tax_on_invoice")),
        "active_sales_tax_account":    yn(raw.get("active_sales_tax_account")),
        "withholding":                 str(raw.get("withholding", "")).strip(),
        "w9":                          yn(raw.get("w9")),
        "business_license":            yn(raw.get("business_license")),
        "aicp_code":                   aicp,
        # Always blank/zero at extraction time -- the automation never
        # disqualifies a line itself, it only flags candidates via notes.
        "qualified":                   "",
        "non_qualified":               0,
        "notes":                       str(raw.get("notes", "")).strip(),
        "website_address":             str(raw.get("website_address", "")).strip(),
        "sourceFile":                  filename,
    }


@app.post("/extract-ga-ap")
async def extract_ga_ap(
    files:           list[UploadFile] = File(...),
    prodco_name:     str              = Form(""),
    prodco_address:  str              = Form(""),
    agency_name:     str              = Form(""),
    work_state:      str              = Form("GA"),
    payer_entities:  str              = Form("[]"),
    x_app_secret:    str              = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    try:
        entities = json.loads(payer_entities)
        if not isinstance(entities, list):
            entities = []
    except Exception:
        entities = []

    # Fall back to scalar fields if payer_entities was not provided
    if not entities:
        if prodco_name.strip():
            entities.append({"role": "prodco", "name": prodco_name.strip(), "address": prodco_address.strip()})
        if agency_name.strip():
            entities.append({"role": "agency", "name": agency_name.strip(), "address": ""})

    files = sorted(files, key=lambda f: (f.filename or "").lower())

    client        = _client()
    system_prompt = _load_ga_ap_prompt(entities, work_state)
    user_text     = "Extract invoice data from these document pages."

    loaded = []
    for uf in files:
        data = await uf.read()
        loaded.append((uf.filename, data))

    loop = asyncio.get_running_loop()
    sem  = asyncio.Semaphore(5)

    async def _extract_one(filename, data):
        async with sem:
            try:
                raw_list = await loop.run_in_executor(
                    None,
                    functools.partial(_extract_from_file, filename, data, system_prompt, client, user_text=user_text),
                )
                return filename, raw_list, None
            except Exception as e:
                return filename, [], str(e)

    extraction_results = await asyncio.gather(*[_extract_one(fn, d) for fn, d in loaded])

    rows, issues, file_summaries = [], [], []

    for filename, raw_list, err in extraction_results:
        errs: list[str] = []
        if err:
            errs.append(err)
            issues.append(f"{filename}: {err}")

        file_rows: list[dict] = []
        if not raw_list:
            errs.append("no GA AP data extracted — review manually")
            issues.append(f"{filename}: no GA AP data extracted")
        else:
            for raw in raw_list:
                try:
                    file_rows.append(normalize_ga_ap_row(raw))
                except Exception as e:
                    errs.append(f"row normalization error: {e}")
                    issues.append(f"{filename}: row normalization error: {e}")

        rows.extend(file_rows)
        vendor_label = file_rows[0]["vendor_name"] if file_rows else "unknown"
        file_summaries.append({
            "file":   filename,
            "rows":   len(file_rows),
            "issues": errs,
        })

    return {"rows": rows, "issues": issues, "files": file_summaries}


# ── GA AP call sheet position matching ───────────────────────────────────────
# Call sheet extraction stays on /extract-call-sheet (Claude, unchanged). This
# endpoint only does the matching step, on OpenAI, per the Claude-extracts /
# OpenAI-matches split: extraction is the hard visual-parsing task, matching a
# name against an already-extracted list is the easy one.

class _CrewEntryIn(BaseModel):
    name:      str       = ""
    positions: list[str] = []
    dates:     list[str] = []

class _MatchApPositionsRequest(BaseModel):
    ap_names: list[str]         = []
    crew:     list[_CrewEntryIn] = []


@app.post("/match-ap-positions")
async def match_ap_positions(
    payload:      _MatchApPositionsRequest,
    x_app_secret: str = Header(default=""),
):
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    if not payload.ap_names or not payload.crew:
        return {"mapping": {}, "issues": []}

    client = _client()

    ap_list = "\n".join(f"  {i+1}. {n}" for i, n in enumerate(payload.ap_names))
    crew_lines = []
    for c in payload.crew:
        if not c.name:
            continue
        pos = ", ".join(c.positions) if c.positions else "(none listed)"
        dates = ", ".join(c.dates) if c.dates else "(none listed)"
        crew_lines.append(f"  - {c.name} — positions: {pos} — dates worked: {dates}")
    crew_list = "\n".join(crew_lines)

    # NOTE: this prompt deliberately forces the model to write out a per-name
    # verdict (candidate + one-line reasoning) for EVERY AP name before we
    # derive the mapping, instead of asking it to jump straight to a compact
    # {ap_name: position} JSON object. Asking for the compact form directly
    # was empirically shown (via a standalone debug script hitting this same
    # prompt/model) to silently skip correct fuzzy/exact matches once the
    # list got past the first several names -- the model has no scratch
    # space to actually check "did I already consider this crew member for
    # a different AP name" and appears to guess/pattern-complete instead.
    # Forcing an explicit reasoning field per name, mirroring how the model
    # behaves when asked to reason about just a few names in isolation,
    # fixed every one of the 3 real-match failures we reproduced.
    user_prompt = (
        "You match crew member names from a production AP (accounts payable) spreadsheet "
        "against a call sheet's crew roster, and return each matched person's position.\n\n"
        f"AP NAMES — you must produce exactly one entry in \"matches\" for EVERY one of these "
        f"{len(payload.ap_names)} names, in this same order, with no omissions.\n{ap_list}\n\n"
        f"CALL SHEET CREW:\n{crew_list}\n\n"
        "MATCHING RULES:\n"
        "- Use fuzzy matching: nicknames (\"Joe\"/\"Joseph\", \"Dave\"/\"David\", \"Sam\"/\"Samuel\"), "
        "spacing/capitalization differences (\"Deschutter\" matches \"De Schutter\"), middle names, "
        "and minor spelling differences all count as the same person.\n"
        "- IMPORTANT: more than one AP name can legitimately match the SAME crew member -- this "
        "happens whenever different invoices spelled the same real person's name differently. "
        "Matching a crew member to one AP name does NOT use them up or make them unavailable for "
        "a different AP name -- even when one of the two is an exact, verbatim match to the crew "
        "entry and the other is a fuzzy variant. A single crew member can be the correct answer "
        "for as many AP names as plausibly refer to them.\n"
        "- If a person has multiple positions listed and they are all the same, use that position. "
        "If they genuinely differ across dates, use the position tied to the most dates worked.\n"
        "- The position you return must always be a SINGLE job title, never more than one combined "
        "together. Sometimes a single call sheet entry already combines two roles for one day with "
        "a separator (e.g. \"2nd Grip/Grip\", \"Set Dresser, Leadman\") because that person double-"
        "hatted that day -- if the position you'd otherwise return contains a \"/\", \",\", \"&\", or "
        "\"and\" joining two roles, pick just the FIRST role listed and drop the rest. Never let a "
        "\"/\", \",\", or \"&\" appear anywhere in the position you return.\n"
        "- If no reasonable match exists on the call sheet for an AP name, set call_sheet_match to "
        "null and leave position empty -- do not guess.\n\n"
        "For EACH AP name, independently: scan the ENTIRE call sheet crew list from top to bottom "
        "(do not stop early just because a similarly-named person elsewhere already matched a "
        "different AP name -- re-scan the full list every time), pick the single best candidate "
        "(or none), write one short sentence of reasoning, and give the single position in Title "
        "Case with NO parentheses and no combined roles (e.g. \"Gaffer\", \"Key Grip\", \"2nd Props\").\n\n"
        "Return ONLY valid JSON, no markdown, no explanation, in exactly this shape:\n"
        "{\"matches\": [{\"ap_name\": \"AP Name As Given\", \"call_sheet_match\": \"Crew Name As "
        "Listed\" or null, \"reasoning\": \"short sentence\", \"position\": \"Title Case Position "
        "or empty string\"}]}"
    )

    loop = asyncio.get_running_loop()
    try:
        result = await loop.run_in_executor(
            None,
            functools.partial(_call_gpt_text_json, user_prompt, client, max_tokens=6000),
        )
    except Exception as e:
        return {"mapping": {}, "issues": [f"Position matching failed: {e}"]}

    matches = result.get("matches") if isinstance(result, dict) else None
    if not isinstance(matches, list):
        matches = []

    # Derive the mapping/issues from the per-name verdicts. A name only
    # lands in the mapping if the model gave it both a call_sheet_match and
    # a non-empty position; everything else becomes an issue note.
    fixed_mapping = {}
    issues = []
    seen_names = set()
    for m in matches:
        if not isinstance(m, dict):
            continue
        ap_name = str(m.get("ap_name") or "").strip()
        if not ap_name:
            continue
        seen_names.add(ap_name)
        pos = str(m.get("position") or "").strip()
        # Defensive: the model was told to never combine roles, but if it
        # still returns e.g. "2nd Grip/Grip" or "Set Dresser, Leadman",
        # keep only the first-listed role rather than passing the combo
        # through -- downstream (PD1/PD2 promotion, description overwrite)
        # expects a single clean title.
        pos = re.split(r"\s*(?:/|,|&|\band\b)\s*", pos, maxsplit=1, flags=re.IGNORECASE)[0].strip()
        call_sheet_match = m.get("call_sheet_match")
        if pos and call_sheet_match:
            if not (pos.startswith("(") and pos.endswith(")")):
                pos = f"({pos})"
            fixed_mapping[ap_name] = pos
        else:
            issues.append(f"{ap_name}: No match found on call sheet.")

    # Belt-and-suspenders: if the model dropped an AP name entirely instead
    # of giving it a null-match verdict, still surface it instead of letting
    # it silently vanish.
    for n in payload.ap_names:
        if n not in seen_names:
            issues.append(f"{n}: No match found on call sheet.")

    return {"mapping": fixed_mapping, "issues": issues}


# ── GA workbook writer ────────────────────────────────────────────────────────

# Column numbers for the GA AP tab (1-indexed, matching the bundled template).
# Row 1 = metadata/labels  Row 2 = col notes  Row 3 = headers  Row 4+ = data
# Confirmed against frontend-bundled georgia.xlsx by Frontend Claude.
# NOTE: /build-ga-workbook is deprioritized; the frontend writes the workbook
# client-side.  This map is kept accurate for reference and for any server-side
# utility use.
_GA_AP_COL = {
    "seq":             1,   # A  — generated (1, 2, 3…)
    "check_number":    2,   # B
    "invoice_number":  3,   # C
    "invoice_date":    4,   # D  — datetime
    "po_number":       5,   # E
    "vendor_name":     6,   # F
    "ff1":             7,   # G  — short code (HT/BX/CR/AF/GX/HF/FA1/FA2/PD1/PD2/LO/NQ)
    "ff2":             8,   # H  — short code (GS/GL/DL/NQ)
    "source_code":     9,   # I  — constant "AP"
    "je_number":      10,   # J
    "description":    11,   # K
    "last_name":      12,   # L
    "first_name":     13,   # M
    "home_state":     14,   # N
    "episode":        15,   # O
    "nights":         16,   # P
    "address":        17,   # Q
    "city":           18,   # R
    "state":          19,   # S
    "zip":            20,   # T
    "org_cur":        21,   # U  — constant "US"
    "amount":         22,   # V  — numeric
    "non_qualified":  23,   # W  — numeric
    "qualified_fml":  24,   # X  — formula =VN-WN
    "payment_method": 25,   # Y
    "proof_of_pay":   26,   # Z
    "payment_entity": 27,   # AA — was missing in original map; shifts everything below
    "received_inv":   28,   # AB
    "loan_out":       29,   # AC
    "loan_out_name":  30,   # AD
    "sales_tax":      31,   # AE
    "active_tax_acct":32,   # AF
    "withholding":    33,   # AG
    "w9":             34,   # AH
    "biz_license":    35,   # AI
    "aicp_code":      36,   # AJ
    # 37 = AICP Category Description (AK) — formula, frontend writes it
    "qualified":      38,   # AL
    "notes":          39,   # AM
    "website":        40,   # AN
    # AO–AR (41–44) = TPC review columns — always blank
}


def _write_ga_ap_rows(ws, rows: list[dict]) -> int:
    """Write extracted GA AP rows into an openpyxl worksheet starting at row 4.

    Keeps rows 1-3 (metadata, col notes, headers) untouched.  Overwrites data
    cells from row 4 onward and blanks any leftover cells below the new data.
    Returns the number of rows written.
    """
    DATA_START = 4   # row 1=metadata, 2=col notes, 3=headers, 4=first data row
    prev_max   = ws.max_row

    def _num(v):
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    def _date(v):
        if not v:
            return None
        if isinstance(v, (_dt,)):
            return v
        s = str(v).strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return _dt.strptime(s, fmt)
            except ValueError:
                continue
        return s  # leave as string if unparseable

    def _txt(v):
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    for i, row in enumerate(rows):
        r = DATA_START + i
        c = ws.cell  # shorthand

        # Blank all columns in this row first so template remnants don't bleed
        # through in skipped columns (FF1, FF2, AICP, Qualified, Ces Review cols).
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None

        c(row=r, column=_GA_AP_COL["seq"]).value            = i + 1
        c(row=r, column=_GA_AP_COL["check_number"]).value   = _txt(row.get("check_number"))
        c(row=r, column=_GA_AP_COL["invoice_number"]).value = _txt(row.get("invoice_number"))
        c(row=r, column=_GA_AP_COL["invoice_date"]).value   = _date(row.get("invoice_date"))
        c(row=r, column=_GA_AP_COL["po_number"]).value      = _txt(row.get("po_number"))
        c(row=r, column=_GA_AP_COL["vendor_name"]).value    = _txt(row.get("vendor_name"))
        c(row=r, column=_GA_AP_COL["ff1"]).value             = _txt(row.get("ff1"))
        c(row=r, column=_GA_AP_COL["ff2"]).value             = _txt(row.get("ff2"))
        c(row=r, column=_GA_AP_COL["source_code"]).value    = "AP"
        c(row=r, column=_GA_AP_COL["je_number"]).value      = _txt(row.get("je_number"))
        c(row=r, column=_GA_AP_COL["description"]).value    = _txt(row.get("distribution_description"))
        c(row=r, column=_GA_AP_COL["last_name"]).value      = _txt(row.get("last_name"))
        c(row=r, column=_GA_AP_COL["first_name"]).value     = _txt(row.get("first_name"))
        c(row=r, column=_GA_AP_COL["home_state"]).value     = _txt(row.get("home_state"))
        c(row=r, column=_GA_AP_COL["episode"]).value        = _txt(row.get("episode"))
        c(row=r, column=_GA_AP_COL["nights"]).value         = row.get("nights")
        c(row=r, column=_GA_AP_COL["address"]).value        = _txt(row.get("address"))
        c(row=r, column=_GA_AP_COL["city"]).value           = _txt(row.get("city"))
        c(row=r, column=_GA_AP_COL["state"]).value          = _txt(row.get("state"))
        c(row=r, column=_GA_AP_COL["zip"]).value            = _txt(row.get("zip"))
        c(row=r, column=_GA_AP_COL["org_cur"]).value        = "US"
        c(row=r, column=_GA_AP_COL["amount"]).value         = _num(row.get("amount"))
        c(row=r, column=_GA_AP_COL["non_qualified"]).value  = _num(row.get("non_qualified", 0))
        c(row=r, column=_GA_AP_COL["qualified_fml"]).value  = f"=V{r}-W{r}"
        c(row=r, column=_GA_AP_COL["payment_method"]).value = _txt(row.get("payment_method"))
        c(row=r, column=_GA_AP_COL["proof_of_pay"]).value   = _txt(row.get("proof_of_payment"))
        c(row=r, column=_GA_AP_COL["payment_entity"]).value = _txt(row.get("payment_entity"))
        c(row=r, column=_GA_AP_COL["received_inv"]).value   = _txt(row.get("received_invoice"))
        c(row=r, column=_GA_AP_COL["loan_out"]).value       = _txt(row.get("loan_out"))
        c(row=r, column=_GA_AP_COL["loan_out_name"]).value  = _txt(row.get("loan_out_individual_name"))
        c(row=r, column=_GA_AP_COL["sales_tax"]).value      = _txt(row.get("sales_tax_on_invoice"))
        c(row=r, column=_GA_AP_COL["active_tax_acct"]).value = _txt(row.get("active_sales_tax_account"))
        c(row=r, column=_GA_AP_COL["withholding"]).value    = _txt(row.get("withholding"))
        c(row=r, column=_GA_AP_COL["w9"]).value             = _txt(row.get("w9"))
        c(row=r, column=_GA_AP_COL["biz_license"]).value    = _txt(row.get("business_license"))
        c(row=r, column=_GA_AP_COL["aicp_code"]).value      = row.get("aicp_code")
        c(row=r, column=_GA_AP_COL["qualified"]).value      = _txt(row.get("qualified"))
        c(row=r, column=_GA_AP_COL["notes"]).value          = _txt(row.get("notes"))
        c(row=r, column=_GA_AP_COL["website"]).value        = _txt(row.get("website_address"))

    # Blank any leftover rows from the original template data.
    # Skip merged-cell slaves — openpyxl raises AttributeError if you try to
    # assign .value on a MergedCell (the non-top-left cells of a merged region).
    last_new = DATA_START + len(rows) - 1
    for r in range(last_new + 1, prev_max + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None

    return len(rows)


@app.post("/build-ga-workbook")
async def build_ga_workbook(
    files:          list[UploadFile] = File(...),
    template:       UploadFile       = File(...),
    prodco_name:    str              = Form(""),
    prodco_address: str              = Form(""),
    agency_name:    str              = Form(""),
    work_state:     str              = Form("GA"),
    payer_entities: str              = Form("[]"),
    project_title:  str              = Form(""),
    x_app_secret:   str              = Header(default=""),
):
    """Extract GA AP rows from uploaded PDFs and write them into the uploaded
    GA template XLSX.  Returns a populated .xlsx file for download.

    Form fields mirror /extract-ga-ap; additionally requires:
      template      — the blank (or prior) GA state submission .xlsx
      project_title — used to name the download file (optional)
    """
    if APP_SHARED_SECRET and x_app_secret != APP_SHARED_SECRET:
        raise HTTPException(401, "Bad or missing X-App-Secret header.")

    try:
        entities = json.loads(payer_entities)
        if not isinstance(entities, list):
            entities = []
    except Exception:
        entities = []
    if not entities:
        if prodco_name.strip():
            entities.append({"role": "prodco", "name": prodco_name.strip(), "address": prodco_address.strip()})
        if agency_name.strip():
            entities.append({"role": "agency", "name": agency_name.strip(), "address": ""})

    # ── Extract rows from PDFs ────────────────────────────────────────────────
    pdf_files     = sorted(files, key=lambda f: (f.filename or "").lower())
    client        = _client()
    system_prompt = _load_ga_ap_prompt(entities, work_state)
    user_text     = "Extract invoice data from these document pages."

    rows: list[dict] = []
    issues: list[str] = []

    for uf in pdf_files:
        data = await uf.read()
        try:
            raw_list = _extract_from_file(uf.filename, data, system_prompt, client, user_text=user_text)
        except Exception as e:
            issues.append(f"{uf.filename}: {e}")
            raw_list = []
        if not raw_list:
            issues.append(f"{uf.filename}: no data extracted — review manually")
        for raw in raw_list:
            try:
                rows.append(normalize_ga_ap_row(raw))
            except Exception as e:
                issues.append(f"{uf.filename}: row normalization error: {e}")

    # ── Write rows into the template ──────────────────────────────────────────
    tmpl_bytes = await template.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(tmpl_bytes))
    except Exception as e:
        raise HTTPException(400, f"Could not open template: {e}")

    if "AP" not in wb.sheetnames:
        raise HTTPException(400, "Template is missing an 'AP' sheet.")

    n_written = _write_ga_ap_rows(wb["AP"], rows)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    title = (project_title.strip() or "GA Submission").replace("/", "-").replace("\\", "-")
    fname = f"{title} - State Submission.xlsx"

    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "X-Rows-Written":      str(n_written),
            "X-Issues":            str(len(issues)),
        },
    )


# ── Frontend ─────────────────────────────────────────────────────────────────

_FRONTEND_DIR   = os.path.join(os.path.dirname(__file__), "frontend")
_FRONTEND_INDEX = os.path.join(_FRONTEND_DIR, "index.html")

@app.get("/", include_in_schema=False)
async def serve_frontend():
    return FileResponse(_FRONTEND_INDEX)

@app.get("/workbook-engine.js", include_in_schema=False)
async def serve_workbook_engine():
    return FileResponse(
        os.path.join(_FRONTEND_DIR, "workbook-engine.js"),
        media_type="application/javascript",
    )
